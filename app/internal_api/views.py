"""
Internal API for React frontend
Provides CRUD operations for all core models with proper filtering and serialization
"""
from rest_framework import viewsets, filters, status, permissions, serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django_filters import FilterSet, CharFilter, DateFilter
from django.db.models import Q
from metadata.models import Item, Collection, Collaborator, Languoid, ItemTitle
from .serializers import (
    InternalItemSerializer, 
    InternalCollectionSerializer, 
    InternalCollaboratorSerializer, 
    InternalLanguoidSerializer,
    InternalItemTitleSerializer
)


class IsAuthenticatedWithEditAccess(permissions.BasePermission):
    """
    Custom permission for internal API:
    - Users must be in Archivist, Museum Staff, or Read-Only group (or be staff/superuser) to view (GET)
    - Staff users OR Archivist/Museum Staff groups can modify (POST, PUT, PATCH, DELETE)
    
    This allows Museum Staff (is_staff=False) to edit via React app while blocking Django admin access.
    Read-Only group members can view but not edit.
    """
    def has_permission(self, request, view):
        # Must be authenticated
        if not request.user or not request.user.is_authenticated:
            return False
        
        # Check if user has view access (must be in a group or be staff/superuser)
        has_view_access = (
            request.user.is_staff or 
            request.user.is_superuser or
            request.user.groups.filter(name__in=['Archivist', 'Museum Staff', 'Read-Only']).exists()
        )
        
        if not has_view_access:
            return False
            
        # For read operations, view access is sufficient
        if request.method in permissions.SAFE_METHODS:
            return True
            
        # For write operations, need edit access (staff OR Archivist/Museum Staff groups)
        return (request.user.is_staff or 
                request.user.groups.filter(name__in=['Archivist', 'Museum Staff']).exists())


class ItemFilter(FilterSet):
    """Custom filter for Items with all the search fields from Django templates"""
    # Text search filters
    catalog_number = CharFilter(field_name='catalog_number', lookup_expr='exact')  # Exact match for uniqueness validation
    catalog_number_contains = CharFilter(field_name='catalog_number', lookup_expr='icontains')
    item_access_level_contains = CharFilter(field_name='item_access_level', lookup_expr='icontains')
    call_number_contains = CharFilter(field_name='call_number', lookup_expr='icontains')
    indigenous_title_contains = CharFilter(method='filter_indigenous_title')
    english_title_contains = CharFilter(method='filter_english_title')
    titles_contains = CharFilter(method='filter_all_titles')
    resource_type_contains = CharFilter(field_name='resource_type', lookup_expr='icontains')
    language_contains = CharFilter(method='filter_language')
    description_scope_and_content_contains = CharFilter(field_name='description_scope_and_content', lookup_expr='icontains')
    genre_contains = CharFilter(field_name='genre', lookup_expr='icontains')
    collaborator_contains = CharFilter(method='filter_collaborator')
    depositor_name_contains = CharFilter(field_name='depositor_name', lookup_expr='icontains')
    keyword_contains = CharFilter(method='filter_keyword')
    
    # Date range filters
    accession_date_min = DateFilter(field_name='accession_date_min', lookup_expr='gte')
    accession_date_max = DateFilter(field_name='accession_date_max', lookup_expr='lte')
    creation_date_min = DateFilter(field_name='creation_date_min', lookup_expr='gte')
    creation_date_max = DateFilter(field_name='creation_date_max', lookup_expr='lte')

    class Meta:
        model = Item
        fields = [
            'catalog_number',  # Exact match for uniqueness validation
            'catalog_number_contains',
            'item_access_level_contains', 
            'call_number_contains',
            'accession_date_min',
            'accession_date_max',
            'indigenous_title_contains',
            'english_title_contains',
            'titles_contains',
            'resource_type_contains',
            'language_contains',
            'creation_date_min',
            'creation_date_max',
            'description_scope_and_content_contains',
            'genre_contains',
            'collaborator_contains',
            'depositor_name_contains',
            'keyword_contains',
        ]

    def filter_indigenous_title(self, queryset, name, value):
        """Filter by indigenous titles"""
        return queryset.filter(title_item__title__icontains=value, title_item__language__name__icontains='indigenous').distinct()
    
    def filter_english_title(self, queryset, name, value):
        """Filter by English titles"""
        return queryset.filter(title_item__title__icontains=value, title_item__language__name__icontains='english').distinct()
    
    def filter_all_titles(self, queryset, name, value):
        """Filter by any title"""
        return queryset.filter(title_item__title__icontains=value).distinct()
    
    def filter_language(self, queryset, name, value):
        """Filter by language name"""
        return queryset.filter(language__name__icontains=value).distinct()
    
    def filter_collaborator(self, queryset, name, value):
        """Filter by collaborator name"""
        return queryset.filter(
            Q(collaborator__firstname__icontains=value) |
            Q(collaborator__lastname__icontains=value) |
            Q(collaborator__name__icontains=value)
        ).distinct()
    
    def filter_keyword(self, queryset, name, value):
        """Search across multiple text fields"""
        return queryset.filter(
            Q(catalog_number__icontains=value) |
            Q(call_number__icontains=value) |
            Q(description_scope_and_content__icontains=value) |
            Q(title_item__title__icontains=value) |
            Q(collaborator__firstname__icontains=value) |
            Q(collaborator__lastname__icontains=value) |
            Q(collaborator__name__icontains=value)
        ).distinct()


class CollectionFilter(FilterSet):
    """Custom filter for Collections with exact match support for uniqueness validation"""
    # Text search filters
    collection_abbr = CharFilter(field_name='collection_abbr', lookup_expr='exact')  # Exact match for uniqueness validation
    collection_abbr_contains = CharFilter(field_name='collection_abbr', lookup_expr='icontains')
    name_contains = CharFilter(field_name='name', lookup_expr='icontains')
    description_contains = CharFilter(field_name='description', lookup_expr='icontains')
    
    class Meta:
        model = Collection
        fields = [
            'collection_abbr',  # Exact match for uniqueness validation
            'collection_abbr_contains',
            'name_contains',
            'description_contains',
        ]


class InternalItemViewSet(viewsets.ModelViewSet):
    """Internal API for Items - used by React frontend"""
    queryset = Item.objects.all().prefetch_related(
        'title_item__language',
        'collaborator',
        'language',
        'collection'
    )
    serializer_class = InternalItemSerializer
    permission_classes = [IsAuthenticatedWithEditAccess]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter]
    filterset_class = ItemFilter
    search_fields = ['catalog_number', 'call_number', 'description_scope_and_content']
    ordering_fields = ['catalog_number', 'accession_date_min', 'creation_date_min']
    ordering = ['catalog_number']

    def get_queryset(self):
        """All authenticated users can see all items, permissions handled by permission_classes"""
        if not self.request.user.is_authenticated:
            return Item.objects.none()
        return super().get_queryset()


class InternalCollectionViewSet(viewsets.ModelViewSet):
    """Internal API for Collections - used by React frontend"""
    queryset = Collection.objects.all().prefetch_related('languages')
    serializer_class = InternalCollectionSerializer
    permission_classes = [IsAuthenticatedWithEditAccess]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter]
    filterset_class = CollectionFilter
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created']
    ordering = ['name']

    def get_queryset(self):
        """All authenticated users can see all collections, permissions handled by permission_classes"""
        if not self.request.user.is_authenticated:
            return Collection.objects.none()
        return super().get_queryset()


class InternalCollaboratorViewSet(viewsets.ModelViewSet):
    """Internal API for Collaborators - used by React frontend"""
    queryset = Collaborator.objects.all().prefetch_related('native_languages', 'other_languages', 'collaborator_collaboratorroles__item__collection')
    serializer_class = InternalCollaboratorSerializer
    permission_classes = [IsAuthenticatedWithEditAccess]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter]
    filterset_fields = ['anonymous', 'gender', 'collaborator_id']
    search_fields = ['firstname', 'lastname', 'name', 'collaborator_id', 'tribal_affiliations']
    ordering_fields = ['firstname', 'lastname', 'name', 'collaborator_id', 'added', 'updated']
    ordering = ['lastname', 'firstname', 'name', 'collaborator_id']

    def get_queryset(self):
        """All authenticated users can see all collaborators, permissions handled by permission_classes"""
        if not self.request.user.is_authenticated:
            return Collaborator.objects.none()
        return super().get_queryset()
    
    def get_serializer_context(self):
        """Add request to serializer context for privacy logic"""
        context = super().get_serializer_context()
        context['request'] = self.request
        return context


class LanguoidFilter(FilterSet):
    """Custom filter for Languoids with hierarchical level filtering and exact match support"""
    
    # Exact match filters for uniqueness validation
    name = CharFilter(field_name='name', lookup_expr='exact')
    glottocode = CharFilter(field_name='glottocode', lookup_expr='exact')
    iso = CharFilter(field_name='iso', lookup_expr='exact')
    
    class Meta:
        model = Languoid
        fields = {
            'name': ['icontains', 'exact'],
            'iso': ['icontains', 'exact'],
            'glottocode': ['icontains', 'exact'],
            'level_nal': ['exact'],
            'level_glottolog': ['exact'],
            'region': ['icontains'],
            'tribes': ['icontains'],
        }
        # Remove alt_names from fields since it's a JSONField that needs special handling


class InternalLanguoidViewSet(viewsets.ModelViewSet):
    """Internal API for Languoids with hierarchical filtering and sorting
    
    Supports lookup by both ID and glottocode:
    - GET /internal/v1/languoids/123/ (numeric ID)
    - GET /internal/v1/languoids/cher1273/ (glottocode)
    """
    queryset = Languoid.objects.select_related(
        'family_languoid', 'parent_languoid', 'pri_subgroup_languoid', 
        'sec_subgroup_languoid'
    ).prefetch_related('child_languoids')
    serializer_class = InternalLanguoidSerializer
    permission_classes = [IsAuthenticatedWithEditAccess]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter]
    filterset_class = LanguoidFilter
    search_fields = ['name', 'iso', 'glottocode', 'region', 'tribes']
    ordering_fields = ['name', 'iso', 'level_nal', 'added', 'updated']
    ordering = ['name']  # Default ordering by name
    lookup_field = 'pk'  # Can be either ID or glottocode
    
    def get_object(self):
        """
        Override to support lookup by both glottocode and ID
        Tries glottocode first (if it matches format), then falls back to ID
        """
        queryset = self.filter_queryset(self.get_queryset())
        lookup_url_kwarg = self.lookup_url_kwarg or self.lookup_field
        identifier = self.kwargs[lookup_url_kwarg]
        
        # Try glottocode first if it looks like a glottocode (8 chars, last 4 numeric)
        # or if it's clearly not a numeric ID
        if not identifier.isdigit():
            try:
                obj = queryset.get(glottocode=identifier)
                self.check_object_permissions(self.request, obj)
                return obj
            except Languoid.DoesNotExist:
                pass
        
        # Fall back to ID lookup
        try:
            obj = queryset.get(pk=identifier)
            self.check_object_permissions(self.request, obj)
            return obj
        except (Languoid.DoesNotExist, ValueError):
            pass
        
        # If neither worked, raise 404
        from rest_framework.exceptions import NotFound
        raise NotFound(f'Languoid with identifier "{identifier}" not found.')

    def get_queryset(self):
        """Enhanced queryset with hierarchical ordering support"""
        if not self.request.user.is_authenticated:
            return Languoid.objects.none()
        
        queryset = super().get_queryset()
        
        # Check for hierarchical ordering parameter
        hierarchical = self.request.query_params.get('hierarchical', 'false').lower() == 'true'
        
        if hierarchical:
            # Custom hierarchical ordering: families first, then their children
            queryset = self._get_hierarchical_queryset(queryset)
        
        return queryset
    
    def _get_hierarchical_queryset(self, queryset):
        """
        Return queryset ordered hierarchically:
        1. Families (level_nal='family') ordered by name
        2. Languages under their families (level_nal='language') 
        3. Dialects under their languages (level_nal='dialect')
        """
        from django.db.models import Case, When, Value, CharField
        
        # Create ordering that puts families first, then languages, then dialects
        # and orders by family hierarchy using FK relationships
        return queryset.annotate(
            hierarchy_order=Case(
                When(level_nal='family', then=Value('1')),
                When(level_nal='language', then=Value('2')),
                When(level_nal='dialect', then=Value('3')),
                default=Value('4'),
                output_field=CharField()
            )
        ).order_by(
            'hierarchy_order',
            'family_languoid__name',  # Group by family name via FK
            'pri_subgroup_languoid__name',  # Then by primary subgroup via FK
            'sec_subgroup_languoid__name',  # Then by secondary subgroup via FK
            'name'  # Finally by name
        )

    def get_serializer_context(self):
        """Add request context for user tracking in serializer"""
        context = super().get_serializer_context()
        context['request'] = self.request
        return context
    
    def list(self, request, *args, **kwargs):
        """
        Override list to use cached response for the full languoid list.
        
        This method:
        - Checks if request is for full list with hierarchical ordering
        - Returns cached response if available (< 1 second)
        - Falls back to building and caching if cache miss (~22 seconds)
        - Works seamlessly with background cache warming
        """
        from django.core.cache import cache
        from rest_framework.response import Response
        import logging
        
        logger = logging.getLogger(__name__)
        
        # Only cache the full list with hierarchical ordering
        # This matches what the frontend requests: ?page_size=10000&hierarchical=true
        is_full_list = (
            request.query_params.get('page_size') == '10000' and
            request.query_params.get('hierarchical', 'false').lower() == 'true'
            # Note: Don't check exact param count - other params like auth tokens may be present
        )
        
        if not is_full_list:
            # Not the full list request, use normal pagination
            logger.debug("[Cache] Not full list request, using normal pagination")
            return super().list(request, *args, **kwargs)
        
        # Try to get from cache
        cache_key = 'languoid_list_full'
        cached_data = cache.get(cache_key)
        
        if cached_data is not None:
            logger.info(f"[Cache] Cache hit! Returning {len(cached_data)} languoids from cache")
            # Wrap cached data in pagination format that frontend expects
            return Response({
                'count': len(cached_data),
                'next': None,
                'previous': None,
                'results': cached_data
            })
        
        # Cache miss - build and cache the response
        logger.warning("[Cache] Cache miss! Building languoid list (this will be slow)...")
        
        # Import the utility function from tasks
        from metadata.tasks import build_languoid_list_cache
        
        # Build the cache data
        data = build_languoid_list_cache()
        
        # Store in cache with 10-minute TTL
        cache.set(cache_key, data, timeout=600)
        
        logger.info(f"[Cache] Built and cached {len(data)} languoids")
        
        # Wrap in pagination format
        return Response({
            'count': len(data),
            'next': None,
            'previous': None,
            'results': data
        })
    
    @action(detail=False, methods=['post'], url_path='validate-field')
    def validate_field(self, request):
        """
        Validate a single field value for batch editing
        
        POST /internal/v1/languoids/validate-field/
        Body: {
            "field_name": "glottocode",
            "value": "stan1295",
            "row_id": "draft-123" or 42 (languoid id),
            "original_value": "stan1295" (optional - value when loaded from DB)
        }
        
        Returns: {
            "valid": true/false,
            "error": "error message if invalid"
        }
        """
        field_name = request.data.get('field_name')
        value = request.data.get('value')
        row_id = request.data.get('row_id')
        original_value = request.data.get('original_value')  # New parameter
        
        if not field_name:
            return Response(
                {'valid': False, 'error': 'field_name is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get the serializer to use its validation methods
        serializer = self.get_serializer()
        
        # Try to validate the field using serializer's field validators
        try:
            # Get the field from the serializer
            if field_name not in serializer.fields:
                return Response(
                    {'valid': False, 'error': f'Unknown field: {field_name}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            field = serializer.fields[field_name]
            
            # Run field-level validation
            field.run_validation(value)
            
            # Check for field-specific validator methods (e.g., validate_glottocode)
            validate_method_name = f'validate_{field_name}'
            if hasattr(serializer, validate_method_name):
                validate_method = getattr(serializer, validate_method_name)
                validate_method(value)
            
            # Field-specific uniqueness checks
            if field_name == 'glottocode' and value:
                # If value equals original_value, it's valid (editing back to self)
                if original_value is not None and value == original_value:
                    return Response({'valid': True})
                
                # Check for uniqueness (exclude current row if editing)
                existing = Languoid.objects.filter(glottocode=value)
                
                # Exclude the current row being edited
                if str(row_id).startswith('draft-'):
                    # Draft row - don't exclude anything (it's a new row)
                    pass
                else:
                    # Existing row - exclude it from uniqueness check
                    try:
                        row_id_int = int(row_id) if isinstance(row_id, str) else row_id
                        existing = existing.exclude(id=row_id_int)
                    except (ValueError, TypeError):
                        # If we can't convert to int, skip exclusion
                        pass
                
                if existing.exists():
                    return Response({
                        'valid': False,
                        'error': f'Glottocode "{value}" already exists.'
                    })
            
            return Response({'valid': True})
            
        except serializers.ValidationError as e:
            # Extract error message
            if isinstance(e.detail, dict):
                error_msg = str(list(e.detail.values())[0][0]) if e.detail else 'Validation error'
            elif isinstance(e.detail, list):
                error_msg = str(e.detail[0])
            else:
                error_msg = str(e.detail)
            
            return Response({
                'valid': False,
                'error': error_msg
            })
        except Exception as e:
            return Response({
                'valid': False,
                'error': str(e)
            })
    
    @action(detail=False, methods=['post'], url_path='save-batch')
    def save_batch(self, request):
        """
        Save multiple languoids in a single transaction with batch-optimized hierarchy updates.
        
        POST /internal/v1/languoids/save-batch/
        Body: {
            "rows": [
                {
                    "id": 42 or "draft-uuid",
                    "name": "English",
                    "glottocode": "stan1293",
                    "level_glottolog": "language",
                    ...
                },
                ...
            ]
        }
        
        Returns: {
            "success": true,
            "saved": [{"id": 42, ...}, ...],
            "errors": []
        }
        
        OPTIMIZATION:
        - Layer 1 (Batch-Aware Signals): Skips individual task scheduling via _skip_async_tasks flag
        - Layer 2 (Hierarchy-Change Detection): Only processes hierarchy-changed languoids
        """
        from django.db import transaction
        from django.core.cache import cache
        import logging
        logger = logging.getLogger(__name__)
        
        rows = request.data.get('rows', [])
        logger.info(f"[BATCH SAVE] Received {len(rows)} rows")
        
        if not rows:
            return Response(
                {'success': False, 'errors': ['No rows provided']},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Define hierarchy fields for Layer 2 detection
        HIERARCHY_FIELDS = {'parent_languoid', 'level_glottolog'}
        
        saved_objects = []
        errors = []
        
        # Track what needs post-batch processing
        hierarchy_changed_languoids = []  # Languoids with hierarchy changes
        old_parent_ids = set()  # Old parents to update
        needs_orphaning = []  # Languoids that changed from language level
        
        try:
            with transaction.atomic():
                for row in rows:
                    row_id = row.get('id')
                    is_draft = isinstance(row_id, str) and str(row_id).startswith('draft-')
                    
                    try:
                        if is_draft:
                            # CREATE NEW LANGUOID
                            logger.debug(f"[BATCH SAVE] Creating new languoid from draft: {row_id}")
                            row_data = {k: v for k, v in row.items() if k not in ['id', '_updated', '_original_values']}
                            serializer = self.get_serializer(data=row_data)
                            serializer.is_valid(raise_exception=True)
                            
                            # Set batch flag to skip individual signals
                            instance = serializer.save()
                            instance._skip_async_tasks = True
                            instance.save()
                            
                            saved_objects.append(instance)
                            # New languoids always need hierarchy processing
                            hierarchy_changed_languoids.append(instance)
                            
                        else:
                            # UPDATE EXISTING LANGUOID
                            logger.debug(f"[BATCH SAVE] Updating existing languoid: {row_id}")
                            try:
                                old_instance = Languoid.objects.get(pk=row_id)
                            except Languoid.DoesNotExist:
                                errors.append(f"Languoid with ID {row_id} not found")
                                continue
                            
                            # CONFLICT DETECTION: Field-level conflict checking
                            # Check if row was modified since UI loaded it
                            client_updated = row.get('_updated')  # Frontend sends DB's 'updated' timestamp
                            conflicting_fields = []
                            
                            if client_updated:
                                # Parse client timestamp (ISO format)
                                from django.utils.dateparse import parse_datetime
                                client_updated_dt = parse_datetime(client_updated)
                                
                                if client_updated_dt:
                                    # Compare exact timestamps (with microsecond precision)
                                    db_updated = old_instance.updated
                                    
                                    logger.info(f"[BATCH SAVE] Conflict check for {row_id}: DB={db_updated}, Client={client_updated_dt}, DB > Client? {db_updated > client_updated_dt}")
                                    
                                    if db_updated > client_updated_dt:
                                        # Row was modified - check which fields have TRUE conflicts
                                        # True conflict = client edited field AND another user also changed it
                                        
                                        # Get client's original values (what they loaded)
                                        client_original_values = row.get('_original_values', {})
                                        logger.info(f"[BATCH SAVE] Checking conflicts for row {row_id}. Client original values: {client_original_values}")
                                        
                                        for field_name in row.keys():
                                            if field_name in ['id', '_updated', '_original_values']:
                                                continue  # Skip metadata fields
                                            
                                            # Get current DB value
                                            db_value = getattr(old_instance, field_name, None)
                                            # Handle FK fields
                                            if hasattr(db_value, 'id'):
                                                db_value = db_value.id
                                            
                                            # Client is trying to update this field
                                            client_new_value = row[field_name]
                                            
                                            # Get what the client originally loaded for this field
                                            client_original_value = client_original_values.get(field_name)
                                            
                                            logger.info(f"[BATCH SAVE] Field '{field_name}': db_value={repr(db_value)}, client_original={repr(client_original_value)}, client_new={repr(client_new_value)}, has_conflict={db_value != client_original_value if client_original_value is not None else 'N/A'}")
                                            
                                            # TRUE CONFLICT: DB value changed from what client loaded
                                            # (Another user edited this field after client loaded it)
                                            if client_original_value is not None and db_value != client_original_value:
                                                logger.warning(f"[BATCH SAVE] TRUE conflict on field '{field_name}': client_original={client_original_value}, db_current={db_value}, client_new={client_new_value}")
                                                conflicting_fields.append({
                                                    'field': field_name,
                                                    'db_value': db_value,
                                                    'client_original': client_original_value,
                                                    'client_new': client_new_value
                                                })
                                        
                                        if conflicting_fields:
                                            logger.warning(f"[BATCH SAVE] Field-level conflicts detected for languoid {row_id}: {[f['field'] for f in conflicting_fields]}")
                                            
                                            # Remove conflicting fields from the update
                                            # (They will not be saved - user must resolve and re-save)
                                            conflicting_field_names = [f['field'] for f in conflicting_fields]
                                            for field_name in conflicting_field_names:
                                                if field_name in row:
                                                    del row[field_name]
                                                    logger.info(f"[BATCH SAVE] Removed conflicting field '{field_name}' from save")
                                            
                                            # Add conflict info to errors
                                            errors.append({
                                                'row_id': row_id,
                                                'type': 'conflict',
                                                'message': f'Languoid {row_id} was modified by another user.',
                                                'conflicting_fields': conflicting_field_names,
                                                'current_data': self.get_serializer(old_instance).data,
                                            })
                                            
                                            # If ALL fields have conflicts, skip the entire row
                                            non_metadata_fields = [k for k in row.keys() if k not in ['id', '_updated', '_original_values']]
                                            if len(non_metadata_fields) == 0:
                                                logger.info(f"[BATCH SAVE] All fields conflicted, skipping row {row_id}")
                                                continue  # Skip this row entirely
                            
                            # LAYER 2: Detect if hierarchy fields changed
                            hierarchy_changed = False
                            for field in HIERARCHY_FIELDS:
                                if field in row:
                                    old_value = getattr(old_instance, field, None)
                                    # Handle FK fields (parent_languoid)
                                    if field == 'parent_languoid':
                                        old_value = old_value.id if old_value else None
                                    new_value = row[field]
                                    if old_value != new_value:
                                        hierarchy_changed = True
                                        # Track old parent for ancestor chain update
                                        if field == 'parent_languoid' and old_value:
                                            old_parent_ids.add(old_value)
                                        break
                            
                            # Check for level change FROM language
                            if old_instance.level_glottolog == 'language' and row.get('level_glottolog') != 'language':
                                needs_orphaning.append(row_id)
                            
                            # Set batch flag before saving
                            old_instance._skip_async_tasks = True
                            
                            # Perform update
                            # Remove metadata fields that aren't part of the model
                            row_data_for_serializer = {k: v for k, v in row.items() if k not in ['_updated', '_original_values', 'id']}
                            serializer = self.get_serializer(old_instance, data=row_data_for_serializer, partial=True)
                            serializer.is_valid(raise_exception=True)
                            instance = serializer.save()
                            saved_objects.append(instance)
                            
                            # Track if needs hierarchy processing
                            if hierarchy_changed:
                                hierarchy_changed_languoids.append(instance)
                    
                    except serializers.ValidationError as e:
                        logger.error(f"[BATCH SAVE] Validation error for row {row_id}: {e.detail}")
                        error_msg = f"Row ID {row_id}: "
                        if isinstance(e.detail, dict):
                            error_msg += ", ".join([f"{k}: {v[0]}" if isinstance(v, list) else f"{k}: {v}" 
                                                   for k, v in e.detail.items()])
                        else:
                            error_msg += str(e.detail)
                        errors.append(error_msg)
                
                # If any VALIDATION errors occurred (not conflicts), rollback transaction
                # Conflicts are non-fatal and should not trigger rollback
                validation_errors = [e for e in errors if not isinstance(e, dict) or e.get('type') != 'conflict']
                if validation_errors:
                    raise Exception("Validation errors occurred")
            
            # TRANSACTION COMMITTED - Now do batch post-processing
            logger.info(
                f"[BATCH SAVE] Transaction committed. "
                f"Hierarchy-changed: {len(hierarchy_changed_languoids)}, "
                f"Needs orphaning: {len(needs_orphaning)}, "
                f"Old parents to update: {len(old_parent_ids)}"
            )
            
            # LAYER 2: Only process hierarchy-changed languoids
            if hierarchy_changed_languoids or old_parent_ids or needs_orphaning:
                # Import tasks
                from metadata.tasks import (
                    recalculate_ancestor_descendents_task,
                    orphan_dialects_batch_task,
                    get_all_ancestors
                )
                
                # Collect all unique ancestors from hierarchy-changed languoids
                affected_ancestors = set()
                for languoid in hierarchy_changed_languoids:
                    # Get ancestors for this languoid
                    ancestors = get_all_ancestors(languoid)
                    for ancestor in ancestors:
                        affected_ancestors.add(ancestor.id)
                    # Include the languoid itself
                    affected_ancestors.add(languoid.id)
                
                # Add old parent chains
                for old_parent_id in old_parent_ids:
                    try:
                        old_parent = Languoid.objects.get(id=old_parent_id)
                        old_ancestors = get_all_ancestors(old_parent)
                        for ancestor in old_ancestors:
                            affected_ancestors.add(ancestor.id)
                        affected_ancestors.add(old_parent.id)
                    except Languoid.DoesNotExist:
                        logger.warning(f"[BATCH SAVE] Old parent {old_parent_id} not found")
                
                # Schedule batch-optimized tasks
                if affected_ancestors:
                    logger.info(f"[BATCH SAVE] Scheduling descendents recalculation for {len(affected_ancestors)} ancestors")
                    recalculate_ancestor_descendents_task.apply_async(
                        args=[list(affected_ancestors)],
                        priority=9  # High priority
                    )
                
                if needs_orphaning:
                    logger.info(f"[BATCH SAVE] Scheduling dialect orphaning for {len(needs_orphaning)} languoids")
                    orphan_dialects_batch_task.apply_async(
                        args=[needs_orphaning],
                        priority=9  # High priority
                    )
            
            # Invalidate cache ONCE after all processing
            from metadata.tasks import invalidate_and_warm_languoid_cache
            logger.info("[BATCH SAVE] Invalidating cache")
            cache.delete('languoid_list_full')
            invalidate_and_warm_languoid_cache.apply_async(priority=8)
            
            # Return response (include errors list for conflicts)
            response_serializer = self.get_serializer(saved_objects, many=True)
            return Response({
                'success': True,
                'saved': response_serializer.data,
                'errors': errors  # Include conflicts and other non-fatal errors
            })
        
        except Exception as e:
            logger.error(f"[BATCH SAVE] Error: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({
                'success': False,
                'saved': [],
                'errors': errors if errors else [str(e)]
            }, status=status.HTTP_400_BAD_REQUEST)


    
    @action(detail=True, methods=['get'], url_path='descendants-tree')
    def descendants_tree(self, request, pk=None):
        """
        Get hierarchical tree of all descendants for a languoid
        
        GET /internal/v1/languoids/{id}/descendants-tree/
        
        Returns: [
            {
                "id": 1,
                "name": "Malayo-Polynesian",
                "glottocode": "mala1545",
                "level_nal": "subfamily",
                "level_display": "Primary Subfamily",
                "children": [...]
            }
        ]
        """
        languoid = self.get_object()
        
        def build_tree(parent_id):
            """Recursively build tree structure from parent-child relationships"""
            children = Languoid.objects.filter(
                parent_languoid_id=parent_id
            ).select_related(
                'parent_languoid'
            ).order_by('name')
            
            tree = []
            for child in children:
                node = {
                    'id': child.id,
                    'name': child.name,
                    'glottocode': child.glottocode,
                    'level_nal': child.level_nal,
                    'level_display': child.get_level_nal_display(),
                    'children': build_tree(child.id)  # Recursive call
                }
                tree.append(node)
            
            return tree
        
        descendants = build_tree(languoid.id)
        
        return Response(descendants)
    
    @action(detail=False, methods=['get'], url_path='last-modified')
    def last_modified(self, request):
        """
        Get the timestamp of the most recently modified languoid.
        Used for cache invalidation on the frontend.
        
        GET /internal/v1/languoids/last-modified/
        
        Returns: {
            "last_modified": "2025-10-28T12:34:56.789Z",
            "count": 1234
        }
        """
        from django.db.models import Max, Count
        
        # Get the most recent update timestamp
        result = Languoid.objects.aggregate(
            last_modified=Max('updated'),
            count=Count('id')
        )
        
        last_modified = result.get('last_modified')
        count = result.get('count', 0)
        
        # If no languoids exist, return current time
        if last_modified is None:
            from django.utils import timezone
            last_modified = timezone.now()
        
        return Response({
            'last_modified': last_modified.isoformat(),
            'count': count
        })
    
    @action(detail=False, methods=['post'], url_path='export')
    def export_languoids(self, request):
        """
        Export languoids to Excel spreadsheet.
        
        POST /internal/v1/languoids/export/
        Body: {
            "mode": "filtered" | "selected",
            "ids": [1, 2, 3, ...]
        }
        
        Returns:
        - For <= 100 languoids: Excel file download (synchronous)
        - For > 100 languoids: Job info for async export (background task)
        """
        import logging
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from datetime import datetime
        import uuid
        
        logger = logging.getLogger(__name__)
        
        mode = request.data.get('mode', 'filtered')
        ids = request.data.get('ids', [])
        
        logger.info(f"[EXPORT] Exporting {len(ids)} languoids (mode: {mode})")
        
        if not ids:
            return Response(
                {'detail': 'No languoids to export'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if we should use async export (> 100 languoids)
        if len(ids) > 100:
            logger.info(f"[EXPORT] Large export ({len(ids)} languoids) - using async task")
            
            # Generate unique export ID
            export_id = str(uuid.uuid4())
            
            # Trigger async task
            from metadata.tasks import generate_languoid_export_task
            task = generate_languoid_export_task.apply_async(
                args=[export_id, mode, ids],
                priority=7  # High priority but not urgent
            )
            
            return Response({
                'async': True,
                'export_id': export_id,
                'task_id': task.id,
                'count': len(ids),
                'message': 'Export is being generated in the background. Check status using the export_id.'
            })
        
        # SYNCHRONOUS EXPORT (≤ 100 languoids)
        try:
            queryset = Languoid.objects.filter(id__in=ids).select_related(
                'parent_languoid',
                'family_languoid',
                'pri_subgroup_languoid',
                'sec_subgroup_languoid'
            )
            
            languoids = list(queryset)
            logger.info(f"[EXPORT] Found {len(languoids)} languoids to export")
            
            # Sort languoids by hierarchy (tree structure)
            # Build a tree structure for sorting
            languoid_dict = {l.id: l for l in languoids}
            
            def get_sort_key(languoid):
                """
                Generate a sort key that preserves tree hierarchy.
                Returns a tuple of (ancestor_names..., own_name) for hierarchical sorting.
                """
                path = []
                current = languoid
                visited = set()
                
                # Walk up the tree to build the full path
                while current:
                    if current.id in visited:
                        # Circular reference protection
                        break
                    visited.add(current.id)
                    path.insert(0, current.name.lower() if current.name else '')
                    
                    # Move to parent
                    if current.parent_languoid and current.parent_languoid.id in languoid_dict:
                        current = languoid_dict[current.parent_languoid.id]
                    elif current.parent_languoid:
                        # Parent exists but not in our export set - use its name
                        path.insert(0, current.parent_languoid.name.lower() if current.parent_languoid.name else '')
                        break
                    else:
                        # No parent - this is a top-level node
                        break
                
                return tuple(path)
            
            # Sort languoids by hierarchy
            languoids.sort(key=get_sort_key)
            logger.info(f"[EXPORT] Sorted {len(languoids)} languoids by hierarchy")
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Languoids"
            
            # Define column headers (all fields from Languoid model)
            headers = [
                'Name',
                'Name Abbreviation',
                'Glottocode',
                'ISO 639-3',
                'Level (Glottolog)',
                'Level (NAL)',
                'Parent Languoid',
                'Parent Languoid Abbreviation',
                'Parent Languoid Glottocode',
                'Family',
                'Family Abbreviation',
                'Family Glottocode',
                'Primary Subfamily',
                'Primary Subfamily Abbreviation',
                'Primary Subfamily Glottocode',
                'Secondary Subfamily',
                'Secondary Subfamily Abbreviation',
                'Secondary Subfamily Glottocode',
                'Alternate Names',
                'Region',
                'Longitude',
                'Latitude',
                'Tribes',
                'Notes',
            ]
            
            # Write header row with styling
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
            
            # Helper function to safely convert values to Excel-compatible format
            def safe_value(value):
                """Convert value to Excel-compatible format"""
                if value is None:
                    return ''
                if isinstance(value, (list, tuple)):
                    # Convert lists/tuples to comma-separated string
                    return ', '.join(str(v) for v in value)
                if isinstance(value, bool):
                    return 'Yes' if value else 'No'
                return str(value)
            
            # Write data rows
            for row_num, languoid in enumerate(languoids, 2):
                ws.cell(row=row_num, column=1).value = safe_value(languoid.name)
                ws.cell(row=row_num, column=2).value = safe_value(languoid.name_abbrev)
                ws.cell(row=row_num, column=3).value = safe_value(languoid.glottocode)
                ws.cell(row=row_num, column=4).value = safe_value(languoid.iso)
                ws.cell(row=row_num, column=5).value = languoid.get_level_glottolog_display() if languoid.level_glottolog else ''
                ws.cell(row=row_num, column=6).value = languoid.get_level_nal_display() if languoid.level_nal else ''
                
                # Parent Languoid + denormalized fields (Name, Abbrev, Glottocode)
                ws.cell(row=row_num, column=7).value = languoid.parent_languoid.name if languoid.parent_languoid else ''
                ws.cell(row=row_num, column=8).value = safe_value(languoid.parent_languoid.name_abbrev) if languoid.parent_languoid else ''
                ws.cell(row=row_num, column=9).value = safe_value(languoid.parent_languoid.glottocode) if languoid.parent_languoid else ''
                
                # Family + denormalized fields (Name, Abbrev, Glottocode)
                ws.cell(row=row_num, column=10).value = languoid.family_languoid.name if languoid.family_languoid else ''
                ws.cell(row=row_num, column=11).value = safe_value(languoid.family_languoid.name_abbrev) if languoid.family_languoid else ''
                ws.cell(row=row_num, column=12).value = safe_value(languoid.family_languoid.glottocode) if languoid.family_languoid else ''
                
                # Primary Subfamily + denormalized fields (Name, Abbrev, Glottocode)
                ws.cell(row=row_num, column=13).value = languoid.pri_subgroup_languoid.name if languoid.pri_subgroup_languoid else ''
                ws.cell(row=row_num, column=14).value = safe_value(languoid.pri_subgroup_languoid.name_abbrev) if languoid.pri_subgroup_languoid else ''
                ws.cell(row=row_num, column=15).value = safe_value(languoid.pri_subgroup_languoid.glottocode) if languoid.pri_subgroup_languoid else ''
                
                # Secondary Subfamily + denormalized fields (Name, Abbrev, Glottocode)
                ws.cell(row=row_num, column=16).value = languoid.sec_subgroup_languoid.name if languoid.sec_subgroup_languoid else ''
                ws.cell(row=row_num, column=17).value = safe_value(languoid.sec_subgroup_languoid.name_abbrev) if languoid.sec_subgroup_languoid else ''
                ws.cell(row=row_num, column=18).value = safe_value(languoid.sec_subgroup_languoid.glottocode) if languoid.sec_subgroup_languoid else ''
                
                # Other fields
                ws.cell(row=row_num, column=19).value = safe_value(languoid.alt_names)
                ws.cell(row=row_num, column=20).value = safe_value(languoid.region)
                ws.cell(row=row_num, column=21).value = str(languoid.longitude) if languoid.longitude else ''
                ws.cell(row=row_num, column=22).value = str(languoid.latitude) if languoid.latitude else ''
                ws.cell(row=row_num, column=23).value = safe_value(languoid.tribes)
                ws.cell(row=row_num, column=24).value = safe_value(languoid.notes)
            
            # Auto-size columns (approximate)
            for col_num in range(1, len(headers) + 1):
                ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = 20
            
            # Save workbook to response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            filename = f'languoids_export_{timestamp}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # Write workbook to response
            wb.save(response)
            
            logger.info(f"[EXPORT] Successfully generated export file: {filename}")
            return response
            
        except Exception as e:
            logger.error(f"[EXPORT] Error generating export: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response(
                {'detail': f'Export failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='export-status/(?P<export_id>[^/.]+)')
    def export_status(self, request, export_id=None):
        """
        Check the status of an async export.
        
        GET /internal/v1/languoids/export-status/{export_id}/
        
        Returns: {
            "status": "pending" | "processing" | "completed" | "failed",
            "filename": "languoids_export_xxx.xlsx" (if completed),
            "error": "error message" (if failed)
        }
        """
        import logging
        import os
        from django.conf import settings
        from celery.result import AsyncResult
        
        logger = logging.getLogger(__name__)
        
        logger.info(f"[EXPORT STATUS] Checking status for export {export_id}")
        
        # Check if file exists (completed)
        exports_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        filename = f'languoids_export_{export_id}.xlsx'
        file_path = os.path.join(exports_dir, filename)
        
        if os.path.exists(file_path):
            logger.info(f"[EXPORT STATUS] Export {export_id} completed")
            return Response({
                'status': 'completed',
                'filename': filename,
                'export_id': export_id
            })
        
        # File doesn't exist yet - check task status
        # Note: We'd need to store task_id to check actual Celery status
        # For now, return processing
        logger.info(f"[EXPORT STATUS] Export {export_id} still processing")
        return Response({
            'status': 'processing',
            'export_id': export_id
        })
    
    @action(detail=False, methods=['get'], url_path='export-download/(?P<export_id>[^/.]+)')
    def export_download(self, request, export_id=None):
        """
        Download a completed async export.
        
        GET /internal/v1/languoids/export-download/{export_id}/
        
        Returns: Excel file download
        """
        import logging
        import os
        from django.conf import settings
        from django.http import HttpResponse, Http404
        
        logger = logging.getLogger(__name__)
        
        logger.info(f"[EXPORT DOWNLOAD] Downloading export {export_id}")
        
        # Check if file exists
        exports_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        filename = f'languoids_export_{export_id}.xlsx'
        file_path = os.path.join(exports_dir, filename)
        
        if not os.path.exists(file_path):
            logger.error(f"[EXPORT DOWNLOAD] Export {export_id} not found")
            raise Http404("Export file not found or has expired")
        
        # Read file and return
        with open(file_path, 'rb') as f:
            response = HttpResponse(
                f.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            logger.info(f"[EXPORT DOWNLOAD] Successfully served export {export_id}")
            return response

class InternalItemTitleViewSet(viewsets.ModelViewSet):
    """Nested API for ItemTitles under Items - used by React frontend"""
    serializer_class = InternalItemTitleSerializer
    permission_classes = [IsAuthenticatedWithEditAccess]
    
    def get_queryset(self):
        """Filter titles by the parent item"""
        if not self.request.user.is_authenticated:
            return ItemTitle.objects.none()
            
        item_pk = self.kwargs.get('item_pk')
        if item_pk:
            return ItemTitle.objects.filter(item_id=item_pk).select_related('language', 'item')
        return ItemTitle.objects.none()
    
    def get_serializer_context(self):
        """Add item to serializer context for validation"""
        context = super().get_serializer_context()
        item_pk = self.kwargs.get('item_pk')
        if item_pk:
            try:
                context['item'] = Item.objects.get(pk=item_pk)
            except Item.DoesNotExist:
                pass
        return context
    
    def perform_create(self, serializer):
        """Handle default title business rule on create"""
        # If this is being set as default, unset other defaults for this item
        if serializer.validated_data.get('default', False):
            item_pk = self.kwargs.get('item_pk')
            ItemTitle.objects.filter(item_id=item_pk, default=True).update(default=False)
        
        serializer.save()
    
    def perform_update(self, serializer):
        """Handle default title business rule on update"""
        # If this is being set as default, unset other defaults for this item
        if serializer.validated_data.get('default', False):
            item_pk = self.kwargs.get('item_pk')
            ItemTitle.objects.filter(
                item_id=item_pk, 
                default=True
            ).exclude(
                id=serializer.instance.id
            ).update(default=False)
        
        serializer.save()
    
    def perform_destroy(self, instance):
        """Handle title deletion with business rule validation"""
        item_pk = self.kwargs.get('item_pk')
        
        # Check if this is the last title - prevent deletion
        title_count = ItemTitle.objects.filter(item_id=item_pk).count()
        if title_count <= 1:
            from rest_framework.exceptions import ValidationError
            raise ValidationError("Cannot delete the only title. Add another title first.")
        
        # If deleting the default title, set another one as default
        if instance.default:
            # Find another title to make default
            new_default = ItemTitle.objects.filter(
                item_id=item_pk
            ).exclude(id=instance.id).first()
            
            if new_default:
                new_default.default = True
                new_default.save()
        
        instance.delete()
    
    @action(detail=True, methods=['post'])
    def set_default(self, request, item_pk=None, pk=None):
        """Set a specific title as the default for the item"""
        try:
            title = self.get_object()
            
            # Unset other defaults for this item
            ItemTitle.objects.filter(
                item_id=item_pk, 
                default=True
            ).exclude(id=title.id).update(default=False)
            
            # Set this title as default
            title.default = True
            title.save()
            
            serializer = self.get_serializer(title)
            return Response(serializer.data)
            
        except ItemTitle.DoesNotExist:
            return Response(
                {'error': 'Title not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
