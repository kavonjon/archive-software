from openpyxl import Workbook, load_workbook
from django.db.models import Count, Sum, Max, Q
from metadata.models import Item, Language, Dialect, DialectInstance, Collaborator, CollaboratorRole, Geographic, Columns_export, Document, Video, ACCESS_CHOICES, ACCESSION_CHOICES, AVAILABILITY_CHOICES, CONDITION_CHOICES, RESOURCE_TYPE_CHOICES, FORMAT_CHOICES, GENRE_CHOICES, MONTH_CHOICES, ROLE_CHOICES, reverse_lookup_choices, validate_date_text
from django.core.management.base import BaseCommand

class Command(BaseCommand):
    def handle(self, **options):
        # now do the things that you want with your models here

        collaborator_list = Collaborator.objects.all()

        new_workbook = Workbook()
        sheet = new_workbook.active

        sheet_column_counter = 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )
        header_cell.value = 'Collaborator Name'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )
        header_cell.value = 'Nickname'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )
        header_cell.value = 'Other Names'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )
        header_cell.value = 'Anonymous?'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )
        header_cell.value = 'Native Languages'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )
        header_cell.value = 'Other Languages'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )
        header_cell.value = 'Clan Society'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )
        header_cell.value = 'Tribal Affiliation'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )
        header_cell.value = 'Place of Origin'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )
        header_cell.value = 'Date of Birth'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )
        header_cell.value = 'Date of Death'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )
        header_cell.value = 'Gender'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )
        header_cell.value = 'Other info'


        for collaborator in collaborator_list:


            for native_language in collaborator.native_languages.all():
                native_dialect_i, created = DialectInstance.objects.get_or_create(collaborator_native=collaborator, language=native_language)
            native_dialect_info = zip(collaborator.native_languages.all().order_by('name'), DialectInstance.objects.filter(collaborator_native=collaborator).order_by('language__name'))

            for other_language in collaborator.other_languages.all():
                other_dialect_i, created = DialectInstance.objects.get_or_create(collaborator_other=collaborator, language=other_language)
            other_dialect_info = zip(collaborator.other_languages.all().order_by('name'), DialectInstance.objects.filter(collaborator_other=collaborator).order_by('language__name'))





            xl_row = []
            xl_row.append(collaborator.name)
            xl_row.append(collaborator.nickname)
            xl_row.append(collaborator.other_names)
            xl_row.append(collaborator.anonymous)
            xl_row.append(", ".join( collaborator.native_languages.values_list('name', flat=True) ) )
            xl_row.append(", ".join( collaborator.other_languages.values_list('name', flat=True) ) )
            xl_row.append(collaborator.clan_society)
            xl_row.append(collaborator.tribal_affiliations)
            xl_row.append(collaborator.origin)
            xl_row.append(collaborator.birthdate)
            xl_row.append(collaborator.deathdate)
            xl_row.append(collaborator.gender)
            xl_row.append(collaborator.other_info)
            sheet.append(xl_row)


        new_workbook.save("collaborators.xlsx")
