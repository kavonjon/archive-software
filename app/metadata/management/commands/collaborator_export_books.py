from openpyxl import Workbook, load_workbook
from django.db.models import Count, Sum, Max, Q
from metadata.models import Item, Language, Dialect, DialectInstance, Collaborator, CollaboratorRole, Geographic, Columns_export, Document, Video, ACCESS_CHOICES, ACCESSION_CHOICES, AVAILABILITY_CHOICES, CONDITION_CHOICES, CONTENT_CHOICES, FORMAT_CHOICES, GENRE_CHOICES, MONTH_CHOICES, ROLE_CHOICES, reverse_lookup_choices, validate_date_text
from django.core.management.base import BaseCommand

class Command(BaseCommand):
    def handle(self, **options):
        # now do the things that you want with your models here

        collab_role_filter = CollaboratorRole.objects.filter(item__resource_type='book')

        collab_role_all_filters = collab_role_filter.filter(role__icontains='author').union(
                                  collab_role_filter.filter(role__icontains='editor'),
                                  collab_role_filter.filter(role__icontains='speaker'))

        collaborator_list = Collaborator.objects.filter(collaborator_collaboratorroles__in=list(collab_role_all_filters.values_list('id', flat=True))).distinct().order_by('name')

        new_workbook = Workbook()
        sheet = new_workbook.active

        sheet_column_counter = 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )
        header_cell.value = 'Collaborator Name'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )
        header_cell.value = 'Tribal Affiliation'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )



        max_item_counts = 1
        for collaborator in collaborator_list:
            xl_row = []
            xl_row.append(collaborator.name)
            xl_row.append(collaborator.tribal_affiliations)


            current_collaborator_role_filter = CollaboratorRole.objects.filter(collaborator=collaborator).filter(item__resource_type='book')

            current_collaborator_role_all_filters = current_collaborator_role_filter.filter(role__icontains='author').union(
                                                    current_collaborator_role_filter.filter(role__icontains='editor'),
                                                    current_collaborator_role_filter.filter(role__icontains='speaker'))

            current_collaborator_item_list = Item.objects.filter(item_collaboratorroles__in=list(current_collaborator_role_all_filters.values_list('id', flat=True))).distinct().order_by('catalog_number')

            for current_collaborator_item in current_collaborator_item_list:
                xl_row.append(current_collaborator_item.catalog_number)
                xl_row.append(str(current_collaborator_item.item_collaboratorroles.get(collaborator=collaborator).role))

            if len(current_collaborator_item_list) > max_item_counts:
                max_item_counts = len(current_collaborator_item_list)

            sheet.append(xl_row)


        column_counter = 1
        while column_counter <= max_item_counts:
            header_cell.value = 'Item %s catalog number' %column_counter
            sheet_column_counter += 1
            header_cell = sheet.cell(row=1, column=sheet_column_counter )
            header_cell.value = 'Item %s collaborator role' %column_counter
            sheet_column_counter += 1
            header_cell = sheet.cell(row=1, column=sheet_column_counter )

            column_counter += 1


        new_workbook.save("collaborators_in_books.xlsx")
