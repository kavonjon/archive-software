"""
Service classes for metadata business logic.
Centralizes shared logic between views and tasks to avoid code duplication.
"""

from django.db.models import Count, Q
from django.contrib.auth.models import User
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.writer.excel import save_virtual_workbook
from datetime import datetime
from typing import Dict, Any, Optional

from .models import Collaborator
from .views import is_member_of_archivist, is_valid_param


class CollaboratorService:
    """
    Service class for collaborator-related business logic.
    Centralizes filtering and export logic to avoid duplication between views and tasks.
    """
    
    @staticmethod
    def build_filtered_queryset(user: User, filter_params: Dict[str, Any]):
        """
        Build a filtered collaborator queryset based on user permissions and filter parameters.
        EXACTLY matches the original working logic from collaborator_index view.
        
        Args:
            user: The user requesting the data (for permission checks)
            filter_params: Dictionary containing filter parameters
            
        Returns:
            QuerySet: Filtered and ordered collaborator queryset
        """
        # Start with all collaborators - EXACTLY like original
        qs = Collaborator.objects.all()
        
        # Extract filter parameters
        collection_contains_query = filter_params.get('collection_contains')
        native_languages_contains_query = filter_params.get('native_languages_contains')
        other_languages_contains_query = filter_params.get('other_languages_contains')
        name_contains_query = filter_params.get('name_contains')
        order_choice = filter_params.get('order_choice', 'name')
        
        # Apply user permission filters - EXACTLY like original
        if not is_member_of_archivist(user):
            qs = qs.exclude(anonymous=True)

        # Apply collection filter - EXACTLY like original
        if is_valid_param(collection_contains_query):
            qs = qs.filter(
                item_collaborators__collection__collection_abbr__icontains=collection_contains_query
            ).distinct()

        # Apply language filters - EXACTLY like original
        if is_valid_param(native_languages_contains_query):
            qs = qs.filter(native_languages__name__icontains=native_languages_contains_query)

        if is_valid_param(other_languages_contains_query):
            qs = qs.filter(other_languages__name__icontains=other_languages_contains_query)

        # Handle name filter with anonymous logic - EXACTLY like original
        anonymous_list = Collaborator.objects.none()
        if is_valid_param(name_contains_query):
            if name_contains_query.lower() == 'anonymous':
                # For anonymous search, get all anonymous collaborators with same filters applied
                anonymous_list = Collaborator.objects.filter(anonymous=True)
                
                # Apply the same non-name filters to anonymous list
                if is_valid_param(collection_contains_query):
                    anonymous_list = anonymous_list.filter(
                        item_collaborators__collection__collection_abbr__icontains=collection_contains_query
                    ).distinct()
                if is_valid_param(native_languages_contains_query):
                    anonymous_list = anonymous_list.filter(native_languages__name__icontains=native_languages_contains_query)
                if is_valid_param(other_languages_contains_query):
                    anonymous_list = anonymous_list.filter(other_languages__name__icontains=other_languages_contains_query)
            
            # Apply name filter to main queryset (this will find collaborators with "anonymous" in their name)
            qs = qs.filter(
                Q(name__icontains=name_contains_query) | 
                Q(nickname__icontains=name_contains_query) | 
                Q(other_names__icontains=name_contains_query)
            )

        # Apply distinct FIRST - EXACTLY like original
        qs = qs.distinct()

        # Apply union AFTER distinct - ONLY if we have anonymous results
        if anonymous_list.exists():
            qs = qs.union(anonymous_list)

        # Apply ordering - EXACTLY like original
        if order_choice == "updated":
            qs = qs.order_by('-updated')
        else:
            qs = qs.order_by('name')
            
        return qs
    
    @staticmethod
    def generate_export_workbook(collaborators_queryset):
        """
        Generate an Excel workbook from a collaborator queryset.
        
        Args:
            collaborators_queryset: QuerySet of collaborators to export
            
        Returns:
            Workbook: Excel workbook ready for saving
        """
        # Create workbook and styling
        new_workbook = Workbook()
        sheet = new_workbook.active
        
        # Define styles
        style_general = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
        style_personal = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
        style_languages = PatternFill(start_color="D7E4BD", end_color="D7E4BD", fill_type="solid")
        
        # Determine dynamic column counts
        # Handle union querysets which don't support annotate()
        try:
            native_language_list = collaborators_queryset.annotate(key_count=Count('native_languages__name'))
            native_language_counts = [entry.key_count for entry in native_language_list]
            max_native_language_counts = max(native_language_counts) if native_language_counts else 1
            
            other_language_list = collaborators_queryset.annotate(key_count=Count('other_languages__name'))
            other_language_counts = [entry.key_count for entry in other_language_list]
            max_other_language_counts = max(other_language_counts) if other_language_counts else 1
        except Exception:
            # Fallback for union querysets - manually count languages
            max_native_language_counts = 1
            max_other_language_counts = 1
            
            for collaborator in collaborators_queryset:
                native_count = collaborator.native_languages.count()
                other_count = collaborator.other_languages.count()
                
                if native_count > max_native_language_counts:
                    max_native_language_counts = native_count
                if other_count > max_other_language_counts:
                    max_other_language_counts = other_count
        
        if max_native_language_counts < 1:
            max_native_language_counts = 1
        if max_other_language_counts < 1:
            max_other_language_counts = 1
        
        # Create headers
        sheet_column_counter = 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter)
        
        headers = [
            ('Collaborator ID', style_general),
            ('Name', style_general),
            ('First Name', style_general),
            ('Last Name', style_general),
            ('Nickname', style_personal),
            ('Other Names', style_personal),
            ('Anonymous', style_personal),
            ('Birth Date', style_personal),
            ('Death Date', style_personal),
            ('Gender', style_personal),
            ('Origin', style_personal),
            ('Clan/Society', style_personal),
            ('Tribal Affiliations', style_personal),
            ('Collections', style_general),
        ]
        
        # Add static headers
        for header_text, style in headers:
            header_cell.value = header_text
            header_cell.fill = style
            sheet_column_counter += 1
            header_cell = sheet.cell(row=1, column=sheet_column_counter)

        # Add dynamic native language headers
        for i in range(1, max_native_language_counts + 1):
            header_cell.value = f'Native Language {i}'
            header_cell.fill = style_languages
            sheet_column_counter += 1
            header_cell = sheet.cell(row=1, column=sheet_column_counter)

        # Add dynamic other language headers
        for i in range(1, max_other_language_counts + 1):
            header_cell.value = f'Other Language {i}'
            header_cell.fill = style_languages
            sheet_column_counter += 1
            header_cell = sheet.cell(row=1, column=sheet_column_counter)

        # Add final header
        header_cell.value = 'Other Information'
        header_cell.fill = style_personal

        # Populate data rows
        for collaborator in collaborators_queryset:
            xl_row = []
            
            # Static data
            xl_row.append(collaborator.collaborator_id)
            xl_row.append(collaborator.name)
            xl_row.append(collaborator.firstname)
            xl_row.append(collaborator.lastname)
            xl_row.append(collaborator.nickname)
            xl_row.append(collaborator.other_names)
            xl_row.append('Yes' if collaborator.anonymous else 'No' if collaborator.anonymous is False else '')
            xl_row.append(collaborator.birthdate)
            xl_row.append(collaborator.deathdate)
            xl_row.append(collaborator.gender)
            xl_row.append(collaborator.origin)
            xl_row.append(collaborator.clan_society)
            xl_row.append(collaborator.tribal_affiliations)
            
            # Collections - get unique collection abbreviations
            collection_abbrs = list(set(
                collaborator.item_collaborators.filter(collection__isnull=False)
                .values_list('collection__collection_abbr', flat=True)
            ))
            xl_row.append(', '.join(sorted(collection_abbrs)))
            
            # Native languages
            native_language_rows = list(
                collaborator.native_languages.all()
                .values_list('name', flat=True)
                .order_by('name')
            )
            native_language_rows.extend([''] * (max_native_language_counts - len(native_language_rows)))
            xl_row.extend(native_language_rows)
            
            # Other languages
            other_language_rows = list(
                collaborator.other_languages.all()
                .values_list('name', flat=True)
                .order_by('name')
            )
            other_language_rows.extend([''] * (max_other_language_counts - len(other_language_rows)))
            xl_row.extend(other_language_rows)
            
            # Other information
            xl_row.append(collaborator.other_info)

            sheet.append(xl_row)

        return new_workbook
    
    @staticmethod
    def generate_export_filename() -> str:
        """
        Generate a timestamped filename for collaborator exports.
        
        Returns:
            str: Filename with timestamp
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f'collaborators-export-{timestamp}.xlsx'
