import os
import shutil
import logging
from django.conf import settings
from celery import shared_task
from datetime import datetime, timedelta
from metadata.models import Collection, Item, Languoid, Collaborator
from django.db.models import Min, Max
from .utils import parse_standardized_date
from .models import File
from .file_utils import ensure_directory_structure, save_item_metadata, list_item_files_by_numbers, get_item_files_path_by_numbers, update_file_metadata
from celery.exceptions import MaxRetriesExceededError
import hashlib
import mimetypes

logger = logging.getLogger(__name__)

@shared_task
def process_scanned_files(file_paths, scan_result):
    """
    Process files after virus scanning
    """
    if settings.SERVER_ROLE == 'public':
        destination = settings.PUBLIC_STORAGE_PATH if scan_result == 'clean' else None
    else:  # private
        destination = settings.MAIN_STORAGE_PATH if scan_result == 'clean' else None
    
    if not destination:
        logger.warning(f"Files failed virus scan: {file_paths}")
        return
    
    for file_path in file_paths:
        filename = os.path.basename(file_path)
        dest_path = os.path.join(destination, filename)
        try:
            shutil.move(file_path, dest_path)
            logger.info(f"Moved {file_path} to {dest_path}")
        except Exception as e:
            logger.error(f"Error moving file {file_path}: {str(e)}")

@shared_task
def cleanup_temp_files():
    """
    Clean up temporary files older than 24 hours
    """
    if settings.SERVER_ROLE != 'public':
        return
    
    now = datetime.now()
    temp_dir = settings.TEMP_STORAGE_PATH
    
    for filename in os.listdir(temp_dir):
        file_path = os.path.join(temp_dir, filename)
        file_modified = datetime.fromtimestamp(os.path.getmtime(file_path))
        if now - file_modified > timedelta(hours=24):
            try:
                os.remove(file_path)
                logger.info(f"Removed temporary file: {file_path}")
            except Exception as e:
                logger.error(f"Error removing temporary file {file_path}: {str(e)}")

@shared_task
def sync_public_files():
    """
    Sync files from private main storage to public storage
    """
    if settings.SERVER_ROLE != 'private':
        return
    
    # This would use a secure method to transfer files to the public server
    # Implementation depends on your specific requirements and infrastructure
    logger.info("Starting public file synchronization")
    # Example: Use rsync, sftp, or a custom API call to transfer files 

@shared_task
def update_collection_item_counts():
    """
    Update the item count for all collections
    """
    logger.info("Starting collection item count update")
    collections = Collection.objects.all()
    update_count = 0
    
    for collection in collections:
        # Count items related to this collection
        count = Item.objects.filter(collection=collection).count()
        
        # Only update if the count has changed
        if collection.item_count != count:
            collection.item_count = count
            collection.save(update_fields=['item_count'])
            update_count += 1
    
    logger.info(f"Updated item counts for {update_count} collections")
    return update_count

@shared_task
def update_collection_date_ranges():
    """
    Update the date range fields for all collections based on their items
    """
    logger.info("==== Collection Date Range Update Started ====")
    collections = Collection.objects.all()
    total_collections = collections.count()
    update_count = 0
    
    for i, collection in enumerate(collections, 1):
        # Get items for this collection
        items = Item.objects.filter(collection=collection)
        item_count = items.count()
        
        if item_count == 0:
            logger.info(f"Collection {collection.collection_abbr} has no associated items")
            continue
            
        logger.info(f"Processing collection: {collection.collection_abbr} with {item_count} items")
        
        # Find the earliest date (minimum of all item min dates)
        collection_min_date = items.exclude(collection_date_min=None).aggregate(Min('collection_date_min'))['collection_date_min__min']
        accession_min_date = items.exclude(accession_date_min=None).aggregate(Min('accession_date_min'))['accession_date_min__min']
        
        # Find the latest date (maximum of all item max dates)
        collection_max_date = items.exclude(collection_date_max=None).aggregate(Max('collection_date_max'))['collection_date_max__max']
        accession_max_date = items.exclude(accession_date_max=None).aggregate(Max('accession_date_max'))['accession_date_max__max']
        
        # Determine the overall min and max dates
        min_date = collection_min_date if collection_min_date else accession_min_date
        max_date = collection_max_date if collection_max_date else accession_max_date
        
        logger.info(f"Collection date values found:")
        logger.info(f"  - Collection date min: {collection_min_date}")
        logger.info(f"  - Collection date max: {collection_max_date}")
        logger.info(f"  - Accession date min: {accession_min_date}")
        logger.info(f"  - Accession date max: {accession_max_date}")
        logger.info(f"  - Overall min date: {min_date}")
        logger.info(f"  - Overall max date: {max_date}")
        
        # Format the date range text
        date_range = format_date_range(min_date, max_date)
        logger.info(f"  - Formatted date range: '{date_range}'")
        
        # Compare with existing values
        if (collection.date_range_min != min_date or 
            collection.date_range_max != max_date or 
            collection.date_range != date_range):
            
            logger.info(f"Updating collection: {collection.collection_abbr} - {collection.name}")
            if collection.date_range != date_range:
                logger.info(f"  Date range changing from '{collection.date_range}' to '{date_range}'")
            
            # Only update if we actually have dates
            if min_date is not None or max_date is not None:
                collection.date_range_min = min_date
                collection.date_range_max = max_date
                collection.date_range = date_range
                collection.save(update_fields=['date_range_min', 'date_range_max', 'date_range'])
                update_count += 1
            else:
                logger.info(f"  Skipping update - no valid dates found")
        
        # Progress indicator
        if i % 10 == 0:
            logger.info(f"Progress: {i}/{total_collections} collections processed")
    
    logger.info(f"==== Collection Date Range Update Completed ====")
    logger.info(f"Updated {update_count} of {total_collections} collections")
    return update_count

@shared_task
def update_item_date_ranges():
    """
    Update the min/max date fields for all items based on their text date fields
    """
    logger.info("=" * 50)
    logger.info("Item Date Range Update Started")
    logger.info("=" * 50)
    
    items = Item.objects.all()
    total_items = items.count()
    update_count = 0
    
    logger.info(f"Found {total_items} items to process")
    
    date_field_pairs = {
        'accession_date': ('accession_date_min', 'accession_date_max'),
        'cataloged_date': ('cataloged_date_min', 'cataloged_date_max'),
        'collection_date': ('collection_date_min', 'collection_date_max'),
        'creation_date': ('creation_date_min', 'creation_date_max'),
        'deposit_date': ('deposit_date_min', 'deposit_date_max'),
    }

    for i, item in enumerate(items, 1):
        item_updated = False
        updates_for_item = []
        
        for text_field, (min_field, max_field) in date_field_pairs.items():
            current_value = getattr(item, text_field)
            if not current_value:
                continue
                
            min_date, max_date = parse_standardized_date(current_value)
            current_min = getattr(item, min_field)
            current_max = getattr(item, max_field)
            
            if min_date != current_min or max_date != current_max:
                updates_for_item.append({
                    'field': text_field,
                    'old_min': current_min,
                    'old_max': current_max,
                    'new_min': min_date,
                    'new_max': max_date,
                    'text_value': current_value
                })
                setattr(item, min_field, min_date)
                setattr(item, max_field, max_date)
                item_updated = True
        
        if item_updated:
            logger.info(f"Updating item: {item.catalog_number}")
            for update in updates_for_item:
                logger.info(
                    f"  {update['field']}:\n"
                    f"    Text: {update['text_value']}\n"
                    f"    Old range: {update['old_min']} to {update['old_max']}\n"
                    f"    New range: {update['new_min']} to {update['new_max']}"
                )
            item.save()
            update_count += 1
            
        # Progress indicator every 100 items
        if i % 100 == 0:
            logger.info(f"Progress: {i}/{total_items} items processed ({(i/total_items)*100:.1f}%)")
    
    logger.info("=" * 50)
    logger.info(f"Item Date Range Update Completed")
    logger.info(f"Updated {update_count} of {total_items} items")
    logger.info("=" * 50)
    
    return update_count

def format_date_range(min_date, max_date):
    """
    Format date range as YYYY/MM/DD-YYYY/MM/DD with appropriate simplifications
    """
    if min_date is None and max_date is None:
        logger.info("  Cannot format date range - both min_date and max_date are None")
        return ""
    
    # If only one date is available, use it for both min and max
    if min_date is None and max_date is not None:
        logger.info("  Only max_date is available, using it for both min and max")
        min_date = max_date
    elif max_date is None and min_date is not None:
        logger.info("  Only min_date is available, using it for both min and max")
        max_date = min_date
    
    try:
        # Format full dates
        min_str = min_date.strftime("%Y/%m/%d")
        max_str = max_date.strftime("%Y/%m/%d")
        
        # Same year
        if min_date.year == max_date.year:
            # Same year, same month
            if min_date.month == max_date.month:
                # Same day (exact date)
                if min_date.day == max_date.day:
                    return f"{min_date.year}/{min_date.month:02d}/{min_date.day:02d}"
                # Same year, same month, different days
                else:
                    return f"{min_date.year}/{min_date.month:02d}/{min_date.day:02d}-{max_date.day:02d}"
            # Same year, different months
            else:
                # Check if it spans the entire year
                if min_date.month == 1 and min_date.day == 1 and max_date.month == 12 and max_date.day == 31:
                    return f"{min_date.year}"
                else:
                    return f"{min_date.year}/{min_date.month:02d}-{max_date.month:02d}"
        # Different years
        else:
            # Check if it spans entire years (Jan 1 to Dec 31)
            if min_date.month == 1 and min_date.day == 1 and max_date.month == 12 and max_date.day == 31:
                return f"{min_date.year}-{max_date.year}"
            else:
                return f"{min_str}-{max_str}"
    except Exception as e:
        logger.error(f"Error formatting date range: {str(e)}")
        return ""

@shared_task(bind=True, max_retries=3)
def export_item_metadata(self, item_id):
    """
    Export an item's metadata to JSON file asynchronously with retry logic
    """
    try:
        item = Item.objects.get(pk=item_id)
        logger.info(f"Starting metadata export for item {item.catalog_number}")
        
        if not item.collection or not item.pk:
            logger.error(f"Cannot export metadata: No collection or primary key for item {item}")
            return False
            
        # Ensure the directory structure exists using new path format
        logger.info(f"Ensuring directory structure for {item.collection.collection_abbr}/{item.catalog_number}")
        ensure_directory_structure(None, None, item.collection.collection_abbr, item.catalog_number)
        
        # Create a dictionary of metadata to export
        metadata = {
            'id': item.pk,
            'catalog_number': item.catalog_number,
            'collection': item.collection.collection_abbr if item.collection else None,
            'english_title': item.english_title,
            'indigenous_title': item.indigenous_title,
            'description': item.description_scope_and_content,
            'collection_date': item.collection_date,
            'creation_date': item.creation_date,
            'languages': [lang.name for lang in item.language.all()],
            'resource_type': item.resource_type,
            'genre': item.genre,
            'access_level': item.item_access_level,
            'modified_by': item.modified_by,
            'last_updated': item.updated.isoformat() if item.updated else None,
        }
        
        # Add the list of available files
        available_files = item.get_item_files()
        logger.info(f"Available files: {available_files}")
        metadata['available_files'] = available_files
        
        # Add detailed file information
        files_data = {}
        total_bytes = 0
        file_objects = File.objects.filter(item=item)
        logger.info(f"File objects: {[f.filename for f in file_objects]}")
        
        for file_obj in file_objects:
            total_bytes += file_obj.filesize or 0
            files_data[file_obj.filename] = {
                'id': str(file_obj.uuid),
                'checksum': file_obj.checksum,
                'ext': file_obj.get_extension(),
                'size': file_obj.filesize,
                'mimetype': file_obj.mimetype,
                'key': file_obj.filename,
                'metadata': file_obj.get_metadata_dict()
            }
        
        metadata['files'] = {
            'count': len(file_objects),
            'total_bytes': total_bytes,
            'entries': files_data
        }
        
        # Save the metadata to a file
        logger.info(f"Saving metadata to file for {item.catalog_number}")
        result = save_item_metadata(item.collection.pk, item.pk, metadata, 
                                  item.collection.collection_abbr, item.catalog_number)
        if result:
            logger.info(f"Metadata saved successfully for {item.catalog_number}")
        else:
            logger.error(f"Failed to save metadata for {item.catalog_number}")
            raise Exception("Failed to save metadata file")
        return result
        
    except Item.DoesNotExist:
        logger.error(f"Item with ID {item_id} does not exist")
        return False
    except Exception as e:
        logger.error(f"Error exporting metadata for item {item_id}: {str(e)}")
        try:
            self.retry(exc=e, countdown=60)  # Retry after 1 minute
        except MaxRetriesExceededError:
            logger.error(f"Failed to export metadata for item {item_id} after max retries")
            return False 

@shared_task(bind=True, max_retries=3)
def save_file_selection(self, item_id, selected_files):
    """
    Save the list of selected files to the item's metadata and create File objects asynchronously
    
    Args:
        item_id: The ID of the item to process
        selected_files: List of filenames that should be included in this item
    """
    try:
        item = Item.objects.get(pk=item_id)
        
        if not item.collection:
            logger.error(f"No collection found for item {item_id}")
            return False
            
        # Ensure the directory structure exists
        ensure_directory_structure(None, None, item.collection.collection_abbr, item.catalog_number)
        
        # Get available files
        available_files = list_item_files_by_numbers(item.collection.collection_abbr, item.catalog_number)
        
        # Process each selected file
        for filename in selected_files:
            if filename in available_files:
                try:
                    # Create relative path
                    rel_path = os.path.join(
                        item.collection.collection_abbr,
                        item.catalog_number,
                        filename
                    )
                    
                    # Get absolute path
                    abs_path = os.path.join(
                        get_item_files_path_by_numbers(item.collection.collection_abbr, item.catalog_number),
                        filename
                    )
                    
                    # Get file stats
                    file_stats = os.stat(abs_path)
                    file_size = file_stats.st_size
                    
                    # Get mimetype
                    mime_type, _ = mimetypes.guess_type(filename)
                    mime_type = mime_type or 'application/octet-stream'
                    
                    # Calculate checksum
                    checksum = ""
                    try:
                        with open(abs_path, 'rb') as f:
                            sha256 = hashlib.sha256()
                            for byte_block in iter(lambda: f.read(4096), b""):
                                sha256.update(byte_block)
                            checksum = sha256.hexdigest()
                    except Exception as e:
                        logger.error(f"Error calculating checksum for {abs_path}: {str(e)}")
                    
                    # Get file extension
                    _, file_ext = os.path.splitext(filename)
                    file_ext = file_ext.lower()[1:] if file_ext else ''
                    
                    # Create or update File object
                    file_obj, created = File.objects.get_or_create(
                        filename=filename,
                        item=item,
                        defaults={
                            'filepath': rel_path,
                            'filetype': file_ext,
                            'filesize': file_size,
                            'checksum': checksum,
                            'mimetype': mime_type,
                            'access_level': item.item_access_level,
                            'title': filename,
                            'modified_by': item.modified_by
                        }
                    )
                    
                    # If file exists, update its metadata
                    if not created:
                        file_obj.filepath = rel_path
                        file_obj.filetype = file_ext
                        file_obj.filesize = file_size
                        file_obj.checksum = checksum
                        file_obj.mimetype = mime_type
                        file_obj.modified_by = item.modified_by
                        file_obj.save()
                        
                except Exception as e:
                    logger.error(f"Error processing file {filename} for item {item_id}: {str(e)}")
                    continue
        
        # Delete File objects for files that are no longer selected
        File.objects.filter(item=item).exclude(filename__in=selected_files).delete()
        
        # Update the file metadata
        update_file_metadata(item.collection.pk, item.pk, selected_files,
                           item.collection.collection_abbr, item.catalog_number)
        
        return True
        
    except Item.DoesNotExist:
        logger.error(f"Item {item_id} not found")
        return False
    except Exception as e:
        logger.error(f"Error processing files for item {item_id}: {str(e)}")
        self.retry(exc=e, countdown=60)  # Retry after 60 seconds

@shared_task(bind=True)
def generate_collaborator_export(self, user_id, filter_params):
    """
    Generate collaborator export spreadsheet asynchronously
    
    Args:
        user_id: ID of the user requesting the export
        filter_params: Dictionary of filter parameters from the request
    """
    import tempfile
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.writer.excel import save_virtual_workbook
    from django.contrib.auth.models import User
    from django.db.models import Count, Q
    from django.core.files.storage import default_storage
    from django.core.files.base import ContentFile
    from datetime import datetime
    
    try:
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"=== STARTING COLLABORATOR EXPORT TASK ===")
        logger.info(f"Task ID: {self.request.id}")
        logger.info(f"User ID: {user_id}")
        logger.info(f"Filter params: {filter_params}")
        
        # Add detailed debugging
        logger.info(f"=== DEBUGGING EXPORT STEP BY STEP ===")
        logger.info(f"Step 1: About to get user object")
        
        # Get the user for permission checks
        user = User.objects.get(pk=user_id)
        logger.info(f"Step 2: Found user: {user.username}")
        
        # Use service to build filtered queryset
        logger.info(f"Step 3: About to call CollaboratorService.build_filtered_queryset")
        from .services import CollaboratorService
        
        logger.info(f"Step 4: Calling service with filters: {filter_params}")
        collaborators_in_qs = CollaboratorService.build_filtered_queryset(user, filter_params)
        
        logger.info(f"Step 5: Service returned queryset, counting results...")
        count = collaborators_in_qs.count()
        logger.info(f"Step 6: Found {count} collaborators to export")
        
        # Test if we can actually iterate over the queryset
        logger.info(f"Step 7: Testing queryset iteration...")
        try:
            first_collaborator = collaborators_in_qs.first()
            if first_collaborator:
                logger.info(f"Step 8: First collaborator: {first_collaborator.name}")
            else:
                logger.info(f"Step 8: No collaborators in queryset")
        except Exception as qs_error:
            logger.error(f"Step 8 ERROR: Cannot iterate queryset: {qs_error}")
            raise
        
        # Use service to generate workbook
        logger.info(f"Step 9: About to generate workbook...")
        new_workbook = CollaboratorService.generate_export_workbook(collaborators_in_qs)
        logger.info(f"Step 10: Workbook generated successfully")
        
        filename = CollaboratorService.generate_export_filename()
        logger.info(f"Step 11: Generated filename: {filename}")
        
        # Convert workbook to bytes for storage
        logger.info(f"Step 12: Converting workbook to bytes...")
        excel_content = save_virtual_workbook(new_workbook)
        logger.info(f"Step 13: Excel content generated, size: {len(excel_content)} bytes")
        
        # Use Django storage with proper error handling and validation
        from django.core.files.storage import default_storage
        from django.core.files.base import ContentFile
        from django.conf import settings
        import os
        
        logger.info("Validating Django storage configuration...")
        
        # Validate Django storage is properly configured
        if not hasattr(settings, 'MEDIA_ROOT'):
            raise Exception("Django MEDIA_ROOT not configured in Celery worker")
        
        if not default_storage.location:
            raise Exception("Django storage location not available in Celery worker")
        
        logger.info(f"Storage backend: {default_storage.__class__.__name__}")
        logger.info(f"Storage location: {default_storage.location}")
        
        # Ensure exports directory exists
        exports_dir = os.path.join(default_storage.location, 'exports')
        if not os.path.exists(exports_dir):
            logger.info(f"Creating exports directory: {exports_dir}")
            os.makedirs(exports_dir, exist_ok=True)
        
        # Verify directory is writable
        if not os.access(exports_dir, os.W_OK):
            raise Exception(f"Exports directory not writable: {exports_dir}")
        
        # Save file using Django storage with atomic operations
        logger.info(f"Saving file using Django storage: exports/{filename}")
        
        try:
            # Use ContentFile for proper Django storage handling
            content_file = ContentFile(excel_content)
            file_path = default_storage.save(f'exports/{filename}', content_file)
            logger.info(f"File saved to Django storage: {file_path}")
            
            # Verify file exists and get size
            if default_storage.exists(file_path):
                file_size = default_storage.size(file_path)
                logger.info(f"File verified in storage: {file_path} ({file_size} bytes)")
            else:
                raise Exception(f"File not found in storage after save: {file_path}")
            
            # Double-check with filesystem
            full_path = os.path.join(default_storage.location, file_path)
            if not os.path.exists(full_path):
                raise Exception(f"File not found on filesystem: {full_path}")
            
            logger.info(f"Collaborator export completed successfully: {file_path}")
            
            return {
                'success': True,
                'file_path': file_path,
                'filename': filename,
                'count': collaborators_in_qs.count()
            }
            
        except Exception as storage_error:
            logger.error(f"Django storage operation failed: {storage_error}")
            raise Exception(f"File save failed: {storage_error}")
        
    except Exception as e:
        logger.error(f"Error generating collaborator export: {str(e)}")
        return {
            'success': False,
            'error': str(e)
        }


# ============================================================================
# LANGUOID HIERARCHY TASKS
# ============================================================================

def get_all_ancestors(languoid, visited=None):
    """
    Get all ancestors (parent, grandparent, etc.) with circular ref protection.
    """
    if visited is None:
        visited = set()
    
    ancestors = []
    current = languoid.parent_languoid
    
    while current and current.id not in visited:
        visited.add(current.id)
        ancestors.append(current)
        current = current.parent_languoid
    
    return ancestors


def get_all_descendents(languoid, visited=None):
    """
    Recursively get all descendents at all levels, with circular reference protection.
    """
    if visited is None:
        visited = set()
    
    if languoid.id in visited:
        return []
    
    visited.add(languoid.id)
    descendents = []
    
    # Get direct children
    children = Languoid.objects.filter(parent_languoid=languoid)
    
    for child in children:
        if child.id not in visited:
            descendents.append(child)
            descendents.extend(get_all_descendents(child, visited))
    
    return descendents


@shared_task(bind=True, max_retries=3)
def update_languoid_hierarchy_task(self, languoid_id, needs_orphaning=False, old_parent_id=None):
    """
    UNIFIED TASK: Update hierarchy relationships for a languoid.
    
    This handles:
    1. Orphaning dialect children (if level changed from language)
    2. Updating descendents M2M for this languoid and NEW parent's ancestors
    3. Updating descendents M2M for OLD parent's ancestors (if parent changed)
    
    Priority 9 (highest) - user is waiting for tree to update.
    
    Args:
        languoid_id: The languoid that was saved
        needs_orphaning: True if level changed from language to other
        old_parent_id: ID of old parent if parent_languoid changed
    """
    try:
        languoid = Languoid.objects.select_related(
            'parent_languoid',
            'parent_languoid__parent_languoid',
            'parent_languoid__parent_languoid__parent_languoid'
        ).get(id=languoid_id)
        
        results = {
            'languoid_name': languoid.name,
            'orphaned_dialects': 0,
            'updated_descendents': 0
        }
        
        # STEP 1: Orphan dialect children if needed
        if needs_orphaning:
            logger.warning(
                f"Orphaning dialects for '{languoid.name}' "
                f"(level changed from language)"
            )
            
            dialect_children = Languoid.objects.filter(
                parent_languoid=languoid,
                level_glottolog='dialect'
            )
            
            orphaned_names = []
            for dialect in dialect_children:
                dialect.parent_languoid = None
                # Save with specific update_fields to trigger hierarchy recalc
                dialect.save(update_fields=[
                    'parent_languoid',
                    'family_languoid',
                    'pri_subgroup_languoid',
                    'sec_subgroup_languoid'
                ])
                orphaned_names.append(dialect.name)
                results['orphaned_dialects'] += 1
            
            if orphaned_names:
                logger.info(f"Orphaned {len(orphaned_names)} dialects: {', '.join(orphaned_names[:5])}")
                if len(orphaned_names) > 5:
                    logger.info(f"  ... and {len(orphaned_names) - 5} more")
        
        # STEP 2: Update descendents M2M for this languoid and all ancestors
        ancestors = get_all_ancestors(languoid)
        ancestors.append(languoid)
        
        for ancestor in ancestors:
            descendents = get_all_descendents(ancestor)
            ancestor.descendents.set(descendents)
            results['updated_descendents'] += 1
        
        # STEP 3: Update OLD parent's ancestor chain (if parent changed)
        if old_parent_id:
            try:
                old_parent = Languoid.objects.select_related(
                    'parent_languoid',
                    'parent_languoid__parent_languoid',
                    'parent_languoid__parent_languoid__parent_languoid'
                ).get(id=old_parent_id)
                
                logger.info(
                    f"Parent changed for '{languoid.name}': "
                    f"updating old parent '{old_parent.name}' ancestor chain"
                )
                
                # Get old parent + all its ancestors
                old_ancestors = get_all_ancestors(old_parent)
                old_ancestors.append(old_parent)
                
                # Recalculate descendents for each (they lost the moved subtree)
                for old_ancestor in old_ancestors:
                    descendents = get_all_descendents(old_ancestor)
                    old_ancestor.descendents.set(descendents)
                    results['updated_descendents'] += 1
                
                logger.info(
                    f"Updated old parent chain: {len(old_ancestors)} languoids "
                    f"(removed '{languoid.name}' and its subtree)"
                )
            except Languoid.DoesNotExist:
                logger.warning(
                    f"Old parent {old_parent_id} not found for '{languoid.name}' "
                    f"(may have been deleted)"
                )
        
        logger.info(
            f"Updated hierarchy for '{languoid.name}': "
            f"{results['orphaned_dialects']} dialects orphaned, "
            f"{results['updated_descendents']} languoids updated"
        )
        
        return results
        
    except Languoid.DoesNotExist:
        logger.error(f"Languoid {languoid_id} not found")
        return None
    except Exception as e:
        logger.error(f"Error updating hierarchy: {e}")
        import traceback
        traceback.print_exc()
        raise self.retry(exc=e, countdown=2 ** self.request.retries)


@shared_task(bind=True, max_retries=3)
def cascade_hierarchy_to_dialects_task(self, languoid_id):
    """
    PRIORITY 2: Cascade hierarchy changes to all dialect descendants.
    This runs AFTER the unified hierarchy task completes.
    
    When a family or language changes hierarchy fields, all descendant dialects
    need to update their hierarchy FKs by re-saving.
    """
    try:
        languoid = Languoid.objects.prefetch_related('descendents').get(id=languoid_id)
        
        # Find all dialect descendants
        dialect_descendants = []
        for descendent in languoid.descendents.all():
            if descendent.level_glottolog == 'dialect':
                dialect_descendants.append(descendent)
        
        if not dialect_descendants:
            logger.info(f"No dialect descendants to cascade for '{languoid.name}'")
            return 0
        
        logger.info(f"Cascading hierarchy to {len(dialect_descendants)} dialects from '{languoid.name}'")
        
        # Update each dialect's hierarchy FKs
        # Save triggers pre_save signal → derive_hierarchy_fks()
        updated_count = 0
        for dialect in dialect_descendants:
            dialect.save(update_fields=[
                'family_languoid',
                'pri_subgroup_languoid', 
                'sec_subgroup_languoid'
            ])
            updated_count += 1
        
        logger.info(f"Cascaded hierarchy updates to {updated_count} dialects")
        return updated_count
        
    except Languoid.DoesNotExist:
        logger.error(f"Languoid {languoid_id} not found")
        return None
    except Exception as e:
        logger.error(f"Error cascading to dialects: {e}")
        raise self.retry(exc=e, countdown=2 ** self.request.retries)


# ============================================================================
# BATCH-OPTIMIZED LANGUOID HIERARCHY TASKS
# ============================================================================

@shared_task(bind=True, max_retries=3)
def recalculate_ancestor_descendents_task(self, ancestor_ids):
    """
    BATCH-OPTIMIZED: Recalculate descendents M2M for multiple ancestors at once.
    
    This task processes a deduplicated list of ancestors that were affected
    by a batch save operation, eliminating redundant recalculations.
    
    Args:
        ancestor_ids: List of languoid IDs whose descendents need recalculation
    
    Returns:
        Number of ancestors updated
    """
    try:
        logger.info(f"[BATCH] Recalculating descendents for {len(ancestor_ids)} ancestors")
        
        updated_count = 0
        for ancestor_id in ancestor_ids:
            try:
                ancestor = Languoid.objects.get(id=ancestor_id)
                descendents = get_all_descendents(ancestor)
                ancestor.descendents.set(descendents)
                updated_count += 1
            except Languoid.DoesNotExist:
                logger.warning(f"[BATCH] Ancestor {ancestor_id} not found (may have been deleted)")
                continue
        
        logger.info(f"[BATCH] Updated descendents for {updated_count}/{len(ancestor_ids)} ancestors")
        return updated_count
        
    except Exception as e:
        logger.error(f"[BATCH] Error recalculating ancestor descendents: {e}")
        import traceback
        traceback.print_exc()
        raise self.retry(exc=e, countdown=2 ** self.request.retries)


@shared_task(bind=True, max_retries=3)
def orphan_dialects_batch_task(self, languoid_ids):
    """
    BATCH-OPTIMIZED: Orphan dialects for multiple languoids that changed from language level.
    
    Args:
        languoid_ids: List of languoid IDs that changed from language to other level
    
    Returns:
        Total number of dialects orphaned
    """
    try:
        logger.info(f"[BATCH] Orphaning dialects for {len(languoid_ids)} languoids")
        
        total_orphaned = 0
        for languoid_id in languoid_ids:
            try:
                languoid = Languoid.objects.get(id=languoid_id)
                
                dialect_children = Languoid.objects.filter(
                    parent_languoid=languoid,
                    level_glottolog='dialect'
                )
                
                orphaned_names = []
                for dialect in dialect_children:
                    dialect.parent_languoid = None
                    # Set batch flag to prevent individual signals
                    dialect._skip_async_tasks = True
                    dialect.save(update_fields=[
                        'parent_languoid',
                        'family_languoid',
                        'pri_subgroup_languoid',
                        'sec_subgroup_languoid'
                    ])
                    orphaned_names.append(dialect.name)
                    total_orphaned += 1
                
                if orphaned_names:
                    logger.info(
                        f"[BATCH] Orphaned {len(orphaned_names)} dialects for '{languoid.name}': "
                        f"{', '.join(orphaned_names[:5])}"
                    )
                    if len(orphaned_names) > 5:
                        logger.info(f"  ... and {len(orphaned_names) - 5} more")
                        
            except Languoid.DoesNotExist:
                logger.warning(f"[BATCH] Languoid {languoid_id} not found")
                continue
        
        logger.info(f"[BATCH] Total orphaned dialects: {total_orphaned}")
        return total_orphaned
        
    except Exception as e:
        logger.error(f"[BATCH] Error orphaning dialects: {e}")
        import traceback
        traceback.print_exc()
        raise self.retry(exc=e, countdown=2 ** self.request.retries)


# ============================================================================
# LANGUOID LIST CACHE WARMING TASKS
# ============================================================================

def build_languoid_list_cache():
    """
    Utility function to build the cached languoid list response.
    
    This function does the expensive work:
    - Queries all languoids
    - Sorts them in tree order (depth-first: parent → all descendants → next sibling)
    - Serializes them to JSON
    - Returns the serialized data ready for caching
    
    Tree ordering ensures:
    - Top-level items (families with no parent) appear first, alphabetically
    - All descendants appear immediately below their parent
    - Siblings are sorted alphabetically
    
    Used by:
    - warm_languoid_list_cache task (background refresh)
    - InternalLanguoidViewSet (on cache miss)
    """
    from internal_api.serializers import InternalLanguoidSerializer
    
    logger.info("[Cache Warming] Building languoid list cache...")
    
    # Query all languoids with optimized select/prefetch
    queryset = Languoid.objects.select_related(
        'family_languoid',
        'parent_languoid',
        'pri_subgroup_languoid',
        'sec_subgroup_languoid'
    ).prefetch_related('child_languoids', 'item_languages')
    
    # Get all languoids as a list for tree building
    all_languoids = list(queryset)
    
    # Build a map for quick parent → children lookup
    children_map = {}
    for languoid in all_languoids:
        parent_id = languoid.parent_languoid_id
        if parent_id not in children_map:
            children_map[parent_id] = []
        children_map[parent_id].append(languoid)
    
    # Sort children alphabetically (by name) for each parent
    for parent_id in children_map:
        children_map[parent_id].sort(key=lambda l: l.name.lower())
    
    # Find top-level items (those without a parent)
    top_level = [l for l in all_languoids if l.parent_languoid_id is None]
    top_level.sort(key=lambda l: l.name.lower())
    
    # Depth-first tree traversal: add parent, then all descendants recursively
    result = []
    processed = set()
    
    def add_languoid_with_children(languoid):
        """Recursively add languoid and all its descendants in tree order"""
        if languoid.id in processed:
            return
        
        processed.add(languoid.id)
        result.append(languoid)
        
        # Add all direct children (already sorted alphabetically)
        children = children_map.get(languoid.id, [])
        for child in children:
            add_languoid_with_children(child)
    
    # Process all top-level items and their descendants
    for languoid in top_level:
        add_languoid_with_children(languoid)
    
    # Add any orphaned items (shouldn't happen, but handle gracefully)
    orphans = [l for l in all_languoids if l.id not in processed]
    if orphans:
        logger.warning(f"[Cache Warming] Found {len(orphans)} orphaned languoids")
        orphans.sort(key=lambda l: l.name.lower())
        for orphan in orphans:
            add_languoid_with_children(orphan)
    
    # Serialize to JSON (same as API response)
    serializer = InternalLanguoidSerializer(result, many=True)
    data = serializer.data
    
    logger.info(f"[Cache Warming] Built cache with {len(data)} languoids in tree order")
    return data


@shared_task(bind=True, max_retries=3)
def warm_languoid_list_cache(self):
    """
    Background task to warm/refresh the languoid list cache.
    
    This task:
    - Builds the full languoid list response
    - Stores it in Redis cache with 10-minute TTL
    - Uses a lock to prevent concurrent rebuilds
    - Runs periodically via Celery Beat (every 9 minutes)
    - Runs immediately after languoid saves/deletes
    
    Users never wait for this - it happens in the background.
    """
    from django.core.cache import cache
    import time
    
    lock_key = 'languoid_list_cache_lock'
    cache_key = 'languoid_list_full'
    
    try:
        # Try to acquire lock (prevents concurrent rebuilds)
        lock_acquired = cache.add(lock_key, 'locked', timeout=120)  # 2-minute lock
        
        if not lock_acquired:
            logger.info("[Cache Warming] Another cache build is in progress, skipping")
            return {'status': 'skipped', 'reason': 'lock_held'}
        
        start_time = time.time()
        logger.info("[Cache Warming] Starting languoid list cache rebuild...")
        
        # Build the cache data
        data = build_languoid_list_cache()
        
        # Store in Redis with 10-minute TTL
        cache.set(cache_key, data, timeout=600)
        
        elapsed = time.time() - start_time
        logger.info(f"[Cache Warming] Cache rebuilt successfully in {elapsed:.2f}s")
        
        return {
            'status': 'success',
            'languoid_count': len(data),
            'elapsed_seconds': elapsed
        }
        
    except Exception as e:
        logger.error(f"[Cache Warming] Error building cache: {e}")
        import traceback
        traceback.print_exc()
        raise self.retry(exc=e, countdown=30)  # Retry after 30 seconds
        
    finally:
        # Always release the lock
        cache.delete(lock_key)


@shared_task
def invalidate_and_warm_languoid_cache():
    """
    Invalidate the languoid list cache and trigger immediate rebuild.
    
    Called by Django signals when a languoid is saved or deleted.
    This ensures users always see fresh data after edits.
    """
    from django.core.cache import cache
    
    cache_key = 'languoid_list_full'
    
    # Invalidate existing cache
    cache.delete(cache_key)
    logger.info("[Cache Invalidation] Languoid list cache invalidated")
    
    # Trigger background rebuild (doesn't block the signal)
    warm_languoid_list_cache.apply_async(priority=8)


# ============================================================================
# COLLABORATOR LIST CACHE WARMING TASKS
# ============================================================================

def build_collaborator_list_cache():
    """
    Utility function to build the cached collaborator list response.
    
    This function does the expensive work:
    - Queries all collaborators with optimized prefetch
    - Sorts them with locale-aware collation (è, é sort near 'e')
    - Serializes them to JSON
    - Returns the serialized data ready for caching
    
    Used by:
    - warm_collaborator_list_cache task (background refresh)
    - InternalCollaboratorViewSet (on cache miss)
    """
    from internal_api.serializers import InternalCollaboratorBatchSerializer
    from django.db.models.functions import Lower, Collate
    
    logger.info("[Cache Warming] Building collaborator list cache...")
    
    # Query all collaborators with optimized prefetch and locale-aware sorting
    queryset = Collaborator.objects.all().prefetch_related(
        'native_languages', 
        'other_languages'
    ).order_by(
        Collate(Lower('last_names'), 'en-US-x-icu'),
        Collate(Lower('first_names'), 'en-US-x-icu'),
        Collate(Lower('full_name'), 'en-US-x-icu'),
        'collaborator_id'
    )
    
    # Serialize to JSON (same as API response)
    serializer = InternalCollaboratorBatchSerializer(queryset, many=True)
    data = serializer.data
    
    logger.info(f"[Cache Warming] Built cache with {len(data)} collaborators")
    return data


@shared_task(bind=True, max_retries=3)
def warm_collaborator_list_cache(self):
    """
    Background task to warm/refresh the collaborator list cache.
    
    This task:
    - Builds the full collaborator list response
    - Stores it in Redis cache with 10-minute TTL
    - Uses a lock to prevent concurrent rebuilds
    - Runs periodically via Celery Beat (every 9 minutes)
    - Runs immediately after collaborator saves/deletes
    
    Users never wait for this - it happens in the background.
    """
    from django.core.cache import cache
    import time
    
    lock_key = 'collaborator_list_cache_lock'
    cache_key = 'collaborator_list_full'
    
    try:
        # Try to acquire lock (prevents concurrent rebuilds)
        lock_acquired = cache.add(lock_key, 'locked', timeout=120)  # 2-minute lock
        
        if not lock_acquired:
            logger.info("[Cache Warming] Another cache build is in progress, skipping")
            return {'status': 'skipped', 'reason': 'lock_held'}
        
        start_time = time.time()
        logger.info("[Cache Warming] Starting collaborator list cache rebuild...")
        
        # Build the cache data
        data = build_collaborator_list_cache()
        
        # Store in Redis with 10-minute TTL
        cache.set(cache_key, data, timeout=600)
        
        elapsed = time.time() - start_time
        logger.info(f"[Cache Warming] Cache rebuilt successfully in {elapsed:.2f}s")
        
        return {
            'status': 'success',
            'collaborator_count': len(data),
            'elapsed_seconds': elapsed
        }
        
    except Exception as e:
        logger.error(f"[Cache Warming] Error building cache: {e}")
        import traceback
        traceback.print_exc()
        raise self.retry(exc=e, countdown=30)  # Retry after 30 seconds
        
    finally:
        # Always release the lock
        cache.delete(lock_key)


@shared_task
def invalidate_and_warm_collaborator_cache():
    """
    Invalidate the collaborator list cache and trigger immediate rebuild.
    
    Called by Django signals when a collaborator is saved or deleted.
    This ensures users always see fresh data after edits.
    """
    from django.core.cache import cache
    
    cache_key = 'collaborator_list_full'
    
    # Invalidate existing cache
    cache.delete(cache_key)
    logger.info("[Cache Invalidation] Collaborator list cache invalidated")
    
    # Trigger background rebuild (doesn't block the signal)
    warm_collaborator_list_cache.apply_async(priority=8)


# ============================================================================
# LANGUOID EXPORT TASKS
# ============================================================================

@shared_task(bind=True, max_retries=3)
def generate_languoid_export_task(self, export_id, mode, ids):
    """
    Background task to generate languoid export for large datasets.
    
    Args:
        export_id: Unique identifier for this export (used for filename and tracking)
        mode: 'filtered' | 'selected'
        ids: List of languoid IDs to export
    
    Returns:
        dict with status, filename, and error info
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from datetime import datetime
    from django.core.files.storage import default_storage
    from django.core.files.base import ContentFile
    from django.conf import settings
    import os
    
    logger.info(f"[EXPORT TASK] Starting export {export_id}: {len(ids)} languoids")
    
    try:
        from django.db.models import Count
        
        # Fetch languoids with optimized queries
        queryset = Languoid.objects.filter(id__in=ids).select_related(
            'parent_languoid',
            'family_languoid',
            'pri_subgroup_languoid',
            'sec_subgroup_languoid'
        ).annotate(
            item_count=Count('item_languages')
        )
        
        languoids = list(queryset)
        logger.info(f"[EXPORT TASK] Found {len(languoids)} languoids")
        
        # Sort languoids by hierarchy (tree structure)
        languoid_dict = {l.id: l for l in languoids}
        
        def get_sort_key(languoid):
            """Generate hierarchical sort key"""
            path = []
            current = languoid
            visited = set()
            
            while current:
                if current.id in visited:
                    break
                visited.add(current.id)
                path.insert(0, current.name.lower() if current.name else '')
                
                if current.parent_languoid and current.parent_languoid.id in languoid_dict:
                    current = languoid_dict[current.parent_languoid.id]
                elif current.parent_languoid:
                    path.insert(0, current.parent_languoid.name.lower() if current.parent_languoid.name else '')
                    break
                else:
                    break
            
            return tuple(path)
        
        languoids.sort(key=get_sort_key)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Languoids"
        
        # Define headers
        headers = [
            'Name', 'Name Abbreviation', 'Glottocode', 'ISO 639-3', '# of Items',
            'Level (Glottolog)', 'Level (NAL)',
            'Parent Languoid', 'Parent Languoid Abbreviation', 'Parent Languoid Glottocode',
            'Family', 'Family Abbreviation', 'Family Glottocode',
            'Primary Subfamily', 'Primary Subfamily Abbreviation', 'Primary Subfamily Glottocode',
            'Secondary Subfamily', 'Secondary Subfamily Abbreviation', 'Secondary Subfamily Glottocode',
            'Alternate Names', 'Region', 'Latitude', 'Longitude', 'Tribes', 'Notes',
        ]
        
        # Write header row with styling
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Helper function for safe values
        def safe_value(value):
            if value is None:
                return ''
            if isinstance(value, (list, tuple)):
                return ', '.join(str(v) for v in value)
            if isinstance(value, bool):
                return 'Yes' if value else 'No'
            return str(value)
        
        # Write data rows
        for row_num, languoid in enumerate(languoids, 2):
            ws.cell(row=row_num, column=1).value = safe_value(languoid.name)
            ws.cell(row=row_num, column=2).value = safe_value(languoid.name_abbrev)
            ws.cell(row=row_num, column=3).value = safe_value(languoid.glottocode)
            ws.cell(row=row_num, column=4).value = safe_value(languoid.iso)
            
            # Item count - blank for families, show count for languages/dialects
            if languoid.level_glottolog == 'family':
                ws.cell(row=row_num, column=5).value = ''
            else:
                ws.cell(row=row_num, column=5).value = languoid.item_count
            
            ws.cell(row=row_num, column=6).value = languoid.get_level_glottolog_display() if languoid.level_glottolog else ''
            ws.cell(row=row_num, column=7).value = languoid.get_level_nal_display() if languoid.level_nal else ''
            
            # Parent Languoid
            ws.cell(row=row_num, column=8).value = languoid.parent_languoid.name if languoid.parent_languoid else ''
            ws.cell(row=row_num, column=9).value = safe_value(languoid.parent_languoid.name_abbrev) if languoid.parent_languoid else ''
            ws.cell(row=row_num, column=10).value = safe_value(languoid.parent_languoid.glottocode) if languoid.parent_languoid else ''
            
            # Family
            ws.cell(row=row_num, column=11).value = languoid.family_languoid.name if languoid.family_languoid else ''
            ws.cell(row=row_num, column=12).value = safe_value(languoid.family_languoid.name_abbrev) if languoid.family_languoid else ''
            ws.cell(row=row_num, column=13).value = safe_value(languoid.family_languoid.glottocode) if languoid.family_languoid else ''
            
            # Primary Subfamily
            ws.cell(row=row_num, column=14).value = languoid.pri_subgroup_languoid.name if languoid.pri_subgroup_languoid else ''
            ws.cell(row=row_num, column=15).value = safe_value(languoid.pri_subgroup_languoid.name_abbrev) if languoid.pri_subgroup_languoid else ''
            ws.cell(row=row_num, column=16).value = safe_value(languoid.pri_subgroup_languoid.glottocode) if languoid.pri_subgroup_languoid else ''
            
            # Secondary Subfamily
            ws.cell(row=row_num, column=17).value = languoid.sec_subgroup_languoid.name if languoid.sec_subgroup_languoid else ''
            ws.cell(row=row_num, column=18).value = safe_value(languoid.sec_subgroup_languoid.name_abbrev) if languoid.sec_subgroup_languoid else ''
            ws.cell(row=row_num, column=19).value = safe_value(languoid.sec_subgroup_languoid.glottocode) if languoid.sec_subgroup_languoid else ''
            
            # Other fields
            ws.cell(row=row_num, column=20).value = safe_value(languoid.alt_names)
            ws.cell(row=row_num, column=21).value = safe_value(languoid.region)
            ws.cell(row=row_num, column=22).value = str(languoid.latitude) if languoid.latitude else ''
            ws.cell(row=row_num, column=23).value = str(languoid.longitude) if languoid.longitude else ''
            ws.cell(row=row_num, column=24).value = safe_value(languoid.tribes)
            ws.cell(row=row_num, column=25).value = safe_value(languoid.notes)
        
        # Auto-size columns
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = 20
        
        # Save to media/exports directory
        exports_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        
        filename = f'languoids_export_{export_id}.xlsx'
        file_path = os.path.join(exports_dir, filename)
        
        # Save workbook
        wb.save(file_path)
        
        logger.info(f"[EXPORT TASK] Successfully generated export: {filename}")
        
        return {
            'status': 'success',
            'filename': filename,
            'count': len(languoids)
        }
        
    except Exception as e:
        logger.error(f"[EXPORT TASK] Error generating export {export_id}: {e}")
        import traceback
        traceback.print_exc()
        raise self.retry(exc=e, countdown=30)


@shared_task(bind=True, max_retries=3, default_retry_delay=30)
def generate_collaborator_export_task(self, export_id, mode, ids):
    """
    Background task to generate collaborator export for large datasets.
    
    Args:
        export_id: Unique identifier for this export (used for filename and tracking)
        mode: 'filtered' | 'selected'
        ids: List of collaborator IDs to export
    
    Returns:
        dict with status, filename, and error info
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from datetime import datetime
    from django.core.files.storage import default_storage
    from django.core.files.base import ContentFile
    from django.conf import settings
    import os
    
    logger.info(f"[EXPORT TASK] Starting collaborator export {export_id}: {len(ids)} collaborators")
    
    try:
        # Fetch collaborators with optimized queries and locale-aware sorting
        from django.db.models.functions import Lower, Collate
        from django.db.models import Count
        
        queryset = Collaborator.objects.filter(id__in=ids).prefetch_related(
            'native_languages',
            'other_languages'
        ).annotate(
            item_count=Count('item_collaborators')
        )
        
        collaborators = list(queryset.order_by(
            Collate(Lower('last_names'), 'en-US-x-icu'),
            Collate(Lower('first_names'), 'en-US-x-icu'),
            'collaborator_id'
        ))
        logger.info(f"[EXPORT TASK] Found {len(collaborators)} collaborators")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Collaborators"
        
        # Define headers (full_name before first_names)
        headers = [
            'Collaborator ID',
            'Full Name',  # Position 2 - export only, read-only
            '# of Items',
            'First and Middle Name(s)',
            'Last Name(s)',
            'Name Suffix',
            'Nickname',
            'Other Names',
            'Anonymous',
            'Native/First Languages',
            'Other Languages',
            'Birth Date',
            'Death Date',
            'Gender',
            'Tribal Affiliations',
            'Clan/Society',
            'Origin',
            'Other Info',
        ]
        
        # Write header row with styling
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Helper function for safe values
        def safe_value(value):
            if value is None:
                return ''
            if isinstance(value, (list, tuple)):
                return ', '.join(str(v) for v in value)
            if isinstance(value, bool):
                return 'Yes' if value else 'No'
            return str(value)
        
        # Write data rows
        for row_num, collab in enumerate(collaborators, 2):
            ws.cell(row=row_num, column=1).value = collab.collaborator_id if collab.collaborator_id else ''
            ws.cell(row=row_num, column=2).value = safe_value(collab.full_name)  # Export only
            ws.cell(row=row_num, column=3).value = collab.item_count  # # of Items
            ws.cell(row=row_num, column=4).value = safe_value(collab.first_names)
            ws.cell(row=row_num, column=5).value = safe_value(collab.last_names)
            ws.cell(row=row_num, column=6).value = safe_value(collab.name_suffix)
            ws.cell(row=row_num, column=7).value = safe_value(collab.nickname)
            ws.cell(row=row_num, column=8).value = safe_value(collab.other_names)
            
            # Anonymous - convert None/True/False to Not specified/Yes/No
            if collab.anonymous is None:
                ws.cell(row=row_num, column=9).value = 'Not specified'
            else:
                ws.cell(row=row_num, column=9).value = 'Yes' if collab.anonymous else 'No'
            
            # Languages - display as comma-separated names
            native_langs = ', '.join([lang.name for lang in collab.native_languages.all()])
            other_langs = ', '.join([lang.name for lang in collab.other_languages.all()])
            ws.cell(row=row_num, column=10).value = native_langs
            ws.cell(row=row_num, column=11).value = other_langs
            
            ws.cell(row=row_num, column=12).value = safe_value(collab.birthdate)
            ws.cell(row=row_num, column=13).value = safe_value(collab.deathdate)
            ws.cell(row=row_num, column=14).value = safe_value(collab.gender)
            ws.cell(row=row_num, column=15).value = safe_value(collab.tribal_affiliations)
            ws.cell(row=row_num, column=16).value = safe_value(collab.clan_society)
            ws.cell(row=row_num, column=17).value = safe_value(collab.origin)
            ws.cell(row=row_num, column=18).value = safe_value(collab.other_info)
        
        # Auto-size columns
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = 20
        
        # Save to media/exports directory
        exports_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        
        filename = f'collaborators_export_{export_id}.xlsx'
        file_path = os.path.join(exports_dir, filename)
        
        # Save workbook
        wb.save(file_path)
        
        logger.info(f"[EXPORT TASK] Successfully generated collaborator export: {filename}")
        
        return {
            'status': 'success',
            'filename': filename,
            'count': len(collaborators)
        }
        
    except Exception as e:
        logger.error(f"[EXPORT TASK] Error generating collaborator export {export_id}: {e}")
        import traceback
        traceback.print_exc()
        raise self.retry(exc=e, countdown=30)