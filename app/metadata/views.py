import os, csv, io, datetime, re, mutagen, librosa, json, zipfile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.writer.excel import save_virtual_workbook
from django.core.files.storage import default_storage
from django.core.paginator import Paginator
from django.conf import settings
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib.admin.utils import flatten
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import Http404, HttpResponse, StreamingHttpResponse
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.forms.models import model_to_dict
from django.db.models import Count, Sum, Max, Q
from django.views.generic.edit import FormView, DeleteView
from rest_framework import generics
from django.contrib.auth.mixins import UserPassesTestMixin, LoginRequiredMixin
from .models import Item, ItemTitle, Language, Dialect, DialectInstance, Collaborator, CollaboratorRole, Geographic, Columns_export, Document, Video, ACCESS_CHOICES, ACCESSION_CHOICES, AVAILABILITY_CHOICES, CONDITION_CHOICES, CONTENT_CHOICES, FORMAT_CHOICES, GENRE_CHOICES, STRICT_GENRE_CHOICES, MONTH_CHOICES, ROLE_CHOICES, LANGUAGE_DESCRIPTION_CHOICES, reverse_lookup_choices, validate_date_text
from .serializers import ItemMigrateSerializer
from .forms import LanguageForm, DialectForm, DialectInstanceForm, DialectInstanceCustomForm, CollaboratorForm, CollaboratorRoleForm, GeographicForm, ItemForm, Columns_exportForm, Columns_export_choiceForm, Csv_format_type, DocumentForm, VideoForm, UploadDocumentForm

def is_member_of_archivist(user):
    return user.groups.filter(name="Archivist").exists()

def custom_error_500(request):
    return render(request, '500.html', {}, status=500)

def trigger_error(request):
    division_by_zero = 1 / 0

class ItemUpdateMigrateView(LoginRequiredMixin, UserPassesTestMixin, generics.UpdateAPIView):
    queryset = Item.objects.all()
    serializer_class = ItemMigrateSerializer

    def test_func(self):
        # return is_member_of_archivist(self.request.user)
        return self.request.user.username == 'kavon'

@login_required
def item_index(request):
    url_path = request.get_full_path()
    qs = Item.objects.all()
    order_choice = request.GET.get("form_control_sort")
    columns_choice_name = request.GET.get("form_control_columns")
    catalog_number_contains_query = request.GET.get('catalog_number_contains')
    item_access_level_contains_query = request.GET.get('item_access_level_contains')
    call_number_contains_query = request.GET.get('call_number_contains')
    accession_date_min_query = request.GET.get('accession_date_min')
    accession_date_max_query = request.GET.get('accession_date_max')
    indigenous_title_contains_query = request.GET.get('indigenous_title_contains')
    english_title_contains_query = request.GET.get('english_title_contains')
    titles_contains_query = request.GET.get('titles_contains')
    general_content_contains_query = request.GET.get('general_content_contains')
    language_contains_query = request.GET.get('language_contains')
    creation_date_min_query = request.GET.get('creation_date_min')
    creation_date_max_query = request.GET.get('creation_date_max')
    description_scope_and_content_contains_query = request.GET.get('description_scope_and_content_contains')
    genre_contains_query = request.GET.get('genre_contains')
    collaborator_contains_query = request.GET.get('collaborator_contains')
    depositor_name_contains_query = request.GET.get('depositor_name_contains')
    keyword_contains_query = request.GET.get('keyword_contains')

    order_choice_last = order_choice
    columns_choice_name_last = columns_choice_name
    catalog_number_contains_query_last = ''
    item_access_level_contains_query_last = ''
    call_number_contains_query_last = ''
    accession_date_min_query_last = ''
    accession_date_max_query_last = ''
    indigenous_title_contains_query_last = ''
    english_title_contains_query_last = ''
    titles_contains_query_last = ''
    general_content_contains_query_last = ''
    language_contains_query_last = ''
    creation_date_min_query_last = ''
    creation_date_max_query_last = ''
    description_scope_and_content_contains_query_last = ''
    genre_contains_query_last = ''
    collaborator_contains_query_last = ''
    depositor_name_contains_query_last = ''
    keyword_contains_query_last = ''



    if is_valid_param(catalog_number_contains_query):
        qs = qs.filter(catalog_number__icontains = catalog_number_contains_query)
        catalog_number_contains_query_last = catalog_number_contains_query

    if is_valid_param(item_access_level_contains_query):
        qs = qs.filter(item_access_level__icontains = item_access_level_contains_query)
        item_access_level_contains_query_last = item_access_level_contains_query

    if is_valid_param(call_number_contains_query):
        qs = qs.filter(call_number__icontains = call_number_contains_query)
        call_number_contains_query_last = call_number_contains_query

    if is_valid_param(accession_date_min_query):
        qs = qs.filter(accession_date__gte = accession_date_min_query)
        accession_date_min_query_last = accession_date_min_query

    if is_valid_param(accession_date_max_query):
        qs = qs.filter(accession_date__lte = accession_date_max_query)
        accession_date_max_query_last = accession_date_max_query

    if is_valid_param(indigenous_title_contains_query):
        qs = qs.filter(indigenous_title__icontains = indigenous_title_contains_query)
        indigenous_title_contains_query_last = indigenous_title_contains_query

    if is_valid_param(english_title_contains_query):
        qs = qs.filter(english_title__icontains = english_title_contains_query)
        english_title_contains_query_last = english_title_contains_query

    if is_valid_param(general_content_contains_query):
        qs = qs.filter(general_content__icontains = general_content_contains_query)
        general_content_contains_query_last = general_content_contains_query

    if is_valid_param(language_contains_query):
        qs = qs.filter(language__name__icontains = language_contains_query)
        language_contains_query_last = language_contains_query

    if is_valid_param(creation_date_min_query):
        qs = qs.filter(creation_date__gte = creation_date_min_query)
        creation_date_min_query_last = creation_date_min_query

    if is_valid_param(creation_date_max_query):
        qs = qs.filter(creation_date__lte = creation_date_max_query)
        creation_date_max_query_last = creation_date_max_query

    if is_valid_param(description_scope_and_content_contains_query):
        qs = qs.filter(description_scope_and_content__icontains = description_scope_and_content_contains_query)
        description_scope_and_content_contains_query_last = description_scope_and_content_contains_query

    if is_valid_param(genre_contains_query):
        qs = qs.filter(genre__icontains = genre_contains_query)
        genre_contains_query_last = genre_contains_query

    if is_valid_param(depositor_name_contains_query):
        qs = qs.filter(depositor_name__icontains = depositor_name_contains_query)
        depositor_name_contains_query_last = depositor_name_contains_query

    if is_valid_param(collaborator_contains_query):
        if ( collaborator_contains_query == 'Anonymous' ) or ( collaborator_contains_query == 'anonymous' ):
            anonymous_list = qs.filter(
                Q(collaborator__anonymous = True) | Q(item_documents__collaborator__anonymous = True)
                )
        else:
            anonymous_list = Item.objects.none()

        if not is_member_of_archivist(request.user):
            qs = qs.exclude(collaborator__anonymous = True).exclude(item_documents__collaborator__anonymous = True)
        collaborator_name_qs = qs.filter(
            Q(collaborator__name__icontains = collaborator_contains_query) | Q(item_documents__collaborator__name__icontains = collaborator_contains_query)
            )
        qs = collaborator_name_qs

        qs = qs.union(anonymous_list)
        collaborator_contains_query_last = collaborator_contains_query


    qs_simple = qs

    if is_valid_param(titles_contains_query):
        indigenous_title_condition = Q(indigenous_title__icontains = titles_contains_query)
        english_title_condition = Q(english_title__icontains=titles_contains_query)
        combined_condition = indigenous_title_condition | english_title_condition
        partial_qs_title = qs_simple.filter(combined_condition)
        qs = qs.intersection(partial_qs_title)
        titles_contains_query_last = titles_contains_query


    if is_valid_param(keyword_contains_query):

        if ( keyword_contains_query == 'Anonymous' ) or ( keyword_contains_query == 'anonymous' ):
            partial_qs_collaborator_anonymous = qs_simple.filter(collaborator__anonymous = True)
            partial_qs_document_collaborator_anonymous = qs_simple.filter(item_documents__collaborator__anonymous = True)

        else:
            partial_qs_collaborator_anonymous = Item.objects.none()
            partial_qs_document_collaborator_anonymous = Item.objects.none()

        partial_qs_collaborator_birthdate = qs_simple.filter(collaborator__birthdate__icontains = keyword_contains_query)
        partial_qs_collaborator_clan_society = qs_simple.filter(collaborator__clan_society__icontains = keyword_contains_query)
        partial_qs_collaborator_deathdate = qs_simple.filter(collaborator__deathdate__icontains = keyword_contains_query)
        partial_qs_collaborator_name = qs_simple.filter(collaborator__name__icontains = keyword_contains_query)
        partial_qs_collaborator_nickname = qs_simple.filter(collaborator__nickname__icontains = keyword_contains_query)
        partial_qs_collaborator_origin = qs_simple.filter(collaborator__origin__icontains = keyword_contains_query)
        partial_qs_collaborator_other_info = qs_simple.filter(collaborator__other_info__icontains = keyword_contains_query)
        partial_qs_collaborator_other_names = qs_simple.filter(collaborator__other_names__icontains = keyword_contains_query)
        partial_qs_collaborator_tribal_affiliations = qs_simple.filter(collaborator__tribal_affiliations__icontains = keyword_contains_query)
        partial_qs_document_collaborator = qs_simple.filter(item_documents__collaborator__name__icontains = keyword_contains_query)
        if not is_member_of_archivist(request.user):
            partial_qs_collaborator_birthdate = partial_qs_collaborator_birthdate.exclude(collaborator__anonymous = True)
            partial_qs_collaborator_clan_society = partial_qs_collaborator_clan_society.exclude(collaborator__anonymous = True)
            partial_qs_collaborator_deathdate = partial_qs_collaborator_deathdate.exclude(collaborator__anonymous = True)
            partial_qs_collaborator_name = partial_qs_collaborator_name.exclude(collaborator__anonymous = True)
            partial_qs_collaborator_nickname = partial_qs_collaborator_nickname.exclude(collaborator__anonymous = True)
            partial_qs_collaborator_origin = partial_qs_collaborator_origin.exclude(collaborator__anonymous = True)
            partial_qs_collaborator_other_info = partial_qs_collaborator_other_info.exclude(collaborator__anonymous = True)
            partial_qs_collaborator_other_names = partial_qs_collaborator_other_names.exclude(collaborator__anonymous = True)
            partial_qs_collaborator_tribal_affiliations = partial_qs_collaborator_tribal_affiliations.exclude(collaborator__anonymous = True)
            partial_qs_document_collaborator = partial_qs_document_collaborator.exclude(item_documents__collaborator__anonymous = True)

        partial_qs_collaborator_native_languages = qs_simple.filter(collaborator__native_languages__name__icontains = keyword_contains_query)
        partial_qs_collaborator_other_languages = qs_simple.filter(collaborator__other_languages__name__icontains = keyword_contains_query)
        partial_qs_access_level_restrictions = qs_simple.filter(access_level_restrictions__icontains = keyword_contains_query)
        partial_qs_accession_date = qs_simple.filter(accession_date__icontains = keyword_contains_query)
        partial_qs_accession_number = qs_simple.filter(accession_number__icontains = keyword_contains_query)
        partial_qs_acquisition_notes = qs_simple.filter(acquisition_notes__icontains = keyword_contains_query)
        partial_qs_additional_digital_file_location = qs_simple.filter(additional_digital_file_location__icontains = keyword_contains_query)
        partial_qs_associated_ephemera = qs_simple.filter(associated_ephemera__icontains = keyword_contains_query)
        partial_qs_availability_status = qs_simple.filter(availability_status__icontains = keyword_contains_query)
        partial_qs_availability_status_notes = qs_simple.filter(availability_status_notes__icontains = keyword_contains_query)
        partial_qs_call_number = qs_simple.filter(call_number__icontains = keyword_contains_query)
        partial_qs_catalog_number = qs_simple.filter(catalog_number__icontains = keyword_contains_query)
        partial_qs_cataloged_by = qs_simple.filter(cataloged_by__icontains = keyword_contains_query)
        partial_qs_cataloged_date = qs_simple.filter(cataloged_date__icontains = keyword_contains_query)
        partial_qs_collaborator_role = qs_simple.filter(item_collaboratorroles__role__icontains = keyword_contains_query)
        partial_qs_collecting_notes = qs_simple.filter(collecting_notes__icontains = keyword_contains_query)
        partial_qs_collection_date_min = qs_simple.filter(collection_date_min__icontains = keyword_contains_query)
        partial_qs_collection_name = qs_simple.filter(collection_name__icontains = keyword_contains_query)
        partial_qs_collector_info = qs_simple.filter(collector_info__icontains = keyword_contains_query)
        partial_qs_collector_name = qs_simple.filter(collector_name__icontains = keyword_contains_query)
        partial_qs_condition = qs_simple.filter(condition__icontains = keyword_contains_query)
        partial_qs_condition_notes = qs_simple.filter(condition_notes__icontains = keyword_contains_query)
        partial_qs_conservation_recommendation = qs_simple.filter(conservation_recommendation__icontains = keyword_contains_query)
        partial_qs_conservation_treatments_performed = qs_simple.filter(conservation_treatments_performed__icontains = keyword_contains_query)
        partial_qs_copyrighted_notes = qs_simple.filter(copyrighted_notes__icontains = keyword_contains_query)
        partial_qs_country_or_territory = qs_simple.filter(country_or_territory__icontains = keyword_contains_query)
        partial_qs_county_or_parish = qs_simple.filter(county_or_parish__icontains = keyword_contains_query)
        partial_qs_creation_date = qs_simple.filter(creation_date__icontains = keyword_contains_query)
        partial_qs_deposit_date = qs_simple.filter(deposit_date__icontains = keyword_contains_query)
        partial_qs_depositor_contact_information = qs_simple.filter(depositor_contact_information__icontains = keyword_contains_query)
        partial_qs_depositor_name = qs_simple.filter(depositor_name__icontains = keyword_contains_query)
        partial_qs_description_scope_and_content = qs_simple.filter(description_scope_and_content__icontains = keyword_contains_query)
        partial_qs_dialect = qs_simple.filter(item_dialectinstances__name__name__icontains = keyword_contains_query)
        partial_qs_digital_file_location = qs_simple.filter(digital_file_location__icontains = keyword_contains_query)
        partial_qs_document_filename = qs_simple.filter(item_documents__filename__icontains = keyword_contains_query)
        partial_qs_document_filetype = qs_simple.filter(item_documents__filetype__icontains = keyword_contains_query)
        partial_qs_document_enumerator = qs_simple.filter(item_documents__enumerator__icontains = keyword_contains_query)
        partial_qs_document_title = qs_simple.filter(item_documents__title__icontains = keyword_contains_query)
        partial_qs_document_duration = qs_simple.filter(item_documents__duration__icontains = keyword_contains_query)
        partial_qs_document_filesize = qs_simple.filter(item_documents__filesize__icontains = keyword_contains_query)
        partial_qs_document_av_spec = qs_simple.filter(item_documents__av_spec__icontains = keyword_contains_query)
        partial_qs_document_language = qs_simple.filter(item_documents__language__name__icontains = keyword_contains_query)
        partial_qs_english_title = qs_simple.filter(english_title__icontains = keyword_contains_query)
        partial_qs_equipment_used = qs_simple.filter(equipment_used__icontains = keyword_contains_query)
        partial_qs_filemaker_legacy_pk_id = qs_simple.filter(filemaker_legacy_pk_id__icontains = keyword_contains_query)
        partial_qs_general_content = qs_simple.filter(general_content__icontains = keyword_contains_query)
        partial_qs_genre = qs_simple.filter(genre__icontains = keyword_contains_query)
        partial_qs_global_region = qs_simple.filter(global_region__icontains = keyword_contains_query)
        partial_qs_indigenous_title = qs_simple.filter(indigenous_title__icontains = keyword_contains_query)
        partial_qs_ipm_issues = qs_simple.filter(ipm_issues__icontains = keyword_contains_query)
        partial_qs_isbn = qs_simple.filter(isbn__icontains = keyword_contains_query)
        partial_qs_item_access_level = qs_simple.filter(item_access_level__icontains = keyword_contains_query)
        partial_qs_language_name = qs_simple.filter(language__name__icontains = keyword_contains_query)
        partial_qs_language_iso = qs_simple.filter(language__iso__icontains = keyword_contains_query)
        partial_qs_language_family = qs_simple.filter(language__family__icontains = keyword_contains_query)
        partial_qs_language_pri_subgroup = qs_simple.filter(language__pri_subgroup__icontains = keyword_contains_query)
        partial_qs_language_sec_subgroup = qs_simple.filter(language__sec_subgroup__icontains = keyword_contains_query)
        partial_qs_language_alt_name = qs_simple.filter(language__alt_name__icontains = keyword_contains_query)
        partial_qs_language_region = qs_simple.filter(language__region__icontains = keyword_contains_query)
        partial_qs_language_notes = qs_simple.filter(language__notes__icontains = keyword_contains_query)
        partial_qs_lender_loan_number = qs_simple.filter(lender_loan_number__icontains = keyword_contains_query)
        partial_qs_loc_catalog_number = qs_simple.filter(loc_catalog_number__icontains = keyword_contains_query)
        partial_qs_location_of_original = qs_simple.filter(location_of_original__icontains = keyword_contains_query)
        partial_qs_migration_file_format = qs_simple.filter(migration_file_format__icontains = keyword_contains_query)
        partial_qs_migration_location = qs_simple.filter(migration_location__icontains = keyword_contains_query)
        partial_qs_municipality_or_township = qs_simple.filter(municipality_or_township__icontains = keyword_contains_query)
        partial_qs_original_format_medium = qs_simple.filter(original_format_medium__icontains = keyword_contains_query)
        partial_qs_other_information = qs_simple.filter(other_information__icontains = keyword_contains_query)
        partial_qs_other_institutional_number = qs_simple.filter(other_institutional_number__icontains = keyword_contains_query)
        partial_qs_permission_to_publish_online = qs_simple.filter(permission_to_publish_online__icontains = keyword_contains_query)
        partial_qs_project_grant = qs_simple.filter(project_grant__icontains = keyword_contains_query)
        partial_qs_public_event = qs_simple.filter(public_event__icontains = keyword_contains_query)
        partial_qs_recorded_on = qs_simple.filter(recorded_on__icontains = keyword_contains_query)
        partial_qs_recording_context = qs_simple.filter(recording_context__icontains = keyword_contains_query)
        partial_qs_software_used = qs_simple.filter(software_used__icontains = keyword_contains_query)
        partial_qs_state_or_province = qs_simple.filter(state_or_province__icontains = keyword_contains_query)
        partial_qs_temporary_accession_number = qs_simple.filter(temporary_accession_number__icontains = keyword_contains_query)
        partial_qs_total_number_of_pages_and_physical_description = qs_simple.filter(total_number_of_pages_and_physical_description__icontains = keyword_contains_query)
        partial_qs_type_of_accession = qs_simple.filter(type_of_accession__icontains = keyword_contains_query)
        partial_qs_modified_by = qs_simple.filter(modified_by__icontains = keyword_contains_query)

        partial_qs_keyword = partial_qs_collaborator_anonymous.union(
                                    partial_qs_document_collaborator_anonymous,
                                    partial_qs_collaborator_name,
                                    partial_qs_collaborator_other_names,
                                    partial_qs_collaborator_nickname,
                                    partial_qs_collaborator_clan_society,
                                    partial_qs_collaborator_tribal_affiliations,
                                    partial_qs_collaborator_native_languages,
                                    partial_qs_collaborator_other_languages,
                                    partial_qs_access_level_restrictions,
                                    partial_qs_accession_date,
                                    partial_qs_accession_number,
                                    partial_qs_acquisition_notes,
                                    partial_qs_additional_digital_file_location,
                                    partial_qs_associated_ephemera,
                                    partial_qs_availability_status,
                                    partial_qs_availability_status_notes,
                                    partial_qs_call_number,
                                    partial_qs_catalog_number,
                                    partial_qs_cataloged_by,
                                    partial_qs_cataloged_date,
                                    partial_qs_collaborator_role,
                                    partial_qs_collecting_notes,
                                    partial_qs_collection_date_min,
                                    partial_qs_collection_name,
                                    partial_qs_collector_info,
                                    partial_qs_collector_name,
                                    partial_qs_condition,
                                    partial_qs_condition_notes,
                                    partial_qs_conservation_recommendation,
                                    partial_qs_conservation_treatments_performed,
                                    partial_qs_copyrighted_notes,
                                    partial_qs_country_or_territory,
                                    partial_qs_county_or_parish,
                                    partial_qs_creation_date,
                                    partial_qs_deposit_date,
                                    partial_qs_depositor_contact_information,
                                    partial_qs_depositor_name,
                                    partial_qs_description_scope_and_content,
                                    partial_qs_dialect,
                                    partial_qs_digital_file_location,
                                    partial_qs_document_filename,
                                    partial_qs_document_filetype,
                                    partial_qs_document_enumerator,
                                    partial_qs_document_title,
                                    partial_qs_document_duration,
                                    partial_qs_document_filesize,
                                    partial_qs_document_av_spec,
                                    partial_qs_document_language,
                                    partial_qs_document_collaborator,
                                    partial_qs_english_title,
                                    partial_qs_equipment_used,
                                    partial_qs_filemaker_legacy_pk_id,
                                    partial_qs_general_content,
                                    partial_qs_genre,
                                    partial_qs_global_region,
                                    partial_qs_indigenous_title,
                                    partial_qs_ipm_issues,
                                    partial_qs_isbn,
                                    partial_qs_item_access_level,
                                    partial_qs_language_name,
                                    partial_qs_language_iso,
                                    partial_qs_language_family,
                                    partial_qs_language_pri_subgroup,
                                    partial_qs_language_sec_subgroup,
                                    partial_qs_language_alt_name,
                                    partial_qs_language_region,
                                    partial_qs_language_notes,
                                    partial_qs_lender_loan_number,
                                    partial_qs_loc_catalog_number,
                                    partial_qs_location_of_original,
                                    partial_qs_migration_file_format,
                                    partial_qs_migration_location,
                                    partial_qs_municipality_or_township,
                                    partial_qs_original_format_medium,
                                    partial_qs_other_information,
                                    partial_qs_other_institutional_number,
                                    partial_qs_permission_to_publish_online,
                                    partial_qs_project_grant,
                                    partial_qs_public_event,
                                    partial_qs_recorded_on,
                                    partial_qs_recording_context,
                                    partial_qs_software_used,
                                    partial_qs_state_or_province,
                                    partial_qs_temporary_accession_number,
                                    partial_qs_total_number_of_pages_and_physical_description,
                                    partial_qs_type_of_accession,
                                    partial_qs_modified_by,
                                    )
        qs = qs.intersection(partial_qs_keyword)
#        qs = Item.objects.get(pk__in=qs.pk)
        keyword_contains_query_last = keyword_contains_query

#    qs = qs.order_by('-pub_date', 'headline')
#    print(type(order_choice))
    if order_choice == "updated":
        qs = qs.order_by('-updated')
    else:
        qs = qs.order_by('catalog_number')

    results_count = qs.count()

    documents_in_qs = Document.objects.filter(item__catalog_number__in=list(qs.values_list('catalog_number', flat=True)))
    results_duration = documents_in_qs.aggregate(sum_duration=Sum('duration')).get('sum_duration')
    if results_duration:
        results_duration_datetime = str(datetime.timedelta(seconds=results_duration))[:-3]
    else:
        results_duration_datetime = 0

    if 'export' in request.GET:

        if re.search('migrate', url_path, flags=re.I):
            # export json files for each item marked for migration, plus a combined json file, as a zip file
            items_to_migrate = Item.objects.filter(migrate=True)
            # items_to_migrate = items_to_migrate.order_by('catalog_number')
            # items_to_migrate = items_to_migrate.prefetch_related('language', 'collaborator', 'item_documents', 'item_documents__language', 'item_documents__collaborator')
            # items_to_migrate = items_to_migrate.prefetch_related('item_dialectinstances', 'item_dialectinstances__name')
            # # items_to_migrate = items_to_migrate.prefetch_related('item_collaboratorroles', 'item_collaboratorroles__role')
            # items_to_migrate = items_to_migrate.prefetch_related('item_geographic')

            # Path to the zip file in the media folder
            zip_path = os.path.join(settings.MEDIA_ROOT, 'items.zip')
            # Directory path
            json_dir_path = os.path.join(settings.MEDIA_ROOT, 'migrate-json')
            # Create the directory if it doesn't exist
            os.makedirs(json_dir_path, exist_ok=True)

            # for each item, create a json file
            with zipfile.ZipFile(zip_path, 'w') as zipf:

                
                for item in items_to_migrate:
                    item_dict = model_to_dict(item)
                    # item_dict['language'] = [model_to_dict(language) for language in item.language.all()]
                    item_dict['language'] = list(item.language.values_list('id', flat=True))
                    # # item_dict['dialect'] = [model_to_dict(dialect) for dialect in item.dialect.all()]
                    # item_dict['collaborator'] = [model_to_dict(collaborator) for collaborator in item.collaborator.all()]
                    item_dict['collaborator'] = list(item.collaborator.values_list('id', flat=True))
                    # item_dict['item_documents'] = [model_to_dict(document) for document in item.item_documents.all()]
                    # item_dict['item_dialectinstances'] = [model_to_dict(dialect) for dialect in item.item_dialectinstances.all()]
                    # item_dict['item_collaboratorroles'] = [model_to_dict(role) for role in item.item_collaboratorroles.all()]
                    # item_dict['item_geographic'] = [model_to_dict(geographic) for geographic in item.item_geographic.all()]
                    # item_dict['item_geographic__parent'] = [model_to_dict(geographic) for geographic in item.item_geographic.all()]

                    # Split the date string by "/"
                    deposit_date_parts = item.deposit_date.split("/")
                    # Reverse the order of the elements
                    deposit_date_parts.reverse()
                    # Add a leading zero to elements with only one digit
                    deposit_date_parts = [part.zfill(2) for part in deposit_date_parts]
                    # Join the elements back together into a string
                    item_dict['deposit_date_formatted'] = "-".join(deposit_date_parts)

                    # construct additional titles
                    item_titles = ItemTitle.objects.filter(item=item)
                    additional_titles = []
                    for title in item_titles:
                        additional_title = {
                            "title": title.title,
                            "type": {"id": "translated-title"},
                            "lang": {"id": title.language.glottocode}
                        }
                        additional_titles.append(additional_title)

                    # Construct metadata
                    metadata = {
                        "resource_type": {
                            "id": item_dict['general_content'].replace('_', '-').replace('audio-video', 'video'),
                        },
                        "creators": [
                            {"person_or_org": {
                                "type": "personal",
                                "name": "na, na",
                                "given_name": "na",
                                "family_name": "na"
                            }}
                        ],
                        "title": item.english_title,
                        "additional_titles": additional_titles,
                        "publication_date": item_dict['deposit_date_formatted'],
                        "description": item.description_scope_and_content
                    }

                    # Construct custom fields
                    custom_fields = {
                        "archive_item:item": item.catalog_number,
                        "archive_item:call_number": item.call_number,
                        "archive_item:access_level": item.get_item_access_level_display(),
                        "archive_item:all_languages": [
                            {"id": each_language.glottocode}
                            for each_language in item.language.all()
                        ],
                        "archive_item:genre": [
                            {"id": each_genre.replace('_', '-'),}
                            for each_genre in item.genre
                            if each_genre not in ["book", "article", "dataset", "document", "educational", "photograph", "thesis"]

                        ],
                        # "archive_item:genre": [
                        #     {"id": each_genre.replace('_', '-'),}
                        #     for each_genre in item.genre
                        #     if each_genre not in ["book", "article", "dataset", "document", "educational", "photograph", "thesis"]
                        # ],
                        "archive_item:lang_desc_type": [
                            {"id": each_language_description_type.replace('_', '-')}
                            for each_language_description_type in item.language_description_type
                        ],
                        "archive_item:associated_ephemera": item.associated_ephemera,
                        "archive_item:access_level_restrictions": item.access_level_restrictions,
                        "archive_item:copyrighted_notes": item.copyrighted_notes,
                        "archive_item:availability_status": {
                            "id": item.availability_status.replace('_', '-'),
                        },
                        "archive_item:availability_status_notes": item.availability_status_notes,

                        "archive_item:collecting_notes": item.collecting_notes,
                        "archive_item:global_region": item.global_region,
                        "archive_item:recording_context": item.recording_context,
                        "archive_item:public_event": True if item.public_event == "Yes" else False,
                        "archive_item:original_format_medium": {
                            "id": item.original_format_medium.replace('_', '-'),
                        },
                        "archive_item:recorded_on": item.recorded_on,
                        "archive_item:equipment_used": item.equipment_used,
                        "archive_item:software_used": item.software_used,
                        "archive_item:location_of_original": item.location_of_original,
                        "archive_item:other_information": item.other_information,
                        "archive_item:publisher": item.publisher,
                        "archive_item:isbn": item.isbn,
                        "archive_item:loc_catalog_number": item.loc_catalog_number,
                        "archive_item:lender_loan_number": item.lender_loan_number,
                        "archive_item:total_number_of_pages_and_physical_description": item.total_number_of_pages_and_physical_description,
                    }

                    # Construct access control details
                    access = {
                        "record": "public",
                        "files": "public",
                        "embargo": {
                            "active": False,
                            "reason": None
                        },
                        "status": "open"
                    }

                    item_output = {
                        "metadata": metadata,
                        "custom_fields": custom_fields,
                        "access": access
                    }

                    item_json = json.dumps(item_output, indent=4)
                    item_filename = item.catalog_number + '.json'
                    item_path = os.path.join(json_dir_path, item_filename)

                    # Write the JSON data to a file
                    with open(item_path, 'w') as f:
                        f.write(item_json)

                    # Add the file to the zip file
                    zipf.write(item_path, arcname=item_filename)


            # Create a generator that reads the file and yields the data
            def file_iterator():
                try:
                    with open(zip_path, 'rb') as f:
                        for chunk in iter(lambda: f.read(4096), b''):
                            yield chunk
                finally:
                    os.remove(zip_path)

            # Create a StreamingHttpResponse that uses the file_iterator
            response = StreamingHttpResponse(file_iterator(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{zip_path}"'

            # # Delete the files from the default storage
            # default_storage.delete(json_file)
            return response


        else:
            # for now this only exports the columns in the main table, not including related tables
            if is_valid_param(columns_choice_name):
                column_choice = Columns_export.objects.get(name=columns_choice_name)

                items_in_qs = Item.objects.filter(catalog_number__in=list(qs.values_list('catalog_number', flat=True)))

                # Determine how many language columns are needed
                language_list = items_in_qs.annotate(key_count=Count('language__name'))
                language_counts = []
                for entry in language_list:
                    language_counts.append(entry.key_count)
                max_language_counts = max(language_counts)
                if max_language_counts < 1:
                    max_language_counts = 1

                # Determine how many collaborator columns are needed
                collaborator_list = items_in_qs.annotate(key_count=Count('collaborator__name'))
                collaborator_counts = []
                for entry in collaborator_list:
                    collaborator_counts.append(entry.key_count)
                max_collaborator_counts = max(collaborator_counts)
                if max_collaborator_counts < 1:
                    max_collaborator_counts = 1

                # Determine how many document columns are needed
                document_list = items_in_qs.annotate(key_count=Count('item_documents'))
                document_counts = []
                for entry in document_list:
                    document_counts.append(entry.key_count)
                max_document_counts = max(document_counts)
                if max_document_counts < 1:
                    max_document_counts = 1

                # Determine how many document language columns are needed
                document_language_list = Document.objects.filter(item__in=items_in_qs).annotate(key_count=Count('language__name'))
                document_language_counts = []
                for entry in document_language_list:
                    document_language_counts.append(entry.key_count)
                if is_valid_param(document_language_counts):
                    max_document_language_counts = max(document_language_counts)
                    if max_document_language_counts < 1:
                        max_document_language_counts = 1
                else:
                    max_document_language_counts = 1

                # Determine how many document language columns are needed
                document_collaborator_list = Document.objects.filter(item__in=items_in_qs).annotate(key_count=Count('collaborator__name'))
                document_collaborator_counts = []
                for entry in document_collaborator_list:
                    document_collaborator_counts.append(entry.key_count)
                if is_valid_param(document_collaborator_counts):
                    max_document_collaborator_counts = max(document_collaborator_counts)
                    if max_document_collaborator_counts < 1:
                        max_document_collaborator_counts = 1
                else:
                    max_document_collaborator_counts = 1

        #        # Create the HttpResponse object with the appropriate CSV header.
        #        response = HttpResponse(content_type='text/csv')
        #        response['Content-Disposition'] = 'attachment; filename="csv_filtered_write.csv"'
        #
        #        writer = csv.writer(response)



                style_general = PatternFill(start_color='00FFFF66',
                                            end_color='00FFFF66',
                                            fill_type='solid')
                style_document_odd = PatternFill(start_color='0080FF80',
                                                end_color='0080FF80',
                                                fill_type='solid')

                style_document_even = PatternFill(start_color='0079D279',
                                                end_color='0079D279',
                                                fill_type='solid')

                style_access = PatternFill(start_color='0099DDFF',
                                        end_color='0099DDFF',
                                        fill_type='solid')

                style_collaborator_odd = PatternFill(start_color='009999FF',
                                                    end_color='009999FF',
                                                    fill_type='solid')

                style_collaborator_even = PatternFill(start_color='006666FF',
                                                    end_color='006666FF',
                                                    fill_type='solid')

                style_browse = PatternFill(start_color='BCAED5',
                                            end_color='BCAED5',
                                            fill_type='solid')

                style_condition = PatternFill(start_color='00CC80FF',
                                            end_color='00CC80FF',
                                            fill_type='solid')

                style_accessions = PatternFill(start_color='00FF99FF',
                                            end_color='00FF99FF',
                                            fill_type='solid')

                style_location = PatternFill(start_color='00FF9999',
                                            end_color='00FF9999',
                                            fill_type='solid')

                style_digitization = PatternFill(start_color='00FFB380',
                                                end_color='00FFB380',
                                                fill_type='solid')

                style_books = PatternFill(start_color='00CCFF66',
                                        end_color='00CCFF66',
                                        fill_type='solid')

                style_external = PatternFill(start_color='00B3FFEC',
                                            end_color='00B3FFEC',
                                            fill_type='solid')

                style_deprecated = PatternFill(start_color='00CCCCCC',
                                            end_color='00CCCCCC',
                                            fill_type='solid')

                new_workbook = Workbook()
                sheet = new_workbook.active

                sheet_column_counter = 1
                header_cell = sheet.cell(row=1, column=sheet_column_counter )

                if column_choice.item_catalog_number:
                    header_cell.value = 'Catalog number'
                    header_cell.fill = style_general
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_item_access_level:
                    header_cell.value = 'Item access level'
                    header_cell.fill = style_general
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_call_number:
                    header_cell.value = 'Call number'
                    header_cell.fill = style_general
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_accession_date:
                    header_cell.value = 'Accession date'
                    header_cell.fill = style_general
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_additional_digital_file_location:
                    header_cell.value = 'Additional digital file location'
                    header_cell.fill = style_general
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_indigenous_title:
                    header_cell.value = 'Indigenous title'
                    header_cell.fill = style_general
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_english_title:
                    header_cell.value = 'English title'
                    header_cell.fill = style_general
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_general_content:
                    header_cell.value = 'General content'
                    header_cell.fill = style_general
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_language:
                    column_counter = 1
                    while column_counter <= max_language_counts:
                        header_cell.value = 'Language name %s' %column_counter
                        header_cell.fill = style_general
                        sheet_column_counter += 1
                        header_cell = sheet.cell(row=1, column=sheet_column_counter )
                        if column_choice.item_dialect:
                            header_cell.value = 'Dialect %s' %column_counter
                            header_cell.fill = style_general
                            sheet_column_counter += 1
                            header_cell = sheet.cell(row=1, column=sheet_column_counter )
                        column_counter += 1
                if column_choice.item_creation_date:
                    header_cell.value = 'Creation date'
                    header_cell.fill = style_general
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_description_scope_and_content:
                    header_cell.value = 'Description scope and content'
                    header_cell.fill = style_general
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_genre:
                    header_cell.value = 'Genre'
                    header_cell.fill = style_general
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_associated_ephemera:
                    header_cell.value = 'Associated ephemera'
                    header_cell.fill = style_general
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )

                if column_choice.item_access_level_restrictions:
                    header_cell.value = 'Access level restrictions'
                    header_cell.fill = style_access
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_copyrighted_notes:
                    header_cell.value = 'Copyrighted notes'
                    header_cell.fill = style_access
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_permission_to_publish_online:
                    header_cell.value = 'Permission to publish online'
                    header_cell.fill = style_access
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_collaborator:
                    column_counter = 1
                    while column_counter <= max_collaborator_counts:
                        header_cell.value = 'Collaborator name %s' %column_counter
                        if column_counter % 2 == 0: #even
                            header_cell.fill = style_collaborator_even
                        else:
                            header_cell.fill = style_collaborator_odd
                        sheet_column_counter += 1
                        header_cell = sheet.cell(row=1, column=sheet_column_counter )
                        if column_choice.item_collaborator_role:
                            header_cell.value = 'Collaborator role %s' %column_counter
                            if column_counter % 2 == 0: #even
                                header_cell.fill = style_collaborator_even
                            else:
                                header_cell.fill = style_collaborator_odd
                            sheet_column_counter += 1
                            header_cell = sheet.cell(row=1, column=sheet_column_counter )
                        column_counter += 1
                if column_choice.item_language_description_type:
                    header_cell.value = 'Language description type'
                    header_cell.fill = style_browse
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_availability_status:
                    header_cell.value = 'Availability status'
                    header_cell.fill = style_condition
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_availability_status_notes:
                    header_cell.value = 'Availability status notes'
                    header_cell.fill = style_condition
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_condition:
                    header_cell.value = 'Condition'
                    header_cell.fill = style_condition
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_condition_notes:
                    header_cell.value = 'Condition notes'
                    header_cell.fill = style_condition
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_ipm_issues:
                    header_cell.value = 'IPM issues'
                    header_cell.fill = style_condition
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_conservation_treatments_performed:
                    header_cell.value = 'Conservation treatments performed'
                    header_cell.fill = style_condition
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_accession_number:
                    header_cell.value = 'Accession number'
                    header_cell.fill = style_accessions
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_accession_date:
                    header_cell.value = 'Accession date'
                    header_cell.fill = style_accessions
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_type_of_accession:
                    header_cell.value = 'Type of accession'
                    header_cell.fill = style_accessions
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_acquisition_notes:
                    header_cell.value = 'Acquisition notes'
                    header_cell.fill = style_accessions
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_project_grant:
                    header_cell.value = 'Project/grant'
                    header_cell.fill = style_accessions
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_collection_name:
                    header_cell.value = 'Collection name'
                    header_cell.fill = style_accessions
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_collector_name:
                    header_cell.value = 'Collector name'
                    header_cell.fill = style_accessions
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_collector_info:
                    header_cell.value = 'Collector information'
                    header_cell.fill = style_accessions
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_collectors_number:
                    header_cell.value = "Collector's number"
                    header_cell.fill = style_accessions
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_collection_date:
                    header_cell.value = 'Collection date'
                    header_cell.fill = style_accessions
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_collecting_notes:
                    header_cell.value = 'Collecting notes'
                    header_cell.fill = style_accessions
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_depositor_name:
                    header_cell.value = 'Depositor name'
                    header_cell.fill = style_accessions
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_depositor_contact_information:
                    header_cell.value = 'Depositor contact information'
                    header_cell.fill = style_accessions
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_deposit_date:
                    header_cell.value = 'Deposit date'
                    header_cell.fill = style_accessions
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_municipality_or_township:
                    header_cell.value = 'Municipality or township'
                    header_cell.fill = style_location
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_county_or_parish:
                    header_cell.value = 'County or parish'
                    header_cell.fill = style_location
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_state_or_province:
                    header_cell.value = 'State or province'
                    header_cell.fill = style_location
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_country_or_territory:
                    header_cell.value = 'Country or territory'
                    header_cell.fill = style_location
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_global_region:
                    header_cell.value = 'Global region'
                    header_cell.fill = style_location
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_recording_context:
                    header_cell.value = 'Recording context'
                    header_cell.fill = style_location
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_public_event:
                    header_cell.value = 'Public event'
                    header_cell.fill = style_location
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_geographic_lat_long:
                    header_cell.value = 'Latitude'
                    header_cell.fill = style_location
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                    header_cell.value = 'Longitude'
                    header_cell.fill = style_location
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_original_format_medium:
                    header_cell.value = 'Original format medium'
                    header_cell.fill = style_digitization
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_recorded_on:
                    header_cell.value = 'Recorded on'
                    header_cell.fill = style_digitization
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_equipment_used:
                    header_cell.value = 'Equipment used'
                    header_cell.fill = style_digitization
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_software_used:
                    header_cell.value = 'Software used'
                    header_cell.fill = style_digitization
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_conservation_recommendation:
                    header_cell.value = 'Conservation recommendation'
                    header_cell.fill = style_digitization
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_location_of_original:
                    header_cell.value = 'Location of original'
                    header_cell.fill = style_digitization
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_other_information:
                    header_cell.value = 'Other information'
                    header_cell.fill = style_digitization
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_publisher:
                    header_cell.value = 'Publisher'
                    header_cell.fill = style_books
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_publisher_address:
                    header_cell.value = 'Publisher address'
                    header_cell.fill = style_books
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_isbn:
                    header_cell.value = 'ISBN'
                    header_cell.fill = style_books
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_loc_catalog_number:
                    header_cell.value = 'LOC catalog number'
                    header_cell.fill = style_books
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_total_number_of_pages_and_physical_description:
                    header_cell.value = 'Total number of pages and physical description'
                    header_cell.fill = style_books
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_temporary_accession_number:
                    header_cell.value = 'Temporary accession number'
                    header_cell.fill = style_external
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_lender_loan_number:
                    header_cell.value = 'Lender loan number'
                    header_cell.fill = style_external
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_other_institutional_number:
                    header_cell.value = 'Other institutional number'
                    header_cell.fill = style_external
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_migration_file_format:
                    header_cell.value = 'Migration file format'
                    header_cell.fill = style_deprecated
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_migration_location:
                    header_cell.value = 'Migration location'
                    header_cell.fill = style_deprecated
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_digital_file_location:
                    header_cell.value = 'Digital file location'
                    header_cell.fill = style_deprecated
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_cataloged_by:
                    header_cell.value = 'Cataloged by'
                    header_cell.fill = style_deprecated
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_cataloged_date:
                    header_cell.value = 'Cataloged date'
                    header_cell.fill = style_deprecated
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                if column_choice.item_filemaker_legacy_pk_id:
                    header_cell.value = 'Filemaker legacy PK ID'
                    header_cell.fill = style_deprecated
                    sheet_column_counter += 1
                    header_cell = sheet.cell(row=1, column=sheet_column_counter )
                column_counter = 1
                while column_counter <= max_document_counts:
                    if column_choice.item_document_filename:
                        header_cell.value = 'Document %s filename' %column_counter
                        if column_counter % 2 == 0: #even
                            header_cell.fill = style_document_even
                        else:
                            header_cell.fill = style_document_odd
                        sheet_column_counter += 1
                        header_cell = sheet.cell(row=1, column=sheet_column_counter )
                    if column_choice.item_document_filetype:
                        header_cell.value = 'Document %s filetype' %column_counter
                        if column_counter % 2 == 0: #even
                            header_cell.fill = style_document_even
                        else:
                            header_cell.fill = style_document_odd
                        sheet_column_counter += 1
                        header_cell = sheet.cell(row=1, column=sheet_column_counter )
                    if column_choice.item_document_access_level:
                        header_cell.value = 'Document %s access level' %column_counter
                        if column_counter % 2 == 0: #even
                            header_cell.fill = style_document_even
                        else:
                            header_cell.fill = style_document_odd
                        sheet_column_counter += 1
                        header_cell = sheet.cell(row=1, column=sheet_column_counter )
                    if column_choice.item_document_enumerator:
                        header_cell.value = 'Document %s enumerator' %column_counter
                        if column_counter % 2 == 0: #even
                            header_cell.fill = style_document_even
                        else:
                            header_cell.fill = style_document_odd
                        sheet_column_counter += 1
                        header_cell = sheet.cell(row=1, column=sheet_column_counter )
                    if column_choice.item_document_title:
                        header_cell.value = 'Document %s title' %column_counter
                        if column_counter % 2 == 0: #even
                            header_cell.fill = style_document_even
                        else:
                            header_cell.fill = style_document_odd
                        sheet_column_counter += 1
                        header_cell = sheet.cell(row=1, column=sheet_column_counter )
                    if column_choice.item_document_duration:
                        header_cell.value = 'Document %s duration' %column_counter
                        if column_counter % 2 == 0: #even
                            header_cell.fill = style_document_even
                        else:
                            header_cell.fill = style_document_odd
                        sheet_column_counter += 1
                        header_cell = sheet.cell(row=1, column=sheet_column_counter )
                    if column_choice.item_document_filesize:
                        header_cell.value = 'Document %s filesize' %column_counter
                        if column_counter % 2 == 0: #even
                            header_cell.fill = style_document_even
                        else:
                            header_cell.fill = style_document_odd
                        sheet_column_counter += 1
                        header_cell = sheet.cell(row=1, column=sheet_column_counter )
                    if column_choice.item_document_av_spec:
                        header_cell.value = 'Document %s A/V specs' %column_counter
                        if column_counter % 2 == 0: #even
                            header_cell.fill = style_document_even
                        else:
                            header_cell.fill = style_document_odd
                        sheet_column_counter += 1
                        header_cell = sheet.cell(row=1, column=sheet_column_counter )
                    if column_choice.item_document_creation_date:
                        header_cell.value = 'Document %s creation date' %column_counter
                        if column_counter % 2 == 0: #even
                            header_cell.fill = style_document_even
                        else:
                            header_cell.fill = style_document_odd
                        sheet_column_counter += 1
                        header_cell = sheet.cell(row=1, column=sheet_column_counter )
                    if column_choice.item_document_language:
                        language_column_counter = 1
                        while language_column_counter <= max_document_language_counts:
                            header_cell.value = ('Document %(document_count)s language name %(language_count)s' %{'document_count': column_counter, 'language_count': language_column_counter})
                            if column_counter % 2 == 0: #even
                                header_cell.fill = style_document_even
                            else:
                                header_cell.fill = style_document_odd
                            sheet_column_counter += 1
                            header_cell = sheet.cell(row=1, column=sheet_column_counter )
                            if column_choice.item_document_dialect:
                                header_cell.value = ('Document %(document_count)s dialect %(language_count)s' %{'document_count': column_counter, 'language_count': language_column_counter})
                                if column_counter % 2 == 0: #even
                                    header_cell.fill = style_document_even
                                else:
                                    header_cell.fill = style_document_odd
                                sheet_column_counter += 1
                                header_cell = sheet.cell(row=1, column=sheet_column_counter )
                            language_column_counter += 1

                    if column_choice.item_document_collaborator:
                        collaborator_column_counter = 1
                        while collaborator_column_counter <= max_document_collaborator_counts:
                            header_cell.value = 'Document %(document_count)s collaborator %(collaborator_count)s' %{'document_count': column_counter, "collaborator_count": collaborator_column_counter}
                            if column_counter % 2 == 0: #even
                                header_cell.fill = style_document_even
                            else:
                                header_cell.fill = style_document_odd
                            sheet_column_counter += 1
                            header_cell = sheet.cell(row=1, column=sheet_column_counter )
                            if column_choice.item_document_collaborator_role:
                                header_cell.value = ('Document %(document_count)s collaborator role %(collaborator_count)s' %{'document_count': column_counter, "collaborator_count": collaborator_column_counter})
                                if column_counter % 2 == 0: #even
                                    header_cell.fill = style_document_even
                                else:
                                    header_cell.fill = style_document_odd
                                sheet_column_counter += 1
                                header_cell = sheet.cell(row=1, column=sheet_column_counter )
                            collaborator_column_counter += 1

                    if column_choice.item_document_geographic_lat_long:
                        header_cell.value = 'Document %s latitude' %column_counter
                        if column_counter % 2 == 0: #even
                            header_cell.fill = style_document_even
                        else:
                            header_cell.fill = style_document_odd
                        sheet_column_counter += 1
                        header_cell = sheet.cell(row=1, column=sheet_column_counter )
                        header_cell.value = 'Document %s longitude' %column_counter
                        if column_counter % 2 == 0: #even
                            header_cell.fill = style_document_even
                        else:
                            header_cell.fill = style_document_odd
                        sheet_column_counter += 1
                        header_cell = sheet.cell(row=1, column=sheet_column_counter )
                    column_counter += 1

                for item in qs:
                    xl_row = []
                    if column_choice.item_catalog_number:
                        xl_row.append(item.catalog_number)
                    if column_choice.item_item_access_level:
                        xl_row.append(item.get_item_access_level_display())
                    if column_choice.item_call_number:
                        xl_row.append(item.call_number)
                    if column_choice.item_accession_date:
                        xl_row.append(item.accession_date)
                    if column_choice.item_additional_digital_file_location:
                        xl_row.append(item.additional_digital_file_location)
                    if column_choice.item_indigenous_title:
                        xl_row.append(item.indigenous_title)
                    if column_choice.item_english_title:
                        xl_row.append(item.english_title)
                    if column_choice.item_general_content:
                        xl_row.append(item.get_general_content_display())
                    if column_choice.item_language:
                        language_rows = []
                        language_rows.extend( item.language.all().values_list('name', flat=True).order_by('name') )
                        language_rows.extend( [''] * ( max_language_counts - len(item.language.all()) ) )
                        if column_choice.item_dialect:
                            dialect_rows = []
                            item_dialects = item.item_dialectinstances.all().order_by('language__name')
                            for item_dialect in item_dialects:
                                dialect_rows.append( "\n".join( item_dialect.name.all().values_list('name', flat=True).order_by('name') ) )
                            #dialect_rows.extend( item.item_dialectinstances.all().values_list('name__name', flat=True).order_by('language__name') )
                            dialect_rows.extend( [''] * ( max_language_counts - len(item.item_dialectinstances.all()) ) )
                            language_zip = zip(language_rows, dialect_rows)
                            language_rows = [i for sublist in language_zip for i in sublist]
                        xl_row.extend(language_rows)
                    if column_choice.item_creation_date:
                        xl_row.append(item.creation_date)
                    if column_choice.item_description_scope_and_content:
                        xl_row.append(item.description_scope_and_content)
                    if column_choice.item_genre:
                        xl_row.append(item.get_genre_display().replace(', ','\n') )
                    if column_choice.item_associated_ephemera:
                        xl_row.append(item.associated_ephemera)


                    if column_choice.item_access_level_restrictions:
                        xl_row.append(item.access_level_restrictions)
                    if column_choice.item_copyrighted_notes:
                        xl_row.append(item.copyrighted_notes)
                    if column_choice.item_permission_to_publish_online:
                        if item.permission_to_publish_online:
                            xl_row.append('yes')
                        elif item.permission_to_publish_online is None:
                            xl_row.append('')
                        else:
                            xl_row.append('no')
                    if column_choice.item_collaborator:
                        collaborator_rows = []
                        collaborator_rows.extend( item.collaborator.all().values_list('name', flat=True).order_by('name') )
                        collaborator_rows.extend( [''] * ( max_collaborator_counts - len(item.collaborator.all()) ) )
                        if column_choice.item_collaborator_role:
                            collaborator_role_rows = []
                            item_collaborator_roles = item.item_collaboratorroles.all().order_by('collaborator__name')
                            for item_collaborator_role in item_collaborator_roles:
                                collaborator_role_rows.append( item_collaborator_role.get_role_display() )
                            #collaborator_role_rows.extend( item.item_collaboratorroles.all().values_list('role', flat=True).order_by('collaborator__name') )
                            collaborator_role_rows.extend( [''] * ( max_collaborator_counts - len(item.item_collaboratorroles.all()) ) )
                            collaborator_role_rows_joined = []
                            for collaborator_role_row in collaborator_role_rows:
                                collaborator_role_rows_joined.append( collaborator_role_row.replace(', ','\n') )
                            collaborator_zip = zip(collaborator_rows, collaborator_role_rows_joined)
                            collaborator_rows = [i for sublist in collaborator_zip for i in sublist]
                        xl_row.extend(collaborator_rows)
                    if column_choice.item_language_description_type:
                        xl_row.append(item.get_language_description_type_display().replace(', ','\n') )
                    if column_choice.item_availability_status:
                        xl_row.append(item.get_availability_status_display())
                    if column_choice.item_availability_status_notes:
                        xl_row.append(item.availability_status_notes)
                    if column_choice.item_condition:
                        xl_row.append(item.get_condition_display())
                    if column_choice.item_condition_notes:
                        xl_row.append(item.condition_notes)
                    if column_choice.item_ipm_issues:
                        xl_row.append(item.ipm_issues)
                    if column_choice.item_conservation_treatments_performed:
                        xl_row.append(item.conservation_treatments_performed)
                    if column_choice.item_accession_number:
                        xl_row.append(item.accession_number)
                    if column_choice.item_accession_date:
                        xl_row.append(item.accession_date)
                    if column_choice.item_type_of_accession:
                        xl_row.append(item.get_type_of_accession_display())
                    if column_choice.item_acquisition_notes:
                        xl_row.append(item.acquisition_notes)
                    if column_choice.item_project_grant:
                        xl_row.append(item.project_grant)
                    if column_choice.item_collection_name:
                        xl_row.append(item.collection_name)
                    if column_choice.item_collector_name:
                        xl_row.append(item.collector_name)
                    if column_choice.item_collector_info:
                        xl_row.append(item.collector_info)
                    if column_choice.item_collectors_number:
                        xl_row.append(item.collectors_number)
                    if column_choice.item_collection_date:
                        xl_row.append(item.collection_date)
                    if column_choice.item_collecting_notes:
                        xl_row.append(item.collecting_notes)
                    if column_choice.item_depositor_name:
                        xl_row.append(item.depositor_name)
                    if column_choice.item_depositor_contact_information:
                        xl_row.append(item.depositor_contact_information)
                    if column_choice.item_deposit_date:
                        xl_row.append(item.deposit_date)
                    if column_choice.item_municipality_or_township:
                        xl_row.append(item.municipality_or_township)
                    if column_choice.item_county_or_parish:
                        xl_row.append(item.county_or_parish)
                    if column_choice.item_state_or_province:
                        xl_row.append(item.state_or_province)
                    if column_choice.item_country_or_territory:
                        xl_row.append(item.country_or_territory)
                    if column_choice.item_global_region:
                        xl_row.append(item.global_region)
                    if column_choice.item_recording_context:
                        xl_row.append(item.recording_context)
                    if column_choice.item_public_event:
                        xl_row.append(item.public_event)
                    if column_choice.item_geographic_lat_long:
                        if item.item_geographic.first():
                            xl_row.append(item.item_geographic.first().lat)
                            xl_row.append(item.item_geographic.first().long)
                        else:
                            xl_row.append('')
                            xl_row.append('')
                    if column_choice.item_original_format_medium:
                        xl_row.append(item.get_original_format_medium_display())
                    if column_choice.item_recorded_on:
                        xl_row.append(item.recorded_on)
                    if column_choice.item_equipment_used:
                        xl_row.append(item.equipment_used)
                    if column_choice.item_software_used:
                        xl_row.append(item.software_used)
                    if column_choice.item_conservation_recommendation:
                        xl_row.append(item.conservation_recommendation)
                    if column_choice.item_location_of_original:
                        xl_row.append(item.location_of_original)
                    if column_choice.item_other_information:
                        xl_row.append(item.other_information)
                    if column_choice.item_publisher:
                        xl_row.append(item.publisher)
                    if column_choice.item_publisher_address:
                        xl_row.append(item.publisher_address)
                    if column_choice.item_isbn:
                        xl_row.append(item.isbn)
                    if column_choice.item_loc_catalog_number:
                        xl_row.append(item.loc_catalog_number)
                    if column_choice.item_total_number_of_pages_and_physical_description:
                        xl_row.append(item.total_number_of_pages_and_physical_description)
                    if column_choice.item_temporary_accession_number:
                        xl_row.append(item.temporary_accession_number)
                    if column_choice.item_lender_loan_number:
                        xl_row.append(item.lender_loan_number)
                    if column_choice.item_other_institutional_number:
                        xl_row.append(item.other_institutional_number)
                    if column_choice.item_migration_file_format:
                        xl_row.append(item.migration_file_format)
                    if column_choice.item_migration_location:
                        xl_row.append(item.migration_location)
                    if column_choice.item_digital_file_location:
                        xl_row.append(item.digital_file_location)
                    if column_choice.item_cataloged_by:
                        xl_row.append(item.cataloged_by)
                    if column_choice.item_cataloged_date:
                        xl_row.append(item.cataloged_date)
                    if column_choice.item_filemaker_legacy_pk_id:
                        xl_row.append(item.filemaker_legacy_pk_id)

                    current_item_documents = item.item_documents.all().order_by('filename')
                    for current_item_document in current_item_documents:
                        if column_choice.item_document_filename:
                            xl_row.append(current_item_document.filename)
                        if column_choice.item_document_filetype:
                            xl_row.append(current_item_document.filetype)
                        if column_choice.item_document_access_level:
                            xl_row.append(current_item_document.get_access_level_display())
                        if column_choice.item_document_enumerator:
                            xl_row.append(current_item_document.enumerator)
                        if column_choice.item_document_title:
                            xl_row.append(current_item_document.title)
                        if column_choice.item_document_duration:
                            xl_row.append(current_item_document.duration)
                        if column_choice.item_document_filesize:
                            xl_row.append(current_item_document.filesize)
                        if column_choice.item_document_av_spec:
                            xl_row.append(current_item_document.av_spec)
                        if column_choice.item_document_creation_date:
                            xl_row.append(current_item_document.creation_date)
                        if column_choice.item_document_language:
                            language_rows = []
                            language_rows.extend( current_item_document.language.all().values_list('name', flat=True).order_by('name') )
                            language_rows.extend( [''] * ( max_document_language_counts - len(current_item_document.language.all()) ) )
                            if column_choice.item_document_dialect:
                                dialect_rows = []
                                current_item_document_dialects = current_item_document.document_dialectinstances.all().order_by('language__name')
                                for current_item_document_dialect in current_item_document_dialects:
                                    dialect_rows.append( "\n".join( current_item_document_dialect.name.all().values_list('name', flat=True).order_by('name') ) )
                                #dialect_rows.extend( current_item_document.document_dialectinstances.all().values_list('name__name', flat=True).order_by('language__name') )
                                dialect_rows.extend( [''] * ( max_document_language_counts - len(current_item_document.document_dialectinstances.all()) ) )
                                language_zip = zip(language_rows, dialect_rows)
                                language_rows = [i for sublist in language_zip for i in sublist]
                            xl_row.extend(language_rows)


                        if column_choice.item_document_collaborator:
                            collaborator_rows = []
                            collaborator_rows.extend( current_item_document.collaborator.all().values_list('name', flat=True).order_by('name') )
                            collaborator_rows.extend( [''] * ( max_document_collaborator_counts - len(current_item_document.collaborator.all()) ) )
                            if column_choice.item_document_collaborator_role:
                                collaborator_role_rows = []
                                current_item_document_collaborator_roles = current_item_document.document_collaboratorroles.all().order_by('collaborator__name')
                                for current_item_document_collaborator_role in current_item_document_collaborator_roles:
                                    collaborator_role_rows.append( current_item_document_collaborator_role.get_role_display() )

                                #collaborator_role_rows.extend( current_item_document.document_collaboratorroles.all().values_list('role', flat=True).order_by('collaborator__name') )
                                collaborator_role_rows.extend( [''] * ( max_document_collaborator_counts - len(current_item_document.document_collaboratorroles.all()) ) )
                                collaborator_role_rows_joined = []
                                for collaborator_role_row in collaborator_role_rows:
                                    collaborator_role_rows_joined.append( collaborator_role_row.replace(', ','\n') )
                                collaborator_zip = zip(collaborator_rows, collaborator_role_rows_joined)
                                collaborator_rows = [i for sublist in collaborator_zip for i in sublist]
                            xl_row.extend(collaborator_rows)
                        if column_choice.item_document_geographic_lat_long:
                            if current_item_document.document_geographic.first():
                                xl_row.append(current_item_document.document_geographic.first().lat)
                                xl_row.append(current_item_document.document_geographic.first().long)
                            else:
                                xl_row.append('')
                                xl_row.append('')


                    for blank_document in range(max_document_counts - len(current_item_documents)):
                        if column_choice.item_document_filename:
                            xl_row.append('')
                        if column_choice.item_document_filetype:
                            xl_row.append('')
                        if column_choice.item_document_access_level:
                            xl_row.append('')
                        if column_choice.item_document_enumerator:
                            xl_row.append('')
                        if column_choice.item_document_title:
                            xl_row.append('')
                        if column_choice.item_document_duration:
                            xl_row.append('')
                        if column_choice.item_document_filesize:
                            xl_row.append('')
                        if column_choice.item_document_av_spec:
                            xl_row.append('')
                        if column_choice.item_document_creation_date:
                            xl_row.append('')
                        if column_choice.item_document_language:
                            xl_row.extend( [''] * max_document_language_counts )
                        if column_choice.item_document_dialect:
                            xl_row.extend( [''] * max_document_language_counts )
                        if column_choice.item_document_collaborator:
                            xl_row.extend( [''] * max_document_collaborator_counts )
                        if column_choice.item_document_collaborator_role:
                            xl_row.extend( [''] * max_document_collaborator_counts )
                        if column_choice.item_document_geographic_lat_long:
                            xl_row.append('')
                            xl_row.append('')


    #####################################################
                    # document_rows_zip = []
                    # if column_choice.item_document_filename:
                    #     document_filename_rows = []
                    #     document_filename_rows.extend( item.item_documents.all().values_list('filename', flat=True) )
                    #     document_filename_rows.extend( [''] * ( max_document_counts - len(item.item_documents.all()) ) )
                    #     document_rows_zip.append('document_filename_rows')
                    # if column_choice.item_document_filetype:
                    #     document_filetype_rows = []
                    #     document_filetype_rows.extend( item.item_documents.all().values_list('filetype', flat=True) )
                    #     document_filetype_rows.extend( [''] * ( max_document_counts - len(item.item_documents.all()) ) )
                    #     document_rows_zip.append('document_filetype_rows')
                    # if column_choice.item_document_enumerator:
                    #     document_enumerator_rows = []
                    #     document_enumerator_rows.extend( item.item_documents.all().values_list('enumerator', flat=True) )
                    #     document_enumerator_rows.extend( [''] * ( max_document_counts - len(item.item_documents.all()) ) )
                    #     document_rows_zip.append('document_enumerator_rows')
                    # if column_choice.item_document_title:
                    #     document_title_rows = []
                    #     document_title_rows.extend( item.item_documents.all().values_list('title', flat=True) )
                    #     document_title_rows.extend( [''] * ( max_document_counts - len(item.item_documents.all()) ) )
                    #     document_rows_zip.append('document_title_rows')
                    # if column_choice.item_document_duration:
                    #     document_duration_rows = []
                    #     document_duration_rows.extend( item.item_documents.all().values_list('duration', flat=True) )
                    #     document_duration_rows.extend( [''] * ( max_document_counts - len(item.item_documents.all()) ) )
                    #     document_rows_zip.append('document_duration_rows')
                    # if column_choice.item_document_filesize:
                    #     document_filesize_rows = []
                    #     document_filesize_rows.extend( item.item_documents.all().values_list('filesize', flat=True) )
                    #     document_filesize_rows.extend( [''] * ( max_document_counts - len(item.item_documents.all()) ) )
                    #     document_rows_zip.append('document_filesize_rows')
                    # if column_choice.item_document_av_spec:
                    #     document_av_spec_rows = []
                    #     document_av_spec_rows.extend( item.item_documents.all().values_list('av_spec', flat=True) )
                    #     document_av_spec_rows.extend( [''] * ( max_document_counts - len(item.item_documents.all()) ) )
                    #     document_rows_zip.append('document_av_spec_rows')
                    # if column_choice.item_document_language:
                    #     document_language_rows = []
                    #     document_languages = item.item_documents.all()
                    #     for each_document_languages in document_languages:
                    #         document_language_rows.append( ", ".join( each_document_languages.language.all().values_list('name', flat=True) ) )
                    #     document_language_rows.extend( [''] * ( max_document_counts - len(item.item_documents.all()) ) )
                    #     document_rows_zip.append('document_language_rows')
                    # if column_choice.item_document_collaborator:
                    #     document_collaborator_rows = []
                    #     document_collaborators = item.item_documents.all()
                    #     for each_document_collaborators in document_collaborators:
                    #         document_collaborator_rows.append( ",".join( each_document_collaborators.collaborator.all().values_list('name', flat=True) ) )
                    #     document_collaborator_rows.extend( [''] * ( max_document_counts - len(item.item_documents.all()) ) )
                    #     document_rows_zip.append('document_collaborator_rows')
                    #
                    # if document_rows_zip:
                    #     document_rows = eval( document_rows_zip[0] )
                    #     if len(document_rows_zip) > 1:
                    #         document_rows = []
                    #         for i in range(max_document_counts):
                    #             for j in range(len(document_rows_zip)):
                    #                 document_rows.append( eval( document_rows_zip[j] )[i] )
                    #
                    #     xl_row.extend( document_rows )
                    #

    ######################################



                    sheet.append(xl_row)


                response = HttpResponse(content=save_virtual_workbook(new_workbook), content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = 'attachment; filename=metadata-export.xlsx'

                return response


    paginator = Paginator(qs, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if re.search('search', url_path, flags=re.I):
        template = 'item_search.html'
    elif re.search('migrate', url_path, flags=re.I):
        template = 'item_migrate.html'
    else:
        template = 'item_index.html'

    context = {
        'queryset': page_obj,
        'results_count' : results_count,
        'results_duration' : results_duration_datetime,
        'order_choice_last' : order_choice_last,
        'columns_choice_name_last' : columns_choice_name_last,
        'catalog_number_contains_query_last' : catalog_number_contains_query_last,
        'item_access_level_contains_query_last' : item_access_level_contains_query_last,
        'call_number_contains_query_last' : call_number_contains_query_last,
        'accession_date_min_query_last' : accession_date_min_query_last,
        'accession_date_max_query_last' : accession_date_max_query_last,
        'indigenous_title_contains_query_last' : indigenous_title_contains_query_last,
        'english_title_contains_query_last' : english_title_contains_query_last,
        'titles_contains_query_last' : titles_contains_query_last,
        'general_content_contains_query_last' : general_content_contains_query_last,
        'language_contains_query_last' : language_contains_query_last,
        'creation_date_min_query_last' : creation_date_min_query_last,
        'creation_date_max_query_last' : creation_date_max_query_last,
        'description_scope_and_content_contains_query_last' : description_scope_and_content_contains_query_last,
        'genre_contains_query_last' : genre_contains_query_last,
        'collaborator_contains_query_last' : collaborator_contains_query_last,
        'depositor_name_contains_query_last' : depositor_name_contains_query_last,
        'keyword_contains_query_last' : keyword_contains_query_last,
        'columns_form_data' : Columns_export.objects.values_list('name', flat=True)
    }
    return render(request, template, context)

@login_required
@user_passes_test(is_member_of_archivist, login_url='/no-permission', redirect_field_name=None)
def item_migrate_list(request):
    qs = Item.objects.filter(migrate=True).order_by('catalog_number')
    paginator = Paginator(qs, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'queryset': page_obj
    }
    return render(request, 'item_migrate.html', context)

@login_required
def item_detail(request, pk):
    item = Item.objects.get(pk=pk)

    # get all the titles for the item
    titles = ItemTitle.objects.filter(item__pk=pk).order_by('title')

    documents = Document.objects.filter(item__pk=pk).order_by('filename')

    for collaborator in item.collaborator.all():
        collaborator_role_i, created = CollaboratorRole.objects.get_or_create(item=item, collaborator=collaborator)
    collaborator_info = zip(item.collaborator.all().order_by('name'), CollaboratorRole.objects.filter(item=item).order_by('collaborator__name'))

    for language in item.language.all():
        dialect_i, created = DialectInstance.objects.get_or_create(item=item, language=language)
    dialect_info = zip(item.language.all().order_by('name'), DialectInstance.objects.filter(item=item).order_by('language__name'))


    geographic_info = item.item_geographic.all()
    geographic_points = []
    for data_point in geographic_info:
        geographic_points.append((data_point.id, data_point.lat, data_point.long))

    context = {
        'item': item,
        'titles' : titles,
        'documents' : documents,
        'collaborator_info' : collaborator_info,
        'dialect_info' : dialect_info,
        'geographic_info' : geographic_info,
        'geographic_points' : geographic_points
    }
    return render(request, 'item_detail.html', context)

@login_required
@user_passes_test(is_member_of_archivist, login_url='/no-permission', redirect_field_name=None)
def item_edit(request, pk):
    item = get_object_or_404(Item, id=pk)
    if request.method == "POST":
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():

            document_prefix_with_space = str(item.catalog_number.lower()) + ' '
            documents_with_space = Document.objects.filter(filename__icontains = document_prefix_with_space)
            document_prefix_with_underscore = str(item.catalog_number.lower()) + '_'
            documents_with_underscore = Document.objects.filter(filename__icontains = document_prefix_with_underscore)
            matching_documents = documents_with_space.union(documents_with_underscore)
            if matching_documents:
                for matching_document in matching_documents:
                    matching_document.item = item
                    matching_document.save()

            item.modified_by = request.user.get_username()
            item.save()
            form.save()
            return redirect("../")
    else:
        form = ItemForm(instance=item)
    return render(request, 'item_edit.html', {'form': form})


class item_add(UserPassesTestMixin, FormView):
    def test_func(self):
        return self.request.user.groups.filter(name="Archivist").exists()
    def handle_no_permission(self):
        return redirect('/no-permission')
    form_class = ItemForm
    template_name = "add.html"
    def form_valid(self, form):
        self.object = form.save()
        pk = self.object.pk
        instance = Item.objects.get(pk=pk)
        instance.modified_by = self.request.user.get_username()
        instance.save()

        document_prefix_with_space = str(instance.catalog_number.lower()) + ' '
        documents_with_space = Document.objects.filter(filename__icontains = document_prefix_with_space)
        document_prefix_with_underscore = str(instance.catalog_number.lower()) + '_'
        documents_with_underscore = Document.objects.filter(filename__icontains = document_prefix_with_underscore)
        matching_documents = documents_with_space.union(documents_with_underscore)
        if matching_documents:
            for matching_document in matching_documents:
                matching_document.item = instance
                matching_document.save()

        return redirect("../%s/" %pk )

class item_delete(UserPassesTestMixin, DeleteView):
    def test_func(self):
        return self.request.user.groups.filter(name="Archivist").exists()
    def handle_no_permission(self):
        return redirect('/no-permission')
    model = Item
    success_url = '/catalog/'



@login_required
def document_index(request):
    qs = Document.objects.all()
    order_choice = request.GET.get("form_control_sort")
    filename_contains_query = request.GET.get('filename_contains')
    filetype_contains_query = request.GET.get('filetype_contains')
    title_contains_query = request.GET.get('title_contains')
    language_contains_query = request.GET.get('language_contains')

    order_choice_last = order_choice
    filename_contains_query_last = ''
    filetype_contains_query_last = ''
    title_contains_query_last = ''
    language_contains_query_last = ''


    if is_valid_param(filename_contains_query):
        qs = qs.filter(filename__icontains = filename_contains_query)
        filename_contains_query_last = filename_contains_query

    if is_valid_param(filetype_contains_query):
        qs = qs.filter(filetype__icontains = filetype_contains_query)
        filetype_contains_query_last = filetype_contains_query

    if is_valid_param(title_contains_query):
        qs = qs.filter(title__icontains = title_contains_query)
        title_contains_query_last = title_contains_query

    if is_valid_param(language_contains_query):
        qs = qs.filter(language__name__icontains = language_contains_query)
        language_contains_query_last = language_contains_query

    qs = qs.distinct()

#    print(type(order_choice))
    if order_choice == "updated":
        qs = qs.order_by('-updated')
    else:
        qs = qs.order_by('filename')

    results_count = qs.count()

    if results_count > 0:
        results_duration = qs.aggregate(sum_duration=Sum('duration')).get('sum_duration')
        if results_duration:
            results_duration_datetime = str(datetime.timedelta(seconds=results_duration))[:-3]
        else:
            results_duration_datetime = 0
    else:
        results_duration_datetime = 0
#        results_duration_datetime = str(results_duration)

    paginator = Paginator(qs, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'queryset': page_obj,
        'results_count' : results_count,
        'results_duration' : results_duration_datetime,
        'order_choice_last' : order_choice_last,
        'filename_contains_query_last' : filename_contains_query_last,
        'filetype_contains_query_last' : filetype_contains_query_last,
        'title_contains_query_last' : title_contains_query_last,
        'language_contains_query_last' : language_contains_query_last,
    }
    return render(request, 'document_index.html', context)

@login_required
def document_detail(request, pk):
    qs = Document.objects.get(pk=pk)

    if not qs.duration:
        pretty_duration = qs.duration
    else:
        pretty_duration = str(datetime.timedelta(seconds=qs.duration))[:-3]

    for collaborator in qs.collaborator.all():
        collaborator_role_i, created = CollaboratorRole.objects.get_or_create(document=qs, collaborator=collaborator)
    collaborator_info = zip(qs.collaborator.all().order_by('name'), CollaboratorRole.objects.filter(document=qs).order_by('collaborator__name'))

    for language in qs.language.all():
        dialect_i, created = DialectInstance.objects.get_or_create(document=qs, language=language)
    dialect_info = zip(qs.language.all().order_by('name'), DialectInstance.objects.filter(document=qs).order_by('language__name'))

    geographic_info = qs.document_geographic.all()
    geographic_points = []
    for data_point in geographic_info:
        geographic_points.append((data_point.id, data_point.lat, data_point.long))

    context = {
        'document': qs,
        'pretty_duration' : pretty_duration,
        'collaborator_info' : collaborator_info,
        'dialect_info' : dialect_info,
        'geographic_info' : geographic_info,
        'geographic_points' : geographic_points
    }
    return render(request, 'document_detail.html', context)

@login_required
@user_passes_test(is_member_of_archivist, login_url='/no-permission', redirect_field_name=None)
def document_edit(request, pk):
    qs = get_object_or_404(Document, id=pk)
    if request.method == "POST":
        form = DocumentForm(request.POST, instance=qs)
        if form.is_valid():

            filename_modified = str(qs.filename).replace(' ','_')
            filename_prefix = filename_modified.split("_")[0]
            matching_items = Item.objects.filter(catalog_number__iexact=filename_prefix)
            if is_valid_param(matching_items):
                existing_item = Item.objects.get(catalog_number=matching_items[0])
                qs.item = existing_item
            else:
                new_item = Item.objects.create(catalog_number=filename_prefix)
                qs.item = new_item
            qs.modified_by = request.user.get_username()
            qs.save()
            form.save()
            return redirect("../")
    else:
        form = DocumentForm(instance=qs)
    return render(request, 'document_edit.html', {'form': form})

class document_add(UserPassesTestMixin, FormView):
    def test_func(self):
        return self.request.user.groups.filter(name="Archivist").exists()
    def handle_no_permission(self):
        return redirect('/no-permission')
    form_class = DocumentForm
    template_name = "add.html"
    def form_valid(self, form):
        self.object = form.save()
        pk = self.object.pk
        instance = Document.objects.get(pk=pk)
        instance.modified_by = self.request.user.get_username()
        instance.save()

        filename_modified = str(instance.filename.lower()).replace(' ','_')
        filename_prefix = filename_modified.split("_")[0]
        matching_items = Item.objects.filter(catalog_number__iexact=filename_prefix)
        if is_valid_param(matching_items):
            existing_item = Item.objects.get(catalog_number=matching_items[0])
            instance.item = existing_item
            instance.save()

        return redirect("../%s/" %pk )

class document_delete(UserPassesTestMixin, DeleteView):
    def test_func(self):
        return self.request.user.groups.filter(name="Archivist").exists()
    def handle_no_permission(self):
        return redirect('/no-permission')
    model = Document
    success_url = '/documents/'






@login_required
def language_index(request):
    qs = Language.objects.all()
    order_choice = request.GET.get("form_control_sort")
    iso_contains_query = request.GET.get('iso_contains')
    glottocode_contains_query = request.GET.get('glottocode_contains')
    name_contains_query = request.GET.get('name_contains')
    family_contains_query = request.GET.get('family_contains')
    primary_subgroup_contains_query = request.GET.get('primary_subgroup_contains')
    region_contains_query = request.GET.get('region_contains')
    has_items_query = request.GET.get('has_items')

    order_choice_last = order_choice
    iso_contains_query_last = ''
    glottocode_contains_query_last = ''
    name_contains_query_last = ''
    family_contains_query_last = ''
    primary_subgroup_contains_query_last = ''
    region_contains_query_last = ''
    has_items_query_last = ''

    EMPTY_KEYWORD = 'EMPTY'

    if is_valid_param(iso_contains_query):
        if iso_contains_query == EMPTY_KEYWORD:
            qs = qs.filter(iso__exact='')
        else:
            qs = qs.filter(iso__icontains=iso_contains_query)
        iso_contains_query_last = iso_contains_query

    if is_valid_param(glottocode_contains_query):
        if glottocode_contains_query == EMPTY_KEYWORD:
            qs = qs.filter(glottocode__exact='')
        else:
            qs = qs.filter(glottocode__icontains = glottocode_contains_query)
        glottocode_contains_query_last = glottocode_contains_query

    if is_valid_param(name_contains_query):
        if name_contains_query == EMPTY_KEYWORD:
            qs = qs.filter(name__exact='')
        else:
            qs = qs.filter(name__icontains = name_contains_query)
        name_contains_query_last = name_contains_query

    if is_valid_param(family_contains_query):
        if family_contains_query == EMPTY_KEYWORD:
            qs = qs.filter(family__exact='')
        else:
            qs = qs.filter(family__icontains = family_contains_query)
        family_contains_query_last = family_contains_query

    if is_valid_param(primary_subgroup_contains_query):
        if primary_subgroup_contains_query == EMPTY_KEYWORD:
            qs = qs.filter(pri_subgroup__exact='')
        else:
            qs = qs.filter(pri_subgroup__icontains = primary_subgroup_contains_query)
        primary_subgroup_contains_query_last = primary_subgroup_contains_query
    
    if is_valid_param(region_contains_query):
        if region_contains_query == EMPTY_KEYWORD:
            qs = qs.filter(region__exact='')
        else:
            qs = qs.filter(region__icontains = region_contains_query)
        region_contains_query_last = region_contains_query

    if is_valid_param(has_items_query):
        if has_items_query == 'items':
            qs = qs.filter(item_languages__isnull=False)
        elif has_items_query == 'no_items':
            qs = qs.filter(item_languages__isnull=True)
        has_items_query_last = has_items_query

    qs = qs.distinct()
    qs = qs.annotate(item_count=Count('item_languages')).distinct()

#    print(type(order_choice))
    if order_choice == "updated":
        qs = qs.order_by('-updated')
    elif order_choice == "tree":
        qs = qs.order_by('family', 'pri_subgroup', 'sec_subgroup')
    elif order_choice == "iso":
        qs = qs.order_by('iso')
    else:
        qs = qs.order_by('name')

    results_count = qs.count()

    if 'export' in request.GET:
    # ported this over from item index view


        items_in_qs = Language.objects.filter(name__in=list(qs.values_list('name', flat=True)))

        new_workbook = Workbook()
        sheet = new_workbook.active

        sheet_column_counter = 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )

        header_cell.value = 'Language ISO code'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )

        header_cell.value = 'Glottocode'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )

        header_cell.value = 'Language name'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )

        header_cell.value = 'Alternative names'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )

        header_cell.value = 'Family'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )

        header_cell.value = 'Family abbreviation'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )

        header_cell.value = 'Family glottocode'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )

        header_cell.value = 'Primary subgroup'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )

        header_cell.value = 'Primary subgroup abbreviation'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )

        header_cell.value = 'Primary subgroup glottocode'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )

        header_cell.value = 'Secondary subgroup'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )

        header_cell.value = 'Secondary subgroup abbreviation'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )

        header_cell.value = 'Secondary subgroup glottocode'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )

        header_cell.value = 'Dialects'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )

        header_cell.value = 'Region'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )

        header_cell.value = 'Latitude'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )

        header_cell.value = 'Longitude'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )

        header_cell.value = 'Tribes'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )

        header_cell.value = 'Notes'
        sheet_column_counter += 1
        header_cell = sheet.cell(row=1, column=sheet_column_counter )


        for language in qs:
            xl_row = []
            xl_row.append(language.iso)
            xl_row.append(language.glottocode)
            xl_row.append(language.name)
            xl_row.append(language.alt_name)
            xl_row.append(language.family)
            xl_row.append(language.family_abbrev)
            xl_row.append(language.family_id)
            xl_row.append(language.pri_subgroup)
            xl_row.append(language.pri_subgroup_abbrev)
            xl_row.append(language.pri_subgroup_id)
            xl_row.append(language.sec_subgroup)
            xl_row.append(language.sec_subgroup_abbrev)
            xl_row.append(language.sec_subgroup_id)

            dialects_in_language = Dialect.objects.filter(language=language).values_list('name', flat=True).order_by('name')

            xl_row.append(", ".join( dialects_in_language ))
            xl_row.append(language.region)
            xl_row.append(language.latitude)
            xl_row.append(language.longitude)
            xl_row.append(language.tribes)
            xl_row.append(language.notes)

            sheet.append(xl_row)


        response = HttpResponse(content=save_virtual_workbook(new_workbook), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=languages-export.xlsx'

        return response

    paginator = Paginator(qs, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'queryset': page_obj,
        'results_count' : results_count,
        'order_choice_last' : order_choice_last,
        'iso_contains_query_last' : iso_contains_query_last,
        'glottocode_contains_query_last' : glottocode_contains_query_last,
        'name_contains_query_last' : name_contains_query_last,
        'family_contains_query_last' : family_contains_query_last,
        'primary_subgroup_contains_query_last' : primary_subgroup_contains_query_last,
        'region_contains_query_last' : region_contains_query_last,
        'has_items_query_last' : has_items_query_last
    }

    return render(request, 'language_index.html', context)

@login_required
def language_detail(request, pk):
    qs = Language.objects.get(pk=pk)

    dialect_info = Dialect.objects.filter(language=qs).order_by('name')

    collaborators_with_language = Collaborator.objects.filter(native_languages=qs).union(Collaborator.objects.filter(other_languages=qs))
    items_with_language = Item.objects.filter(language=qs)

    context = {
        'language': qs,
        'dialect_info' : dialect_info,
        'collaborators_with_language': collaborators_with_language,
        'items_with_language': items_with_language
    }

    return render(request, 'language_detail.html', context)

@login_required
@user_passes_test(is_member_of_archivist, login_url='/no-permission', redirect_field_name=None)
def language_edit(request, pk):
    qs = get_object_or_404(Language, id=pk)
    glcodes = []
    languoids = []
    if request.method == "POST":
        form = LanguageForm(request.POST, instance=qs)
        if form.is_valid():
            form.save(modified_by=request.user.get_username())
            return redirect("../")
    else:
        # Get the path to the CSV file
        csv_path = os.path.join(settings.STATIC_ROOT, 'codelist.csv')
        # Open the CSV file and read its contents
        with open(csv_path, 'r') as file:
            reader = csv.DictReader(file)
            for row in reader:
                glcodes.append(dict(row)['glottocode'])
                languoids.append(dict(row))
        form = LanguageForm(instance=qs)
    context = {

        'form': form,
        'glcodes': json.dumps(glcodes),
        'languoids': json.dumps(languoids),
        'title': 'Edit language'
    }
    return render(request, 'language_edit.html', context)

class language_add(UserPassesTestMixin, FormView):
    def test_func(self):
        return self.request.user.groups.filter(name="Archivist").exists()
    def handle_no_permission(self):
        return redirect('/no-permission')
    form_class = LanguageForm
    template_name = "language_edit.html"
    def form_valid(self, form):
        self.object = form.save(modified_by=self.request.user.get_username())
        pk = self.object.pk
        return redirect("../%s/" %pk )
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        glcodes = []
        languoids = []
        csv_path = os.path.join(settings.STATIC_ROOT, 'codelist.csv')
        with open(csv_path, 'r') as file:
            reader = csv.DictReader(file)
            for row in reader:
                glcodes.append(dict(row)['glottocode'])
                languoids.append(dict(row))
        context['glcodes'] = json.dumps(glcodes)
        context['languoids'] = json.dumps(languoids)
        context['title'] = 'Add a new language'
        return context

class language_delete(UserPassesTestMixin, DeleteView):
    def test_func(self):
        return self.request.user.groups.filter(name="Archivist").exists()
    def handle_no_permission(self):
        return redirect('/no-permission')
    model = Language
    success_url = '/languages/'

@login_required
@user_passes_test(is_member_of_archivist, login_url='/no-permission', redirect_field_name=None)
def language_stats(request):

    # Get a list of languages and the number of items associated with each language
    languages = Language.objects.annotate(num_items=Count('item_languages')).order_by('-num_items')

    context = {
        'languages' : languages
    }

    return render(request, 'language_stats.html', context)

@login_required
@user_passes_test(is_member_of_archivist, login_url='/no-permission', redirect_field_name=None)
def dialect_edit(request, pk):
    qs = get_object_or_404(Dialect, id=pk)
    if request.method == "POST":
        form = DialectForm(request.POST, instance=qs)
        if form.is_valid():
            qs.modified_by = request.user.get_username()
            qs.save()
            form.save()
            url = "../../../languages/%s/" %qs.language.pk
            return redirect(url)
    else:
        form = DialectForm(instance=qs)
    return render(request, 'dialect_edit.html', {'form': form})

class dialect_add(UserPassesTestMixin, FormView):
    def test_func(self):
        return self.request.user.groups.filter(name="Archivist").exists()
    def handle_no_permission(self):
        return redirect('/no-permission')
    form_class = DialectForm
    template_name = "add.html"
    def form_valid(self, form):
        language_pk = self.kwargs['lang_pk']
        self.object = form.save(commit=False)
        self.object.language_id = language_pk
        self.object.save()
        pk = self.object.pk
        instance = Dialect.objects.get(pk=pk)
        instance.modified_by = self.request.user.get_username()
        instance.save()
        return redirect("../../")

class dialect_delete(UserPassesTestMixin, DeleteView):
    def test_func(self):
        return self.request.user.groups.filter(name="Archivist").exists()
    def handle_no_permission(self):
        return redirect('/no-permission')
    model = Dialect
    success_url = '/languages/'

@login_required
@user_passes_test(is_member_of_archivist, login_url='/no-permission', redirect_field_name=None)
def dialect_instance_edit(request, pk):
    qs = get_object_or_404(DialectInstance, id=pk)
    if request.method == "POST":
        form = DialectInstanceForm(request.POST, instance=qs)
        if form.is_valid():
            qs.modified_by = request.user.get_username()
            qs.save()
            form.save()
            url = "/"
            if qs.document:
                url = "../../../documents/%s/" %qs.document.pk
            if qs.item:
                url = "../../../catalog/%s/" %qs.item.pk
            if qs.collaborator_native:
                url = "../../../collaborators/%s/" %qs.collaborator_native.pk
            if qs.collaborator_other:
                url = "../../../collaborators/%s/" %qs.collaborator_other.pk
            return redirect(url)
    else:
        form = DialectInstanceForm(instance=qs)
    return render(request, 'dialect_instance_edit.html', {'form': form})

@login_required
def collaborator_index(request):
    qs = Collaborator.objects.all()
    order_choice = request.GET.get("form_control_sort")
    name_contains_query = request.GET.get('name_contains')
    collection_contains_query = request.GET.get('collection_contains')
    native_languages_contains_query = request.GET.get('native_languages_contains')
    other_languages_contains_query = request.GET.get('other_languages_contains')

    order_choice_last = order_choice
    name_contains_query_last = ''
    collection_contains_query_last = ''
    native_languages_contains_query_last = ''
    other_languages_contains_query_last = ''


    if not is_member_of_archivist(request.user):
        qs = qs.exclude(anonymous = True)

    if is_valid_param(collection_contains_query):
        collection_contains_query_last = collection_contains_query

    if is_valid_param(native_languages_contains_query):
        qs = qs.filter(native_languages__name__icontains = native_languages_contains_query)
        native_languages_contains_query_last = native_languages_contains_query

    if is_valid_param(other_languages_contains_query):
        qs = qs.filter(other_languages__name__icontains = other_languages_contains_query)
        other_languages_contains_query_last = other_languages_contains_query

    anonymous_list = Collaborator.objects.none()
    if is_valid_param(name_contains_query):
        if ( name_contains_query == 'Anonymous' ) or ( name_contains_query == 'anonymous' ):
            anonymous_list = Collaborator.objects.filter(anonymous = True)
        qs = qs.filter(
            Q(name__icontains = name_contains_query) | Q(nickname__icontains = name_contains_query) | Q(other_names__icontains = name_contains_query)
            )

        name_contains_query_last = name_contains_query

    qs = qs.distinct()

    qs = qs.union(anonymous_list)

#    print(type(order_choice))
    if order_choice == "updated":
        qs = qs.order_by('-updated')
    else:
        qs = qs.order_by('name')

    results_count = qs.count()

    paginator = Paginator(qs, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'queryset': page_obj,
        'results_count' : results_count,
        'order_choice_last' : order_choice_last,
        'name_contains_query_last' : name_contains_query_last,
        'collection_contains_query_last' : collection_contains_query_last,
        'native_languages_contains_query_last' : native_languages_contains_query_last,
        'other_languages_contains_query_last' : other_languages_contains_query_last,
    }
    return render(request, 'collaborator_index.html', context)

@login_required
def collaborator_detail(request, pk):
    qs = Collaborator.objects.get(pk=pk)

    for native_language in qs.native_languages.all():
        native_dialect_i, created = DialectInstance.objects.get_or_create(collaborator_native=qs, language=native_language)
    native_dialect_info = zip(qs.native_languages.all().order_by('name'), DialectInstance.objects.filter(collaborator_native=qs).order_by('language__name'))

    for other_language in qs.other_languages.all():
        other_dialect_i, created = DialectInstance.objects.get_or_create(collaborator_other=qs, language=other_language)
    other_dialect_info = zip(qs.other_languages.all().order_by('name'), DialectInstance.objects.filter(collaborator_other=qs).order_by('language__name'))

    items_with_collaborator = Item.objects.filter(collaborator=qs)

    context = {
        'collaborator': qs,
        'native_dialect_info': native_dialect_info,
        'other_dialect_info': other_dialect_info,
        'items_with_collaborator': items_with_collaborator
    }

    return render(request, 'collaborator_detail.html', context)

@login_required
@user_passes_test(is_member_of_archivist, login_url='/no-permission', redirect_field_name=None)
def collaborator_edit(request, pk):
    qs = get_object_or_404(Collaborator, id=pk)
    if request.method == "POST":
        form = CollaboratorForm(request.POST, instance=qs)
        if form.is_valid():
            qs.modified_by = request.user.get_username()
            qs.save()
            form.save()
            return redirect("../")
    else:
        form = CollaboratorForm(instance=qs)
    return render(request, 'collaborator_edit.html', {'form': form})

@method_decorator([login_required], name='dispatch')
class collaborator_add(UserPassesTestMixin, FormView):
    def test_func(self):
        return self.request.user.groups.filter(name="Archivist").exists()
    def handle_no_permission(self):
        return redirect('/no-permission')
    form_class = CollaboratorForm
    template_name = "add.html"
    def form_valid(self, form):
        self.object = form.save()
        pk = self.object.pk
        instance = Collaborator.objects.get(pk=pk)
        instance.modified_by = self.request.user.get_username()
        instance.save()
        return redirect("../%s/" %pk )

class collaborator_delete(UserPassesTestMixin, DeleteView):
    def test_func(self):
        return self.request.user.groups.filter(name="Archivist").exists()
    def handle_no_permission(self):
        return redirect('/no-permission')
    model = Collaborator
    success_url = '/collaborators/'

@login_required
@user_passes_test(is_member_of_archivist, login_url='/no-permission', redirect_field_name=None)
def collaborator_role_edit(request, pk):
    qs = get_object_or_404(CollaboratorRole, id=pk)
    if request.method == "POST":
        form = CollaboratorRoleForm(request.POST, instance=qs)
        if form.is_valid():
            qs.modified_by = request.user.get_username()
            qs.save()
            form.save()
            if qs.document:
                url = "../../../documents/%s/" %qs.document.pk
            if qs.item:
                url = "../../../catalog/%s/" %qs.item.pk
            return redirect(url)
    else:
        form = CollaboratorRoleForm(instance=qs)
    return render(request, 'collaborator_role_edit.html', {'form': form})


class geographic_add(UserPassesTestMixin, FormView):
    def test_func(self):
        return self.request.user.groups.filter(name="Archivist").exists()
    def handle_no_permission(self):
        return redirect('/no-permission')
    form_class = GeographicForm
    template_name = "add.html"
    def form_valid(self, form):
        url_path = self.request.get_full_path()
        self.object = form.save()
        pk = self.object.pk
        instance = Geographic.objects.get(pk=pk)
        catalog_url_match = re.search('catalog/([0-9]{1,})', url_path, flags=re.I)
        if catalog_url_match:
            # print(catalog_url_match.groups(1))
            parent_item = Item.objects.get(id=catalog_url_match.groups(1)[0])
            instance.item = parent_item
        document_url_match = re.search('documents/([0-9]{1,})', url_path, flags=re.I)
        if document_url_match:
            # print(document_url_match.groups(1))
            parent_item = Document.objects.get(id=document_url_match.groups(1)[0])
            instance.document = parent_item


        instance.modified_by = self.request.user.get_username()
        instance.save()
        return redirect("../../")

@login_required
@user_passes_test(is_member_of_archivist, login_url='/no-permission', redirect_field_name=None)
def geographic_edit(request, pk):
    qs = get_object_or_404(Geographic, id=pk)
    if request.method == "POST":
        form = GeographicForm(request.POST, instance=qs)
        if form.is_valid():
            qs.modified_by = request.user.get_username()
            qs.save()
            form.save()
            if qs.document:
                url = "../../../documents/%s/" %qs.document.pk
            if qs.item:
                url = "../../../catalog/%s/" %qs.item.pk
            return redirect(url)
    else:
        form = GeographicForm(instance=qs)
    return render(request, 'geographic_edit.html', {'form': form})

class geographic_delete(UserPassesTestMixin, DeleteView):
    def test_func(self):
        return self.request.user.groups.filter(name="Archivist").exists()
    def handle_no_permission(self):
        return redirect('/no-permission')
    model = Geographic
    success_url = '/catalog/'





@login_required
def columns_export_index(request):
    qs = Columns_export.objects.all()

    context = {
        'queryset': qs,
    }
    return render(request, 'columns_export_index.html', context)

@login_required
def columns_export_detail(request, pk):
    qs = Columns_export.objects.get(pk=pk)
    context = {
        'columns_export': qs
    }
    return render(request, 'columns_export_detail.html', context)

@login_required
@user_passes_test(is_member_of_archivist, login_url='/no-permission', redirect_field_name=None)
def columns_export_edit(request, pk):
    qs = get_object_or_404(Columns_export, id=pk)
    if request.method == "POST":
        form = Columns_exportForm(request.POST, instance=qs)
        if form.is_valid():
            qs.modified_by = request.user.get_username()
            qs.save()
            form.save()
            return redirect("/export-columns/%s/" %pk)
    else:
        form = Columns_exportForm(instance=qs)
    return render(request, 'columns_export_edit.html', {'form': form})

class columns_export_add(UserPassesTestMixin, FormView):
    def test_func(self):
        return self.request.user.groups.filter(name="Archivist").exists()
    def handle_no_permission(self):
        return redirect('/no-permission')
    form_class = Columns_exportForm
    template_name = "add.html"
    def form_valid(self, form):
        self.object = form.save()
        pk = self.object.pk
        instance = Columns_export.objects.get(pk=pk)
        instance.modified_by = self.request.user.get_username()
        instance.save()
        return redirect("../%s/" %pk )




def is_valid_param(param):
    if ( param == 0 or param == [0] ) :
        return True
    else:
        return param != '' and param != 'None' and param is not None and param


def list_string_find_indices(the_list, search_string):
    matched_indices = []
    i = 0
    length = len(the_list)
    while i < length:
        if the_list[i]:
            if re.search(search_string, the_list[i], flags=re.I):
                matched_indices.append(i)
        i += 1
    return(matched_indices)


def date_processing(date): #this is probably deprecated
    if isinstance(date, str): #for cases like a date field imported from a spreadsheet
        cleaned_date = re.sub(r"([0-9]{2})/([0-9]{2})/([0-9]{4})", r"\3-\1-\2", date)
    elif ( isinstance(date, list) ) or ( isinstance(date, tuple) ): # for cases when 3 part date is passed
        pass
    return cleaned_date # will replace with a dic of the 3 char fields and start/end dates

##### reverse choices was here

def import_field(request, model_field, human_fields, headers, row, object_instance, model, choices=None, multiselect=False, yesno=False, validate_glottocode=False, validate_coord=False):
    # model_field: string, human_fields: tuple, headers: list, row: list, object_instance: object instance,
    # choices: tuple of tuples from django models
    automate_glottocode_flag = False
    object_instance_name = ''
    if model == "Item":
        object_instance_name = 'Item: ' + str(object_instance.catalog_number)
    elif ( model == "Document" ) or ( model == "Item document" ):
        if is_valid_param(object_instance.filename):
            object_instance_name = 'Document: ' + str(object_instance.filename)
        elif is_valid_param(object_instance.title):
            object_instance_name = 'Document with title: ' + str(object_instance.title)
        else:
            object_instance_name = 'Document with unique ID: ' + str(object_instance.id)
    elif ( model == "Collaborator native" ) or ( model == "Collaborator other" ) or ( model == "Collaborator" ):
        object_instance_name = 'Collaborator: ' + str(object_instance.name) + ' (' + str(object_instance.collaborator_id) + ')'
    elif ( model == "Language" ):
        object_instance_name = 'Language: ' + str(object_instance.name)

    for human_field in human_fields:
        human_field_indeces = list_string_find_indices(headers, human_field)
        if is_valid_param(human_field_indeces):
            model_field_value = row[human_field_indeces[0]]
            stripped_human_field = human_field.replace('^', '').replace('$', '')
            if is_valid_param(model_field_value):
                old_value = getattr(object_instance, model_field)
                if yesno:
                    if re.search(r"yes", model_field_value, flags=re.I):
                        model_field_value = True
                    elif re.search(r"no", model_field_value, flags=re.I):
                        model_field_value = False
                if choices:
                    if multiselect:
                        model_field_value = re.sub(r"\n", r", ", model_field_value)
                        model_field_value = re.sub(r"[,\s]*$", r"", model_field_value)
                        model_field_value = reverse_lookup_choices(choices, model_field_value, strict=True)
                        if model_field_value is None:
                            messages.warning(request, object_instance_name + " was not added/updated (all changes were reverted): " + stripped_human_field + " has an invalid value")
                            return False
                    else:
                        model_field_value = reverse_lookup_choices(choices, model_field_value)
                        computer_readable_list = [choice[0] for choice in choices]
                        if not model_field_value in computer_readable_list:
                            stripped_human_field = human_field.replace('^', '').replace('$', '')
                            messages.warning(request, object_instance_name + " was not added/updated (all changes were reverted): " + stripped_human_field + " has an invalid value")
                            return False
                if validate_glottocode == "single":
                    test_value = str(model_field_value).strip()
                    if len(test_value) != 8 or not re.match(r'^.*\d{4}$', test_value):
                        messages.warning(request, object_instance_name + " was not added/updated (all changes were reverted): " + stripped_human_field + " has an invalid value")
                        return False
                if validate_glottocode == "multiple":
                    test_value = str(model_field_value).strip()
                    glottocodes = [code.strip() for code in test_value.split(',')]
                    # Validate each glottocode
                    for glottocode in glottocodes:
                        if len(glottocode) != 8 or not re.match(r'^.*\d{4}$', glottocode):
                            messages.warning(request, object_instance_name + " was not added/updated (all changes were reverted): " + stripped_human_field + " has an invalid value")
                            return False
                if validate_coord:
                    if isinstance(model_field_value, (int, float)):
                        pass  # model_field_value is an int or float, so it's valid
                    elif isinstance(model_field_value, str) and model_field_value.isdigit():
                        pass  # model_field_value is a string of digits, so it's valid
                    else:
                        messages.warning(request, object_instance_name + " was not added/updated (all changes were reverted): " + stripped_human_field + " has an invalid value")
                        return False
                setattr(object_instance, model_field, model_field_value)
                if model_field == 'family':
                    print('family after: ' + str(object_instance.family))
                try:
                    object_instance.clean()
                except:
                    #object_instance.full_clean()
                    setattr(object_instance, model_field, old_value)
                    stripped_human_field = human_field.replace('^', '').replace('$', '')
                    messages.warning(request, object_instance_name + " was not added/updated (all changes were reverted): " + stripped_human_field + " has an invalid value")
                    return False
            else:
                automate_glottocode_flag = True
        else:
            automate_glottocode_flag = True
        if automate_glottocode_flag:
            if validate_glottocode:
                if human_fields[0] == '^Glottocode$':
                    automate_glottocode('glottocode', 'name', object_instance)
                elif human_fields[0] == '^Family glottocode$' or human_fields[0] == '^Family$':
                    automate_glottocode('family_id', 'family', object_instance)
                elif human_fields[0] == '^Primary subgroup glottocode$':
                    automate_glottocode('pri_subgroup_id', 'pri_subgroup', object_instance)
                elif human_fields[0] == '^Secondary subgroup glottocode$':
                    automate_glottocode('sec_subgroup_id', 'sec_subgroup', object_instance)
                elif human_fields[0] == '^Dialect glottocodes$':
                    automate_glottocode('dialects_ids', 'dialects', object_instance)
    return True



def import_child_field(request, parent_name_field, child_model_field, human_fields, headers, row, object_instance, model, num=False, choices=None, multiselect=False, yesno=False):

    object_instance_name = ''
    if model == "Item":
        object_instance_name = 'Item: ' + str(object_instance.catalog_number)
    elif ( model == "Document" ) or ( model == "Item document" ):
        if is_valid_param(object_instance.filename):
            object_instance_name = 'Document: ' + str(object_instance.filename)
        elif is_valid_param(object_instance.title):
            object_instance_name = 'Document with title: ' + str(object_instance.title)
        else:
            object_instance_name = 'Document with unique ID: ' + str(object_instance.id)
    elif ( model == "Collaborator native" ) or ( model == "Collaborator other" ) or ( model == "Collaborator" ):
        object_instance_name = 'Collaborator: ' + str(object_instance.name) + ' (' + str(object_instance.collaborator_id) + ')'

    valid_values = [False] * len(human_fields)
    child_model_field_values = [None] * len(human_fields)
    for i, human_field in enumerate(human_fields):
        human_field_indeces = list_string_find_indices(headers, human_field)
        if is_valid_param(human_field_indeces):
            child_model_field_values[i] = row[human_field_indeces[0]]
            if is_valid_param(child_model_field_values[i]):
                valid_values[i] = True
                if num:
                    try:
                        child_model_field_values[i] = float(child_model_field_values[i])
                    except:
                          stripped_human_field = 'Latitude/Longitude'
                          messages.warning(request, object_instance_name + " was not added/updated (all changes were reverted): " + stripped_human_field + " has an invalid value")
                          return False



    if all(valid_values):

        old_geographics = []
        for old_geographic in Geographic.objects.filter(**{parent_name_field: object_instance}):
            old_geographics.append(model_to_dict(old_geographic))
            old_geographic.delete()

        new_geographic = {'lat': child_model_field_values[0],
                          'long': child_model_field_values[1]
        }
        if model == "Item":
            new_geographic['item'] = object_instance
        elif ( model == "Document" ) or ( model == "Item document" ):
            new_geographic['document'] = object_instance

        try:
            new_geographic_instance = Geographic.objects.create(**new_geographic)
        except:
            for old_geographic in old_geographics:
                new_geographic_instance = Geographic.objects.create(**old_geographic)
            stripped_human_field = 'Latitude/Longitude'
            messages.warning(request, object_instance_name + " was not added/updated (all changes were reverted): " + stripped_human_field + " has an invalid value")
            return False
    return True





def import_language_field(request, headers, row, object_instance, model, document_prefix=''):
    # returns True for success and False for failure. Failure = an error with row info that leads to not adding that row
    # headers: list, row: list, object_instance: object instance, model: sting 'Item' 'Document' or 'Collaborator'
    return_object = None

    object_instance_name = ''
    if model == "Item":
        object_instance_name = 'Item: ' + str(object_instance.catalog_number)
    elif ( model == "Document" ) or ( model == "Item document" ):
        if is_valid_param(object_instance.filename):
            object_instance_name = 'Document: ' + str(object_instance.filename)
        elif is_valid_param(object_instance.title):
            object_instance_name = 'Document with title: ' + str(object_instance.title)
        else:
            object_instance_name = 'Document with unique ID: ' + str(object_instance.id)
    elif ( model == "Collaborator native" ) or ( model == "Collaborator other" ) or ( model == "Collaborator" ):
        object_instance_name = 'Collaborator: ' + str(object_instance.name) + ' (' + str(object_instance.collaborator_id) + ')'

    # language and dialect, a manytomany one, with a double foreign key one
    if model == "Item document":
        iso_indexes = list_string_find_indices(headers, document_prefix + ' ISO(\s)?([0-9])*$')
        if not is_valid_param(iso_indexes):
            iso_indexes = list_string_find_indices(headers, document_prefix + ' ISO Indicator(\s)?([0-9])*$')
    else:
        iso_indexes = list_string_find_indices(headers,'^ISO(\s)?([0-9])*$')
        if not is_valid_param(iso_indexes):
            iso_indexes = list_string_find_indices(headers,'^ISO Indicator(\s)?([0-9])*$')

    if ( model == "Item" ) or ( model == "Language") or ( model == "Document"):
        language_indexes = list_string_find_indices(headers,'^Language Name(\s)?([0-9])*$')
    elif model == "Item document":
        language_indexes = list_string_find_indices(headers, document_prefix + ' Language Name(\s)?([0-9])*$')
    elif model == "Collaborator native":
        language_indexes = list_string_find_indices(headers,'^Native/First Languages?(\s)?([0-9])*$')
    elif model == "Collaborator other":
        language_indexes = list_string_find_indices(headers,'^Other Languages?( Spoken)?(\s)?([0-9])*$')

    if ( is_valid_param(iso_indexes) or is_valid_param(language_indexes) ):
        indeces_list = [iso_indexes, language_indexes]
        list_len = [len(i) for i in indeces_list]
        max_list_len = max(list_len)

        iso_indexes.extend( [None] * ( max_list_len - len(iso_indexes) ) )
        language_indexes.extend( [None] * ( max_list_len - len(language_indexes) ) )

        if ( model == "Item" ) or ( model == "Document" ):
            dialect_indexes = list_string_find_indices(headers,'^Dialect(\s)?([0-9])*$')
        elif model == "Item document":
            dialect_indexes = list_string_find_indices(headers, document_prefix + ' Dialect(\s)?([0-9])*$')
        elif model == "Collaborator native":
            dialect_indexes = list_string_find_indices(headers,'^Native/First Languages? Dialect(\s)?([0-9])*$')
        elif model == "Collaborator other":
            dialect_indexes = list_string_find_indices(headers,'^Other Languages?( Spoken)? Dialect(\s)?([0-9])*$')
        elif model == "Language":
            dialect_indexes = list_string_find_indices(headers,'^Dialects(\s)?([0-9])*$')

        if is_valid_param(dialect_indexes):
            dialect_indexes.extend( [None] * ( len(iso_indexes) - len(dialect_indexes) ) ) # if import file has more Language Name fields than Dialect fields, make dummy dialect fields
        else:
            dialect_indexes = [None] * len(iso_indexes) # if import file has Language Name fields but no Dialect fields, make dummy dialect fields

        if ( model == "Collaborator native" ):
            object_instance.native_languages.clear()
        elif ( model == "Collaborator other" ):
            object_instance.other_languages.clear()
        elif model != "Language":
            object_instance.language.clear()

        indexes = zip(iso_indexes, language_indexes, dialect_indexes)
        for iso_index, language_index, dialect_index in indexes:
            iso_value = ''
            language_object_with_iso_value = ''
            language_value = ''
            language_object_with_language_value = ''

            if is_valid_param(iso_index):
                iso_value = row[iso_index]
                if is_valid_param(iso_value):
                    try:
                        language_object_with_iso_value = Language.objects.get(iso=iso_value)
                    except:
                        pass
                else:
                    iso_value = '' # reset this value to '' to help create object_instance_name for the case of language_import

            if is_valid_param(language_index):
                language_value = row[language_index]
                if is_valid_param(language_value):
                    try:
                        language_object_with_language_value = Language.objects.get(name=language_value)
                    except:
                        pass
                else:
                    language_value = '' # reset this value to '' to help create object_instance_name for the case of language_import

            if model == "Language":
                object_instance_name = 'Language ' + language_value + ' (ISO Indicator: ' + iso_value + ')'


            if is_valid_param(iso_value):
                if is_valid_param(language_object_with_iso_value): # thus ISO is defined and exists in database
                    if is_valid_param(language_value):
                        if is_valid_param(language_object_with_language_value):
                            # thus Language Name is defined and exists in database
#                            print(language_object_with_iso_value)
#                            print(language_object_with_language_value)
                            if language_object_with_iso_value.pk == language_object_with_language_value.pk:
                                # 1) ISO and Language name refer to the same entry in the database, so all good
                                if model == "Language":
                                    #language_object_with_iso_value.language_dialects.delete()
                                    Dialect.objects.filter(language=language_object_with_iso_value).delete()
                                    return_object = language_object_with_iso_value
                                elif ( model == "Collaborator native" ):
                                    object_instance.native_languages.add(language_object_with_iso_value)
                                elif ( model == "Collaborator other" ):
                                    object_instance.other_languages.add(language_object_with_iso_value)
                                else:
                                    object_instance.language.add(language_object_with_iso_value)

                                if model == "Item":
                                    dialect, created = DialectInstance.objects.get_or_create(item=object_instance, language=language_object_with_iso_value, defaults={'modified_by': request.user.get_username()})
                                elif ( model == "Document") or ( model == "Item document" ):
                                    dialect, created = DialectInstance.objects.get_or_create(document=object_instance, language=language_object_with_iso_value, defaults={'modified_by': request.user.get_username()})
                                elif model == "Collaborator native":
                                    dialect, created = DialectInstance.objects.get_or_create(collaborator_native=object_instance, language=language_object_with_iso_value, defaults={'modified_by': request.user.get_username()})
                                elif model == "Collaborator other":
                                    dialect, created = DialectInstance.objects.get_or_create(collaborator_other=object_instance, language=language_object_with_iso_value, defaults={'modified_by': request.user.get_username()})
                                    #create dialect object based on language, regardless of value

                                if is_valid_param(dialect_index):
                                    dialect_value = row[dialect_index]
                                    if is_valid_param(dialect_value):
                                        dialect_value = re.sub(r"\n", r", ", dialect_value)
                                        dialect_value = re.sub(r"[,\s]*$", r"", dialect_value)
                                        dialect_values = dialect_value.strip().split(',')
                                        dialect_values = [i.strip() for i in dialect_values]
                                        if model == "Language":
                                            for each_dialect in dialect_values:
                                                dialect_object_with_dialect_value, created = Dialect.objects.get_or_create(name=each_dialect, language=language_object_with_iso_value)
                                                #language_object_with_iso_value.language_dialects.add(dialect_object_with_dialect_value)
                                            # return_object = language_object_with_iso_value
                                        else:
                                            dialect.name.clear()
                                            for each_dialect in dialect_values:
                                                accepted_dialects = Dialect.objects.filter(language=language_object_with_iso_value).values_list('name', flat=True)
                                                if each_dialect in list(accepted_dialects):
                                                    dialect_object_with_dialect_value = Dialect.objects.get(name=each_dialect, language=language_object_with_iso_value)
                                                    dialect.name.add(dialect_object_with_dialect_value)

                                                else:
                                                    messages.warning(request, object_instance_name + " was not added/updated (all changes were reverted): Dialect provided (" + str(each_dialect) + ") is not an available option for the corresponding language in the database")
                                                    return False, return_object

                            else: # 2) ISO and Language Name correspond to different entries
                                messages.warning(request, object_instance_name + " was not added/updated (all changes were reverted): ISO Indicator and Language Name provided correspond to different languages in the database")
                                return False, return_object
                        else: # 3) thus Language Name is defined but does not exist in database
                            messages.warning(request, object_instance_name + " was not added/updated (all changes were reverted): ISO Indicator is associated with a language in the database but Language Name is not")
                            return False, return_object

                    else: # 4) thus Language Name is not defined
                        if model == "Language":
                            #language_object_with_iso_value.language_dialects.clear()
                            Dialect.objects.filter(language=language_object_with_iso_value).delete()
                            return_object = language_object_with_iso_value
                        elif ( model == "Collaborator native" ):
                            object_instance.native_languages.add(language_object_with_iso_value)
                        elif ( model == "Collaborator other" ):
                            object_instance.other_languages.add(language_object_with_iso_value)
                        else:
                            object_instance.language.add(language_object_with_iso_value)

                        if model == "Item":
                            dialect, created = DialectInstance.objects.get_or_create(item=object_instance, language=language_object_with_iso_value, defaults={'modified_by': request.user.get_username()})
                        elif ( model == "Document") or ( model == "Item document" ):
                            dialect, created = DialectInstance.objects.get_or_create(document=object_instance, language=language_object_with_iso_value, defaults={'modified_by': request.user.get_username()})
                        elif model == "Collaborator native":
                            dialect, created = DialectInstance.objects.get_or_create(collaborator_native=object_instance, language=language_object_with_iso_value, defaults={'modified_by': request.user.get_username()})
                        elif model == "Collaborator other":
                            dialect, created = DialectInstance.objects.get_or_create(collaborator_other=object_instance, language=language_object_with_iso_value, defaults={'modified_by': request.user.get_username()})
                            #create dialect object based on language, regardless of value
                        if is_valid_param(dialect_index):
                            dialect_value = row[dialect_index]
                            if is_valid_param(dialect_value):
                                dialect_value = re.sub(r"\n", r", ", dialect_value)
                                dialect_value = re.sub(r"[,\s]*$", r"", dialect_value)
                                dialect_values = dialect_value.strip().split(',')
                                dialect_values = [i.strip() for i in dialect_values]
                                if model == "Language":
                                    for each_dialect in dialect_values:
                                        dialect_object_with_dialect_value, created = Dialect.objects.get_or_create(name=each_dialect, language=language_object_with_iso_value)
                                        #language_object_with_iso_value.language_dialects.add(dialect_object_with_dialect_value)
                                else:
                                    dialect.name.clear()
                                    for each_dialect in dialect_values:
                                        accepted_dialects = Dialect.objects.filter(language=language_object_with_iso_value).values_list('name', flat=True)
                                        if each_dialect in list(accepted_dialects):
                                            dialect_object_with_dialect_value = Dialect.objects.get(name=each_dialect, language=language_object_with_iso_value)
                                            dialect.name.add(dialect_object_with_dialect_value)
                                        else:
                                            messages.warning(request, object_instance_name + " was not added/updated (all changes were reverted): Dialect provided (" + str(each_dialect) + ") is not an available option for the corresponding language in the database")
                                            return False, return_object

                else: # thus ISO is defined but does not exist in database
                    if is_valid_param(language_value):
                        if is_valid_param(language_object_with_language_value):
                            # 5) thus Language Name is defined and exists in database
                            messages.warning(request, object_instance_name + " was not added/updated (all changes were reverted): Language Name is associated with a language in the database but ISO Indicator is not")
                            return False, return_object

                        else: # 6) thus Language Name is defined but does not exist in database
                            language, created = Language.objects.get_or_create(iso=iso_value,
                                defaults={'name': language_value},)
                            if model == "Language":
                                messages.info(request, 'A new language entry was created (ISO indicator: ' + language.iso + ', Name: ' + language.name + ')')
                                return_object = language
                            elif ( model == "Collaborator native" ):
                                object_instance.native_languages.add(language)
                                messages.info(request, 'A new language entry was created (ISO indicator: ' + language.iso + ', Name: ' + language.name + '), and it was added to ' + object_instance_name)
                            elif ( model == "Collaborator other" ):
                                object_instance.other_languages.add(language)
                                messages.info(request, 'A new language entry was created (ISO indicator: ' + language.iso + ', Name: ' + language.name + '), and it was added to ' + object_instance_name)
                            else:
                                object_instance.language.add(language)
                                messages.info(request, 'A new language entry was created (ISO indicator: ' + language.iso + ', Name: ' + language.name + '), and it was added to ' + object_instance_name)
                            if model == "Item":
                                dialect, created = DialectInstance.objects.get_or_create(item=object_instance, language=language, defaults={'modified_by': request.user.get_username()})
                            elif ( model == "Document") or ( model == "Item document" ):
                                dialect, created = DialectInstance.objects.get_or_create(document=object_instance, language=language, defaults={'modified_by': request.user.get_username()})
                            elif model == "Collaborator native":
                                dialect, created = DialectInstance.objects.get_or_create(collaborator_native=object_instance, language=language, defaults={'modified_by': request.user.get_username()})
                            elif model == "Collaborator other":
                                dialect, created = DialectInstance.objects.get_or_create(collaborator_other=object_instance, language=language, defaults={'modified_by': request.user.get_username()})
                                #create dialect object based on language, regardless of value
                            if is_valid_param(dialect_index):
                                dialect_value = row[dialect_index]
                                if is_valid_param(dialect_value):
                                    dialect_value = re.sub(r"\n", r", ", dialect_value)
                                    dialect_value = re.sub(r"[,\s]*$", r"", dialect_value)
                                    dialect_values = dialect_value.strip().split(',')
                                    dialect_values = [i.strip() for i in dialect_values]
                                    if model == "Language":
                                        for each_dialect in dialect_values:
                                            dialect_object_with_dialect_value, created = Dialect.objects.get_or_create(name=each_dialect, language=language)
                                            #language.language_dialects.add(dialect_object_with_dialect_value)
                                    else:
                                        dialect.name.clear()
                                        for each_dialect in dialect_values:
                                            dialect_object_with_dialect_value, created = Dialect.objects.get_or_create(name=each_dialect, language=language)
                                            dialect.name.add(dialect_object_with_dialect_value)
                                            messages.info(request, 'A new dialect (' + dialect_object_with_dialect_value.name + ') was created for ' + language.name + ' (ISO indicator: ' + language.iso + '), and it was added to ' + object_instance_name)

                    else: # 7) thus Language Name is not defined
                        messages.warning(request, object_instance_name + " was not added/updated (all changes were reverted): ISO indicator is not associated with a language in the database, and Language Name is not defined. A new language cannot be created without a Language Name.")
                        return False, return_object

            else: # thus ISO is not defined
                if is_valid_param(language_value):
                    if is_valid_param(language_object_with_language_value):
                        # 8) thus Language Name is defined and exists in database
                        if model == "Language":
                            #language_object_with_language_value.language_dialects.clear()
                            Dialect.objects.filter(language=language_object_with_language_value).delete()
                            return_object = language_object_with_language_value
                        elif ( model == "Collaborator native" ):
                            object_instance.native_languages.add(language_object_with_language_value)
                        elif ( model == "Collaborator other" ):
                            object_instance.other_languages.add(language_object_with_language_value)
                        else:
                            object_instance.language.add(language_object_with_language_value)
                        if model == "Item":
                            dialect, created = DialectInstance.objects.get_or_create(item=object_instance, language=language_object_with_language_value, defaults={'modified_by': request.user.get_username()})
                        elif ( model == "Document") or ( model == "Item document" ):
                            dialect, created = DialectInstance.objects.get_or_create(document=object_instance, language=language_object_with_language_value, defaults={'modified_by': request.user.get_username()})
                        elif model == "Collaborator native":
                            dialect, created = DialectInstance.objects.get_or_create(collaborator_native=object_instance, language=language_object_with_language_value, defaults={'modified_by': request.user.get_username()})
                        elif model == "Collaborator other":
                            dialect, created = DialectInstance.objects.get_or_create(collaborator_other=object_instance, language=language_object_with_language_value, defaults={'modified_by': request.user.get_username()})
                            #create dialect object based on language, regardless of value
                        if is_valid_param(dialect_index):
                            dialect_value = row[dialect_index]
                            if is_valid_param(dialect_value):
                                dialect_value = re.sub(r"\n", r", ", dialect_value)
                                dialect_value = re.sub(r"[,\s]*$", r"", dialect_value)
                                dialect_values = dialect_value.strip().split(',')
                                dialect_values = [i.strip() for i in dialect_values]
                                if model == "Language":
                                    for each_dialect in dialect_values:
                                        dialect_object_with_dialect_value, created = Dialect.objects.get_or_create(name=each_dialect, language=language_object_with_language_value)
                                        #language_object_with_language_value.language_dialects.add(dialect_object_with_dialect_value)
                                else:
                                    dialect.name.clear()
                                    for each_dialect in dialect_values:
                                        accepted_dialects = Dialect.objects.filter(language=language_object_with_language_value).values_list('name', flat=True)
                                        if each_dialect in list(accepted_dialects):
                                            dialect_object_with_dialect_value = Dialect.objects.get(name=each_dialect, language=language_object_with_language_value)
                                            dialect.name.add(dialect_object_with_dialect_value)
                                        else:
                                            messages.warning(request, object_instance_name + " was not added/updated (all changes were reverted): Dialect provided (" + str(each_dialect) + ") is not an available option for the corresponding language in the database")
                                            return False, return_object

                    else: # 9) thus Language Name is defined but does not exist in database
                        messages.warning(request, object_instance_name + " was not added/updated (all changes were reverted): Language Name (" + language_value + ") is not associated with a language in the database, and ISO indicator is not defined. A new language cannot be created without an ISO indictor")
                        return False, return_object

                else: # 10) thus Language Name is not defined
                    pass

    return True, return_object



def import_date_field(model_field, human_fields, headers, row, object_instance):
    # model_field: string, human_fields: tuple, headers: list, row: list, object_instance: object instance
    if len(human_fields) == 1:
        date_value = ''
        human_field_indeces = list_string_find_indices(headers, human_fields[0])
        if not is_valid_param(human_field_indeces):
            return True
        else:
            date_value = row[human_field_indeces[0]]
            if not is_valid_param(date_value):
                return True
            elif is_valid_param(date_value):
                date_re = re.search(r"([0-9]{4})([0-9]{4})", str(date_value), flags=re.I)
                if date_re:
                    setattr(object_instance, model_field, date_re.group(1)+'-'+date_re.group(2))
#                    print(object_instance.catalog_number + " got to A")
                    return True
                date_re = re.search(r"([0-9]+[a-z]{2} century\?*)", str(date_value), flags=re.I)
                if date_re:
                    setattr(object_instance, model_field, date_re.group(1))
#                    print(object_instance.catalog_number + " got to B")
                    return True
                date_re = re.search(r"([0-9]{4})['’]*(s\?*)", str(date_value), flags=re.I)
                if date_re:
                    setattr(object_instance, model_field, date_re.group(1)+date_re.group(2))
#                    print(object_instance.catalog_number + " got to C")
                    return True
                if not re.search(r"([0-9]{4})", str(date_value), flags=re.I): # no date assignment because there is no 4-digit year
                    return False
                date_re = re.search(r"^([0-9]{4}\?)$", str(date_value), flags=re.I)
                if date_re:
                    setattr(object_instance, model_field, date_re.group(1))
#                    print(object_instance.catalog_number + " got to C")
                    return True
                date_re = re.search(r"(ca *[0-9]{4})", str(date_value), flags=re.I)
                if date_re:
                    setattr(object_instance, model_field, date_re.group(1))
#                    print(object_instance.catalog_number + " got to D")
                    return True
                #######################befores and afters
                date_re = re.search(r"([0-9]{1,2}/[0-9]{1,2}/[0-9]{4}-[0-9]{1,2}/[0-9]{1,2}/[0-9]{4})", str(date_value), flags=re.I)
                if date_re:
                    setattr(object_instance, model_field, date_re.group(1))
#                    print(object_instance.catalog_number + " got to D")
                    return True
                date_re = re.search(r"([0-9]{1,2}/[0-9]{4}-[0-9]{1,2}/[0-9]{4})", str(date_value), flags=re.I)
                if date_re:
                    setattr(object_instance, model_field, date_re.group(1))
#                    print(object_instance.catalog_number + " got to E")
                    return True
                date_re = re.search(r"^([0-9]{4})$", str(date_value), flags=re.I)
                if date_re:
                    setattr(object_instance, model_field, date_re.group(1))
#                    print(object_instance.catalog_number + " got to F")
                    return True
                date_re = re.search(r"^([0-9]{4}-[0-9]{4})$", str(date_value), flags=re.I)
                if date_re:
                    setattr(object_instance, model_field, date_re.group(1))
#                    print(object_instance.catalog_number + " got to G")
                    return True

                date_value = reverse_lookup_choices(MONTH_CHOICES, str(date_value)) # convert alpha months to numeric
                date_value = re.sub(r"^(\d)([\s\D])", r"0\g<1>\g<2>", date_value) # add leading zeros to an single digit number, part 1 (exception for beginning of line)
                date_value = re.sub(r"([\s\D])(\d)([\s\D])", r"\g<1>0\g<2>\g<3>", date_value) # add leading zeros to an single digit number, part 2 (main)
                date_value = re.sub(r"([\s\D])(\d)$", r"\g<1>0\g<2>" r"\1\\0\2", date_value) # add leading zeros to an single digit number, part 3 (exception for end of line)

#                print('here' + date_value)
                if not ( re.search(r"^([0-9]{2})[\s\D]", str(date_value), flags=re.I) or re.search(r"[\s\D]([0-9]{2})[\s\D]", str(date_value), flags=re.I) or re.search(r"[\s\D]([0-9]{2})$", str(date_value), flags=re.I) ):
                    if re.search(r"[a-z?]", date_value, flags=re.I): # if it's only years, and if there are any words or question marks (i.e. weirdness), this date is invalid
                        return False
                    date_re = re.sub(r"[,\;\-]", ",", str(date_value))
                    date_re = re.sub(" ", "", date_re)
                    date_re = date_re.split(",")
                    setattr(object_instance, model_field, min(date_re)+'-'+max(date_re))
#                    print(object_instance.catalog_number + " got to H")
                    return True

                date_value = re.sub(r"([0-9]{4})[ ,\;\-]*([0-9]{4})", r"\1-\2", str(date_value))
                date_value = re.sub(r"([0-9]{1,2})[a-z]+", r"\1", date_value) # remove any "st" "nd" "rd" "th" from numbers
                date_value = re.sub(",", "", date_value) # get rid of commas that might be used in a verbose date format
                date_value = re.sub(" ", "/", date_value) # convert spaces to date format (spaces from a verbose date format)
                if re.search(r"[a-z?]", date_value, flags=re.I): # at this point, if there are any words left (or question marks), this date is invalid
                    return False

                if len(re.findall("-", date_value)) <= 1:
                    if len(re.findall("-", date_value)) == 1: # if date has range
                        date_1 = re.sub(r"([0-9]+)-([0-9]+)", r"\1", date_value) # make first date in range
                        date_2 = re.sub(r"([0-9]+)-([0-9]+)", r"\2", date_value) # make second date in range
                        date_value = date_1 + '-' + date_2
                    setattr(object_instance, model_field, date_value)
#                    print(object_instance.catalog_number + " got to I")
                    return True


    elif len(human_fields) == 3:
        date_year_value = ''
        date_month_value = ''
        date_day_value = ''

        human_field_input_year_indeces = list_string_find_indices(human_fields, 'year')
        if is_valid_param(human_field_input_year_indeces):
            human_field_input_year = human_fields[human_field_input_year_indeces[0]]
            human_field_year_index = list_string_find_indices(headers, human_field_input_year)
            if is_valid_param(human_field_year_index):
                date_year_value = str(row[human_field_year_index[0]])

        human_field_input_month_indeces = list_string_find_indices(human_fields, 'month')
        if is_valid_param(human_field_input_month_indeces):
            human_field_input_month = human_fields[human_field_input_month_indeces[0]]
            human_field_month_index = list_string_find_indices(headers, human_field_input_month)
            if is_valid_param(human_field_month_index):
                date_month_value = str(row[human_field_month_index[0]])

        human_field_input_day_indeces = list_string_find_indices(human_fields, 'day')
        if is_valid_param(human_field_input_day_indeces):
            human_field_input_day = human_fields[human_field_input_day_indeces[0]]
            human_field_day_index = list_string_find_indices(headers, human_field_input_day)
            if is_valid_param(human_field_day_index):
                date_day_value = str(row[human_field_day_index[0]])

        if not ( is_valid_param(date_year_value) or is_valid_param(date_month_value) or is_valid_param(date_day_value) ):
            return True
        elif is_valid_param(date_year_value): # check if year value is valid
            date_re = re.search(r"([0-9]{8})", str(date_year_value), flags=re.I)
            if date_re:
                setattr(object_instance, model_field, date_re.group(0)[:-4]+'-'+date_re.group(0)[-4:])
#                print(object_instance.catalog_number + "got to A")
                return True
            date_re = re.search(r"([0-9]+[a-z]{2} century\?*)", str(date_year_value), flags=re.I)
            if date_re:
                setattr(object_instance, model_field, date_re.group(0))
#                 print(object_instance.catalog_number + "got to B")
                return True
            if not re.search(r"([0-9]{4})", str(date_year_value), flags=re.I): # no date assignment because there is no 4-digit year
                return False
            date_re = re.search(r"([0-9]{4})['’]*(s\?*)", str(date_year_value), flags=re.I)
            if date_re:
                setattr(object_instance, model_field, date_re.group(1)+date_re.group(2))
#                print(object_instance.catalog_number + "got to C")
                return True
            date_re = re.search(r"^([0-9]{4}\?)$", str(date_year_value), flags=re.I)
            if date_re:
                setattr(object_instance, model_field, date_re.group(1))
#                print(object_instance.catalog_number + "got to D")
                return True
            date_re = re.search(r"(ca *[0-9]{4})", str(date_year_value), flags=re.I)
            if date_re:
                setattr(object_instance, model_field, date_re.group(1))
#                 print(object_instance.catalog_number + "got to E")
                return True
            #######################befores and afters
            if not re.search(r"([0-9]{1,2})", str(date_year_value), flags=re.I):
                date_re = re.sub(r"[,\;\-]", ",", str(date_year_value))
                date_re = re.sub(" ", "", date_re)
                date_re = date_re.split(",")
                setattr(object_instance, model_field, min(date_re)+'-'+max(date_re))
#                print(object_instance.catalog_number + "got to F")
                return True

            date_year_value = re.sub(r"[,\;\-]", "-", str(date_year_value))
            date_year_value = re.sub(" ", "", date_year_value)


            if is_valid_param(date_month_value):
                date_month_value = reverse_lookup_choices(MONTH_CHOICES, date_month_value)
                date_month_value = re.sub(r"\D(\d)\D", r"0\1", date_month_value) # add leading zeros to an single digit number
                date_month_value = re.sub(r"[,\;\-]", "-", str(date_month_value))
                date_month_value = re.sub(" ", "", date_month_value)

            if is_valid_param(date_day_value):
                date_day_value = re.sub(r"[a-z]+", "", str(date_day_value))
                date_day_value = re.sub(r"\D(\d)\D", r"0\1", date_day_value) # add leading zeros to an single digit number
                date_day_value = re.sub(r"[,\;\-]", "-", date_day_value)
                date_day_value = re.sub(" ", "", date_day_value)

            date_year_1 = ''
            date_year_2 = ''
            date_month_1 = ''
            date_month_2 = ''
            date_day_1 = ''
            date_day_2 = ''

#            print(object_instance.catalog_number + " " + str(date_year_value))
            date_re = re.search(r"(^[0-9]{4}$)", str(date_year_value), flags=re.I)
            if date_re:
                date_year_1 = date_re.group(1)
                date_year_2 = date_re.group(1)

            date_re = re.search(r"([0-9]{4})-([0-9]{4})", str(date_year_value), flags=re.I)
            if date_re:
                date_year_1 = date_re.group(1)
                date_year_2 = date_re.group(2)

            if is_valid_param(date_month_value):

                date_re = re.search(r"^([0-9]{1,2})$", str(date_month_value), flags=re.I)
                if date_re:
                    date_month_1 = date_re.group(1)
                    date_month_1 = ('0' + str(date_month_1))[-2:]
                    date_month_2 = date_month_1

                date_re = re.search(r"([0-9]{1,2})-([0-9]{1,2})", str(date_month_value), flags=re.I)
                if date_re:
                    date_month_1 = date_re.group(1)
                    date_month_1 = ('0' + str(date_month_1))[-2:]
                    date_month_2 = date_re.group(2)
                    date_month_2 = ('0' + str(date_month_2))[-2:]

            if is_valid_param(date_day_value):

                date_re = re.search(r"^([0-9]{1,2})$", str(date_day_value), flags=re.I)
                if date_re:
                    date_day_1 = date_re.group(1)
                    date_day_1 = ('0' + str(date_day_1))[-2:]
                    date_day_2 = date_day_1

                date_re = re.search(r"([0-9]{1,2})-([0-9]{1,2})", str(date_day_value), flags=re.I)
                if date_re:
                    date_day_1 = date_re.group(1)
                    date_day_1 = ('0' + str(date_day_1))[-2:]
                    date_day_2 = date_re.group(2)
                    date_day_2 = ('0' + str(date_day_2))[-2:]
#            print(object_instance.catalog_number + " date_month_value: " + date_month_value)
#            print(object_instance.catalog_number + " date_month_1: " + date_month_1)
#            print(object_instance.catalog_number + " date_month_2: " + date_month_2)
            if date_month_1 != date_month_2:
                if date_day_1 != date_day_2:
                    pass
                else:
                    date_day_1 = ''
                    date_day_2 = ''
            if date_year_1 != date_year_2:
                if date_month_1 != date_month_2:
                    pass
                else:
                    date_month_1 = ''
                    date_month_2 = ''
                    date_day_1 = ''
                    date_day_2 = ''

            date_1 = date_month_1 + '/' + date_day_1 + '/' + date_year_1
            date_1 = re.sub(r"/+", "/", date_1)
            date_1 = re.sub(r"^/*", "", date_1)

            date_2 = date_month_2 + '/' + date_day_2 + '/' + date_year_2
            date_2 = re.sub(r"/+", "/", date_2)
            date_2 = re.sub(r"^/*", "", date_2)

            if date_1 == date_2:
                if date_1 != "":
                    setattr(object_instance, model_field, date_1)
#                    print(object_instance.catalog_number + "got to G")
                    return True
            else:
                setattr(object_instance, model_field, date_1+ "-" + date_2)
#                print(object_instance.catalog_number + "got to H")
                return True
    return False

def automate_glottocode(model_field_glottocode, model_field_name, object_instance):
    global languoids
    # automate = True
    # human_field_indeces = list_string_find_indices(headers, human_field_glottocode)
    # if is_valid_param(human_field_indeces):
    #     model_field_value = row[human_field_indeces[0]]
    #     if is_valid_param(model_field_value):
    #         automate = False
    # if automate:
    search_items = getattr(object_instance, model_field_name).lower().split(',')
    results = []
    for search in search_items:
        search = search.strip()
        languoid_results = [languoid['glottocode'] for languoid in languoids if languoid['name'].lower() == search]
        if len(languoid_results) == 1:
            results.append(languoid_results[0])
    setattr(object_instance, model_field_glottocode, ', '.join(results))


@login_required
@user_passes_test(is_member_of_archivist, login_url='/no-permission', redirect_field_name=None)
def ImportView(request):
    url_path = request.get_full_path()
    if re.search('catalog', url_path, flags=re.I):
        template = 'import.html'
    elif re.search('collaborators', url_path, flags=re.I):
        template = 'import_collaborator.html'
    else:
        ## redirect to the language page (disable import) while I am converting to glottocodes
        return redirect('/languages/')
        template = 'import_language.html'

    prompt_message = 'Order of CSV should be...'

    if request.method == 'GET':
        csv_format_form = Csv_format_type()
        context = {
            'order': prompt_message,
            'csv_format_form' : csv_format_form,
        }
        return render(request, template, context)

    try:
        xls_file = request.FILES['file']
    except Exception as e:
        csv_format_form = Csv_format_type()
        context = {
            'order': prompt_message,
            'csv_format_form' : csv_format_form,
        }
        return render(request, template, context)


    if not ( xls_file.name.endswith('.xlsx') ) or ( xls_file.name.endswith('.xls') ):
        messages.error(request, 'Error: File must be a .xls(x) file. Nothing was uploaded.')
        csv_format_form = Csv_format_type()
        context = {
            'order': prompt_message,
            'csv_format_form' : csv_format_form,
        }
        return render(request, template, context)

    csv_format_form = Csv_format_type(request.POST)

    file_name = default_storage.save('imports/' + str(datetime.datetime.now()).replace(':', '-').replace('.', '-') + '_' + str(request.user.get_username()) + '_' + str(xls_file.name), xls_file)

    workbook = load_workbook(xls_file)
    sheet = workbook.active
    for first_row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(first_row)
    countess = 0
    for row_tuple in sheet.iter_rows(min_row=2, values_only=True):
        row = list(row_tuple)
        countess += 1
        #print(countess)

#    data_set = csv_file.read().decode('UTF-8')
#    io_string = io.StringIO(data_set)
#    reader = csv.reader(csv_file, delimiter=',', quotechar='|')
#    biline = next(reader)
#    headers = next(reader)
#    for row in reader:
##        row.extend([''] * ( len(headers)-len(row) ) ) # this makes sure the current row has as many elements as the header, in case values are empty
#        print(list_string_find_indices(headers,'^Catalog Number'))

        if re.search('catalog', url_path, flags=re.I): # check if you are on the item import page based on URL

            catalog_number_indexes = list_string_find_indices(headers,'^Catalog Number$')
            if not is_valid_param(catalog_number_indexes):
                messages.warning(request, 'Speadsheet must have a row with header "Catalog Number"')
                csv_format_form = Csv_format_type()
                context = {
                    'order': prompt_message,
                    'csv_format_form' : csv_format_form,
                }
                return render(request, template, context)
            else:
                catalog_number_value = row[catalog_number_indexes[0]]

            if not catalog_number_value:
                messages.warning(request, 'Warning: One line item had no catalog number, and was ignored')
            else:

                item, item_created = Item.objects.get_or_create(catalog_number=catalog_number_value, defaults={'modified_by': request.user.get_username()})
#                original_item = item
#                item.modified_by = request.user.get_username()
#                item.save()

                if not item_created:
                    dialect_instances = DialectInstance.objects.filter(item=item)
                    old_dialect_instances = { dialect_instance.language : Dialect.objects.filter(pk__in=list(dialect_instance.name.all().values_list('id', flat=True))) for dialect_instance in dialect_instances }

                import_language_field_success, return_object = import_language_field(request, headers, row, item, model = 'Item')

## document metadata import on item import has been disabled for now

                # adding documents

                document_prefix = ''

                document_headers = re.findall('Document ([0-9]{1,})', ', '.join([x for x in headers if x is not None])) #filter any NoneTypes in column header list, as input to finall
                unique_document_headers = list(set(document_headers))
                number_of_unique_documents = len(unique_document_headers)

                if number_of_unique_documents > 0:
                    document_prefix = '^Document '
                else:
                    document_headers = re.findall('Track ([0-9]{1,})', ', '.join([x for x in headers if x is not None])) #filter any NoneTypes in column header list, as input to finall
                    unique_document_headers = list(set(document_headers))
                    number_of_unique_documents = len(unique_document_headers)

                    if number_of_unique_documents > 0:
                        document_prefix = '^Track '

                # document_filename_indeces = list_string_find_indices(headers, '^Document ([0-9])* Filename$')
                # document_enumerator_indeces = list_string_find_indices(headers, '^Document ([0-9])* Enumerator$')
                # document_title_indeces = list_string_find_indices(headers, '^Document ([0-9])* Title$')
                # document_access_level_indeces = list_string_find_indices(headers, '^Document ([0-9])* Access Level$')
                # document_av_spec_indeces = list_string_find_indices(headers, '^Document ([0-9])* A/V Specs$')
                # document_language_indeces = list_string_find_indices(headers, '^Document ([0-9])* Language Name(\s)?([0-9])*$')
                # document_dialect_indeces = list_string_find_indices(headers, '^Document ([0-9])* Dialect(\s)?([0-9])*$')
                # document_collaborator_indeces = list_string_find_indices(headers, '^Document ([0-9])* Collaborator(\s)?([0-9])*$')
                # document_collaborator_role_indeces = list_string_find_indices(headers, '^Document ([0-9])* Collaborator Role(\s)?([0-9])*$')
                # document_creation_date_indeces = list_string_find_indices(headers, '^Document ([0-9])* Creation Date$')
                #
                # if ( document_filename_indeces or document_enumerator_indeces or document_title_indeces or document_access_level_indeces or document_av_spec_indeces or document_language_indeces or document_dialect_indeces or document_collaborator_indeces or document_collaborator_role_indeces or document_creation_date_indeces):
                #     document_prefix = '^Document '
                # else:
                #     document_filename_indeces = list_string_find_indices(headers, '^Track ([0-9])* Filename$')
                #     document_enumerator_indeces = list_string_find_indices(headers, '^Track ([0-9])* Enumerator$')
                #     document_title_indeces = list_string_find_indices(headers, '^Track ([0-9])* Title$')
                #     document_access_level_indeces = list_string_find_indices(headers, '^Track ([0-9])* Access Level$')
                #     document_av_spec_indeces = list_string_find_indices(headers, '^Track ([0-9])* A/V Specs$')
                #     document_language_indeces = list_string_find_indices(headers, '^Track ([0-9])* Language Name(\s)?([0-9])*$')
                #     document_dialect_indeces = list_string_find_indices(headers, '^Track ([0-9])* Dialect(\s)?([0-9])*$')
                #     document_collaborator_indeces = list_string_find_indices(headers, '^Track ([0-9])* Collaborator(\s)?([0-9])*$')
                #     document_collaborator_role_indeces = list_string_find_indices(headers, '^Track ([0-9])* Collaborator Role(\s)?([0-9])*$')
                #     document_creation_date_indeces = list_string_find_indices(headers, '^Track ([0-9])* Creation Date$')
                #
                #     if ( document_filename_indeces or document_enumerator_indeces or document_title_indeces or document_access_level_indeces or document_av_spec_indeces or document_language_indeces or document_dialect_indeces or document_collaborator_indeces or document_collaborator_role_indeces or document_creation_date_indeces):
                #         document_prefix = '^Track '

                if is_valid_param(document_prefix):

                    # indeces_list = [ document_filename_indeces, document_enumerator_indeces, document_title_indeces, document_access_level_indeces, document_av_spec_indeces, document_language_indeces, document_dialect_indeces, document_collaborator_indeces, document_collaborator_role_indeces, document_creation_date_indeces ]
                    # list_len = [len(i) for i in indeces_list]
                    # max_list_len = max(list_len)

#                    document_filename_indeces = list_string_find_indices(headers, '^Document ([0-9])+ Filename$')
#                    document_filename_indeces.extend( [None] * ( number_of_unique_documents - len(document_filename_indeces) ) )

                    old_dialect_instances_for_item_documents = {}
                    all_document_dialect_instances = DialectInstance.objects.none()
                    for unique_document_header in unique_document_headers:
                        document_prefix_number = document_prefix + str(unique_document_header)
                        document_filename_indexes = list_string_find_indices(headers, document_prefix_number + ' Filename$')
                        if not is_valid_param(document_filename_indexes): #this test of document failure is necesary now because this loop depends on a Filename column in the spreadsheet. Other failure tests are at the end of this for loop.
                            document_import_success = False
                            messages.warning(request, item.catalog_number + ' was not added/updated (all changes were reverted): Document ' + str(unique_document_header) + ' had no column with header "Filename."')
                            break
                        document_filename_value = row[document_filename_indexes[0]]

                        if is_valid_param(document_filename_value):
                            document, document_created = Document.objects.get_or_create(filename=document_filename_value, defaults={'item': item, 'modified_by': request.user.get_username()})
                        else:
                            #document = Document.objects.create(item=item, modified_by=request.user.get_username())
                            #document_created = True
                            document_import_success = True
                            continue # this breaks to next document without doing anything when filename is just blank



                        if not document_created:
                            document_dialect_instances = DialectInstance.objects.filter(document=document)
                            all_document_dialect_instances |= document_dialect_instances
                            document_old_dialect_instances = { document_dialect_instance.language : Dialect.objects.filter(pk__in=list(document_dialect_instance.name.all().values_list('id', flat=True))) for document_dialect_instance in document_dialect_instances }
                            old_dialect_instances_for_item_documents[document] = document_old_dialect_instances


                        import_field(request, 'title', (document_prefix_number + ' Title$',), headers, row, document, model = 'Item document')
                        document_import_language_field_success, return_object = import_language_field(request, headers, row, document, model = 'Item document', document_prefix=document_prefix_number)


                        # for documents: collaborator and collaborator role, a manytomany one, with a double foreign key one (that is a multiselect field)
                        document_collaborator_indexes = list_string_find_indices(headers, document_prefix_number + ' Collaborator(\s)?([0-9])*$')
                        if is_valid_param(document_collaborator_indexes): # don't do anything if import file doesn't have Collaborator Name fields
                            document_collaborator_role_indexes = list_string_find_indices(headers, document_prefix_number + ' Collaborator Roles*(\s)?([0-9])*$')
                            if is_valid_param(document_collaborator_role_indexes):
                                document_collaborator_role_indexes.extend( [None] * ( len(document_collaborator_indexes) - len(document_collaborator_role_indexes) ) ) # if import file has more Collaborator Name fields than Collaborator Role fields, make dummy Collaborator Role fields
                            else:
                                document_collaborator_role_indexes = [None] * len(document_collaborator_indexes) # if import file has Collaborator fields but no Collaborator Role fields, make dummy Collaborator Role fields

                            document.collaborator.clear()
                            document_indexes = zip(document_collaborator_indexes, document_collaborator_role_indexes)
                            for document_collaborator_index, document_collaborator_role_index in document_indexes:
                                if is_valid_param(document_collaborator_index):
                                    document_collaborator_value = row[document_collaborator_index]
                                    if is_valid_param(document_collaborator_value):
                                        collaborator_next_id = 1
                                        try:
                                            collaborator_next_id = Collaborator.objects.all().aggregate(Max('collaborator_id'))['collaborator_id__max'] + 1
                                        except:
                                            pass
                                        try:
                                            collaborator =  Collaborator.objects.get(name=document_collaborator_value)
                                            created = False
                                        except:
                                            try:
                                                collaborator =  Collaborator.objects.get(other_names=document_collaborator_value)
                                                created = False
                                            except:
                                                collaborator, created = Collaborator.objects.get_or_create(name=document_collaborator_value,
                                                    defaults={'collaborator_id': collaborator_next_id},)
                                        # need to build warning here if there is a duplicate
                                        document.collaborator.add(collaborator)
                                        if created:
                                            messages.info(request, 'A new collaborator (' + collaborator.name + ') was created and added to Collaborators for Document ' + document.filename)
                                        collaborator_role, created = CollaboratorRole.objects.get_or_create(document=document, collaborator=collaborator) #create collaborator role object based on collaborator, regardless of value
                                        if is_valid_param(document_collaborator_role_index):
                                            document_collaborator_role_value = row[document_collaborator_role_index]
                                            if is_valid_param(document_collaborator_role_value):
                                                document_cleaned_collaborator_role_value = reverse_lookup_choices(ROLE_CHOICES, document_collaborator_role_value)
                                                document_cleaned_collaborator_role_value = re.sub(r"\n", r", ", document_cleaned_collaborator_role_value)
                                                document_cleaned_collaborator_role_value = re.sub(r"[,\s]*$", r"", document_cleaned_collaborator_role_value)
                                                collaborator_role.role = document_cleaned_collaborator_role_value
                                                collaborator_role.modified_by = request.user.get_username()
                                                collaborator_role.save()


                        document_geographic_success = import_child_field(request, 'document', ('lat', 'long'), (document_prefix_number + ' Latitude$', document_prefix_number + ' Longitude$'), headers, row, document, model = 'Item document', num=True)

                        document_access_level_success = import_field(request, 'access_level', (document_prefix_number + ' Access Level$',), headers, row, document, model = 'Item document', choices=ACCESS_CHOICES)
                        document_enumerator_success = import_field(request, 'enumerator', (document_prefix_number + ' Enumerator$',), headers, row, document, model = 'Item document')
                        import_field(request, 'av_spec', (document_prefix_number + ' A/V Specs$',), headers, row, document, model = 'Item document')
                        document_date_created_1part_success = import_date_field('creation_date', (document_prefix_number + ' Date Created$',), headers, row, document)
                        document_date_created_3part_success = import_date_field('creation_date', (document_prefix_number + ' Date Created Year$', document_prefix_number + ' Date Created Month$', document_prefix_number + ' Date Created Day$'), headers, row, document)
                        document_created_date_1part_success = import_date_field('creation_date', (document_prefix_number + ' Created Date$',), headers, row, document)
                        document_created_date_3part_success = import_date_field('creation_date', (document_prefix_number + ' Created Date, Year$', document_prefix_number + ' Created Date, Month$', document_prefix_number + ' Created Date, Day$'), headers, row, document)
                        document_creation_date_1part_success = import_date_field('creation_date', (document_prefix_number + ' Creation Date$',), headers, row, document)
                        document_creation_date_3part_success = import_date_field('creation_date', (document_prefix_number + ' Creation Date, Year$', document_prefix_number + ' Creation Date, Month$', document_prefix_number + ' Creation Date, Day$'), headers, row, document)
                        document_creation_date_success = document_date_created_1part_success and document_date_created_3part_success and document_created_date_1part_success and document_created_date_3part_success and document_creation_date_1part_success and document_creation_date_3part_success
                        if not document_creation_date_success:
                            messages.warning(request, "Creation Date for Document " + str(unique_document_header) + " was not set, because it was an invalid format")


                        document_import_success = ( document_import_language_field_success and
                                                    document_geographic_success and
                                                    document_access_level_success and
                                                    document_enumerator_success and
                                                    document_creation_date_success )

                        if not document_import_success:
                            if document_created:
                                document.delete()
                            messages.warning(request, item.catalog_number + ' was not added/updated (all changes were reverted): Document ' + str(unique_document_header) + ' had one or more invalid values')
                            break

                        document.modified_by = request.user.get_username()
                        document.save()

                    if not document_import_success:
                        all_document_dialect_instances.delete()
                        for old_dialect_instance_for_item_documents in old_dialect_instances_for_item_documents:
                            old_dialect_instance_document_name = old_dialect_instance_for_item_documents
                            document_old_dialect_instances = old_dialect_instances_for_item_documents[old_dialect_instance_for_item_documents]
                            for document_old_dialect_instance in document_old_dialect_instances:
                                dialect_instance, created = DialectInstance.objects.get_or_create(document=old_dialect_instance_document_name, language=document_old_dialect_instance, defaults={'modified_by': request.user.get_username()})
                                dialect_instance.name.set(document_old_dialect_instances[document_old_dialect_instance], clear=True)


                else:
                    document_import_success = True

#
#
#
#                 document_filename_indeces.extend( [None] * ( max_list_len - len(document_filename_indeces) ) )
#                 document_enumerator_indeces.extend( [None] * ( max_list_len - len(document_enumerator_indeces) ) )
#                 document_title_indeces.extend( [None] * ( max_list_len - len(document_title_indeces) ) )
#                 document_access_level_indeces.extend( [None] * ( max_list_len - len(document_access_level_indeces) ) )
#                 document_av_spec_indeces.extend( [None] * ( max_list_len - len(document_av_spec_indeces) ) )
#                 document_language_indeces.extend( [None] * ( max_list_len - len(document_language_indeces) ) )
#                 document_dialect_indeces.extend( [None] * ( max_list_len - len(document_dialect_indeces) ) )
#                 document_collaborator_indeces.extend( [None] * ( max_list_len - len(document_collaborator_indeces) ) )
#                 document_collaborator_role_indeces.extend( [None] * ( max_list_len - len(document_collaborator_role_indeces) ) )
#                 document_creation_date_indeces.extend( [None] * ( max_list_len - len(document_creation_date_indeces) ) )
#
#                 for document_number in range(max_list_len):
#                     document_title_value, document_collaborator_value, document_language_value, document_av_spec_value, document_creation_date_value = None, None, None, None, None
#                     document_title_index = document_title_indeces[document_number]
#                     if document_title_index:
#                         document_title_value = row[document_title_index]
#                     document_collaborator_index = document_collaborator_indeces[document_number]
#                     if document_collaborator_index:
#                         document_collaborator_value = row[document_collaborator_index]
#                     document_language_index = document_language_indeces[document_number]
#                     if document_language_index:
#                         document_language_value = row[document_language_index]
#                     document_av_spec_index = document_av_spec_indeces[document_number]
#                     if document_av_spec_index:
#                         document_av_spec_value = row[document_av_spec_index]
#                     document_creation_date_index = document_creation_date_indeces[document_number]
#                     if document_creation_date_index:
#                         document_creation_date_value = row[document_creation_date_index]
#
#                     if document_collaborator_value:
#                         collaborator_next_id = Collaborator.objects.all().aggregate(Max('collaborator_id'))['collaborator_id__max'] + 1
#                         get_collaborator_by_value, created = Collaborator.objects.get_or_create(name=document_collaborator_value,
#                             defaults={'collaborator_id': collaborator_next_id},)
#                     if document_language_value:
# ##################### needs to get converted to special language treatment
#                         get_language_by_value, created = Language.objects.get_or_create(name=document_language_value)
#                     if is_valid_param(document_title_value):
#                         document, created = Document.objects.get_or_create(title=document_title_value, item=item,
#                             defaults={'filename': 'no filename'},)
#                     else:
#                         document_qs = Document.objects.exclude(title=document_title_value, item=item)
#                         if document_collaborator_value:
#                             document_qs = document_qs.filter(collaborator=get_collaborator_by_value)
#                         if document_language_value:
#                             document_qs = document_qs.filter(language=get_language_by_value)
#                         if document_av_spec_value:
#                             document_qs = document_qs.filter(av_spec=document_av_spec_value)
#                         if document_creation_date_value:
#                             document_qs = document_qs.filter(creation_date=document_creation_date_value)
#                         document = document_qs.first()
#                         if not document:
#                             document = Document.objects.create(item=item)
#                         if not document.filename:
#                             document.filename = 'no filename'
#                     if document_collaborator_value:
#                         document.collaborator.add(get_collaborator_by_value)
#                     if document_language_value:
# ##################### needs to get converted to special language treatment
#                         document.language.add(get_language_by_value)
#                     if document_av_spec_value:
#                         document.av_spec = document_av_spec_value
#                     if document_creation_date_value:
#                         try:
#                             document_creation_date_value = validate_date_text(document_creation_date_value)
#                             document.creation_date = document_creation_date_value
#                         except:
#                             messages.warning(request, "Creation Date for Document:" + document.title + "(" + item.catalog_number + ") was not set, because it was an invalid format")
#                     document.save()


                # collaborator and collaborator role, a manytomany one, with a double foreign key one (that is a multiselect field)

                if not item_created:
                    collaborator_roles = CollaboratorRole.objects.filter(item=item)
                    old_collaborator_roles = { collaborator_role.collaborator : collaborator_role.role for collaborator_role in collaborator_roles }



                collaborator_indexes = list_string_find_indices(headers,'^Collaborator Name(\s)?([0-9])*$')
                if is_valid_param(collaborator_indexes): # don't do anything if import file doesn't have Collaborator Name fields
                    collaborator_role_indexes = list_string_find_indices(headers,'^Collaborator Roles*(\s)?([0-9])*$')
                    if is_valid_param(collaborator_role_indexes):
                        collaborator_role_indexes.extend( [None] * ( len(collaborator_indexes) - len(collaborator_role_indexes) ) ) # if import file has more Collaborator Name fields than Collaborator Role fields, make dummy Collaborator Role fields
                    else:
                        collaborator_role_indexes = [None] * len(collaborator_indexes) # if import file has Collaborator fields but no Collaborator Role fields, make dummy Collaborator Role fields
                    # print("Before clear: " + str(item.collaborator.all()))
                    item.collaborator.clear()
                    CollaboratorRole.objects.filter(item=item).delete()
                    # print("After clear: " + str(item.collaborator.all()))
                    indexes = zip(collaborator_indexes, collaborator_role_indexes)
                    for collaborator_index, collaborator_role_index in indexes:
                        if is_valid_param(collaborator_index):
                            collaborator_value = row[collaborator_index]
                            if is_valid_param(collaborator_value):
                                collaborator_next_id = 1
                                try:
                                    collaborator_next_id = Collaborator.objects.all().aggregate(Max('collaborator_id'))['collaborator_id__max'] + 1
                                except:
                                    pass
                                try:
                                    collaborator =  Collaborator.objects.get(name=collaborator_value)
                                    created = False
                                except:
                                    try:
                                        collaborator =  Collaborator.objects.get(other_names=collaborator_value)
                                        created = False
                                    except:
                                        collaborator, created = Collaborator.objects.get_or_create(name=collaborator_value,
                                            defaults={'collaborator_id': collaborator_next_id},)
                                # need to build warning here if there is a duplicate
                                # print("Before add: " + str(item.collaborator.all()))

                                item.collaborator.add(collaborator)
                                print("After add: " + str(item.collaborator.all()))

                                if created:
                                    print("Collaborator was created")
                                    messages.info(request, 'A new collaborator (' + collaborator.name + ') was created and added to Collaborators for ' + item.catalog_number)
                                print("collaborator roles" + str(CollaboratorRole.objects.filter(item=item, collaborator=collaborator))) #create collaborator role object based on collaborator, regardless of value
                                print("collaborator" + str(collaborator)) #create collaborator role object based on collaborator, regardless of value

                                collaborator_role, created = CollaboratorRole.objects.get_or_create(item=item, collaborator=collaborator) #create collaborator role object based on collaborator, regardless of value
                                if is_valid_param(collaborator_role_index):
                                    collaborator_role_value = row[collaborator_role_index]
                                    if is_valid_param(collaborator_role_value):
                                        cleaned_collaborator_role_value = reverse_lookup_choices(ROLE_CHOICES, collaborator_role_value)
                                        cleaned_collaborator_role_value = re.sub(r"\n", r", ", cleaned_collaborator_role_value)
                                        cleaned_collaborator_role_value = re.sub(r"[,\s]*$", r"", cleaned_collaborator_role_value)
                                        collaborator_role.role = cleaned_collaborator_role_value
                                        collaborator_role.modified_by = request.user.get_username()
                                        collaborator_role.save()

                geographic_success = import_child_field(request, 'item', ('lat', 'long'), ('^Latitude$', '^Longitude$'), headers, row, item, model = 'Item', num=True)


                ## the rest, in alphabetical order
                accession_date_1part_success = import_date_field('accession_date', ('^Accession Date$',), headers, row, item)
                if not accession_date_1part_success:
                    messages.warning(request, "Accession Date for " + item.catalog_number + " was not set, because it was an invalid format")
                import_field(request, 'access_level_restrictions', ('^Access Level Restrictions$',), headers, row, item, model = 'Item')
                import_field(request, 'accession_number', ('^Accession Number$',), headers, row, item, model = 'Item')
                import_field(request, 'acquisition_notes', ('^Acquisition Notes$',), headers, row, item, model = 'Item')
                import_field(request, 'additional_digital_file_location', ('^Additional Digital File Location$',), headers, row, item, model = 'Item')
                import_field(request, 'associated_ephemera', ('^Associated Ephemera$',), headers, row, item, model = 'Item')
                availability_status_success = import_field(request, 'availability_status', ('^Availability Status$',), headers, row, item, model = 'Item', choices = AVAILABILITY_CHOICES)
                import_field(request, 'availability_status_notes', ('^Availability Status Notes$',), headers, row, item, model = 'Item')
                import_field(request, 'call_number', ('^Call Number$',), headers, row, item, model = 'Item')
                import_field(request, 'cataloged_by', ('^Cataloged By$',), headers, row, item, model = 'Item')
                cataloged_date_1part_success = import_date_field('cataloged_date', ('^Cataloged Date$',), headers, row, item)
                cataloged_3part_success = import_date_field('cataloged_date', ('^Year Cataloged$','^Month Cataloged$','^Day Cataloged$'), headers, row, item)
                cataloged_date_3part_success = import_date_field('cataloged_date', ('^Cataloged Date, Year$','^Cataloged Date, Month$','^Cataloged Date, Day$'), headers, row, item)
                cataloged_date_success = cataloged_date_1part_success and cataloged_3part_success and cataloged_date_3part_success
                if not cataloged_date_success:
                    messages.warning(request, "Cataloged Date for " + item.catalog_number + " was not set, because it was an invalid format")
                import_field(request, 'collecting_notes', ('^Collecting Notes$',), headers, row, item, model = 'Item')
                import_field(request, 'collection_name', ('^Collection Name$',), headers, row, item, model = 'Item')
                import_field(request, 'collector_info', ('^Collector Information$',), headers, row, item, model = 'Item')
                import_field(request, 'collector_name', ('^Collector Name$',), headers, row, item, model = 'Item')
                import_field(request, 'collectors_number', ('^Collector.?s Number$',), headers, row, item, model = 'Item')
                condition_success = import_field(request, 'condition', ('^Condition$',), headers, row, item, model = 'Item', choices=CONDITION_CHOICES)
                import_field(request, 'condition_notes', ('^Condition Notes$',), headers, row, item, model = 'Item')
                import_field(request, 'conservation_recommendation', ('^Conservation Recommendation$',), headers, row, item, model = 'Item')
                import_field(request, 'conservation_treatments_performed', ('^Conservation Treatments Performed$',), headers, row, item, model = 'Item')
                import_field(request, 'copyrighted_notes', ('^Copyrighted Notes$',), headers, row, item, model = 'Item')
                import_field(request, 'country_or_territory', ('^Country or Territory$',), headers, row, item, model = 'Item')
                import_field(request, 'county_or_parish', ('^County or Parish$',), headers, row, item, model = 'Item')
                collection_date_1part_success = import_date_field('collection_date', ('^Collection Date$',), headers, row, item)
                collection_3part_success = import_date_field('collection_date', ('^Collection Year$','^Collection Month$','^Collection Day$'), headers, row, item)
                collection_date_3part_success = import_date_field('collection_date', ('^Collection Date, Year$','^Collection Date, Month$','^Collection Date, Day$'), headers, row, item)
                collection_date_success = collection_date_1part_success and collection_3part_success and collection_date_3part_success
                if not collection_date_success:
                    messages.warning(request, "Collection Date for " + item.catalog_number + " was not set, because it was an invalid format")
                date_created_1part_success = import_date_field('creation_date', ('^Date Created$',), headers, row, item)
                date_created_3part_success = import_date_field('creation_date', ('^Date Created Year$','^Date Created Month$','^Date Created Day$'), headers, row, item)
                created_date_1part_success = import_date_field('creation_date', ('^Created Date$',), headers, row, item)
                created_date_3part_success = import_date_field('creation_date', ('^Created Date, Year$','^Created Date, Month$','^Created Date, Day$'), headers, row, item)
                creation_date_1part_success = import_date_field('creation_date', ('^Creation Date$',), headers, row, item)
                creation_date_3part_success = import_date_field('creation_date', ('^Creation Date, Year$','^Creation Date, Month$','^Creation Date, Day$'), headers, row, item)
                creation_date_success = date_created_1part_success and date_created_3part_success and created_date_1part_success and created_date_3part_success and creation_date_1part_success and creation_date_3part_success
                if not creation_date_success:
                    messages.warning(request, "Creation Date for " + item.catalog_number + " was not set, because it was an invalid format")
                deposit_date_1part_success = import_date_field('deposit_date', ('^Deposit Date$',), headers, row, item)
                if not deposit_date_1part_success:
                    messages.warning(request, "Deposit Date for " + item.catalog_number + " was not set, because it was an invalid format")
                import_field(request, 'depositor_contact_information', ('^Depositor Contact Information$',), headers, row, item, model = 'Item')
                import_field(request, 'depositor_name', ('^Depositor Name$',), headers, row, item, model = 'Item')
                import_field(request, 'description_scope_and_content', ('^Description Scope and Content$',), headers, row, item, model = 'Item')
                import_field(request, 'digital_file_location', ('^Digital File Location$',), headers, row, item, model = 'Item')
                import_field(request, 'english_title', ('^English Title$',), headers, row, item, model = 'Item')
                import_field(request, 'equipment_used', ('^Equipment Used$',), headers, row, item, model = 'Item')
                import_field(request, 'filemaker_legacy_pk_id', ('PK_ID',), headers, row, item, model = 'Item')
                import_field(request, 'filemaker_legacy_pk_id', ('Filemaker Legacy PK ID',), headers, row, item, model = 'Item')
                general_content_success = import_field(request, 'general_content', ('^General Content$',), headers, row, item, model = 'Item', choices=CONTENT_CHOICES)
                genre_success = import_field(request, 'genre', ('^Genre$',), headers, row, item, model = 'Item', choices=STRICT_GENRE_CHOICES, multiselect=True)
                import_field(request, 'global_region', ('^Global Region$',), headers, row, item, model = 'Item')
                import_field(request, 'indigenous_title', ('^Indigenous Title$',), headers, row, item, model = 'Item')
                import_field(request, 'ipm_issues', ('^IPM Issues$',), headers, row, item, model = 'Item')
                import_field(request, 'isbn', ('^ISBN$',), headers, row, item, model = 'Item')
                access_level_success = import_field(request, 'item_access_level', ('^Item Access Level$',), headers, row, item, model = 'Item', choices=ACCESS_CHOICES)
                import_field(request, 'lender_loan_number', ('^Lender Loan Number$',), headers, row, item, model = 'Item')
                import_field(request, 'loc_catalog_number', ('^LCCN$',), headers, row, item, model = 'Item')
                import_field(request, 'loc_catalog_number', ('^LOC Catalog Number$',), headers, row, item, model = 'Item')
                import_field(request, 'location_of_original', ('^Location of Original$',), headers, row, item, model = 'Item')
                import_field(request, 'migration_file_format', ('^Migration File Format$',), headers, row, item, model = 'Item')
                import_field(request, 'migration_location', ('^Migration Location$',), headers, row, item, model = 'Item')
                import_field(request, 'municipality_or_township', ('^Municipality or Township$',), headers, row, item, model = 'Item')
                original_format_medium_success_physical = import_field(request, 'original_format_medium', ('^Original Physical Format$',), headers, row, item, model = 'Item', choices=FORMAT_CHOICES)
                original_format_medium_success_medium = import_field(request, 'original_format_medium', ('^Original Format Medium$',), headers, row, item, model = 'Item', choices=FORMAT_CHOICES)
                original_format_medium_success = original_format_medium_success_physical and original_format_medium_success_medium
                import_field(request, 'other_information', ('^Other Information$',), headers, row, item, model = 'Item')
                import_field(request, 'other_institutional_number', ('^Other Institutional Number$',), headers, row, item, model = 'Item')
                permission_to_publish_online_success = import_field(request, 'permission_to_publish_online', ('^Permission to Publish Online$',), headers, row, item, model = 'Item', yesno=True)
                import_field(request, 'project_grant', ('^Project/Grant$',), headers, row, item, model = 'Item')
                import_field(request, 'public_event', ('^Public Event$',), headers, row, item, model = 'Item')
                import_field(request, 'publisher', ('^Publisher$',), headers, row, item, model = 'Item')
                import_field(request, 'publisher', ('^Publisher Name$',), headers, row, item, model = 'Item')
                import_field(request, 'publisher_address', ('^Publisher Address$',), headers, row, item, model = 'Item')
                import_field(request, 'recorded_on', ('^Recorded On$',), headers, row, item, model = 'Item')
                import_field(request, 'recording_context', ('^Recording Context$',), headers, row, item, model = 'Item')
                import_field(request, 'software_used', ('^Software Used$',), headers, row, item, model = 'Item')
                import_field(request, 'state_or_province', ('^State or Province$',), headers, row, item, model = 'Item')
                import_field(request, 'temporary_accession_number', ('^Temporary Accession Number$',), headers, row, item, model = 'Item')
                import_field(request, 'total_number_of_pages_and_physical_description', ('^Total Number of Pages and Physical Description$',), headers, row, item, model = 'Item')
                type_of_accession_success = import_field(request, 'type_of_accession', ('^Type of Accession$',), headers, row, item, model = 'Item', choices=ACCESSION_CHOICES)
                import_field(request, 'educational_materials', ('^Educational Materials$',), headers, row, item, model = 'Item')
                language_description_type_success = import_field(request, 'language_description_type', ('^Language Description Type$',), headers, row, item, model = 'Item', choices=LANGUAGE_DESCRIPTION_CHOICES, multiselect=True)

                import_success = ( import_language_field_success and
                                   document_import_success and
                                   geographic_success and
                                   accession_date_1part_success and
                                   availability_status_success and
                                   cataloged_date_success and
                                   condition_success and
                                   collection_date_success and
                                   creation_date_success and
                                   deposit_date_1part_success and
                                   general_content_success and
                                   genre_success and
                                   access_level_success and
                                   original_format_medium_success and
                                   permission_to_publish_online_success and
                                   type_of_accession_success and
                                   language_description_type_success )

                if not import_success:
                    if item_created:
                        item.delete()
                    else:
#                        item = original_item
#                        item.save()
                        dialect_instances.delete()
                        for old_dialect_instance in old_dialect_instances:
                            dialect_instance, created = DialectInstance.objects.get_or_create(item=item, language=old_dialect_instance, defaults={'modified_by': request.user.get_username()})
                            dialect_instance.name.set(old_dialect_instances[old_dialect_instance], clear=True)

                        collaborator_roles.delete()
                        for old_collaborator_role in old_collaborator_roles:
                            collaborator_role, created = CollaboratorRole.objects.get_or_create(item=item, collaborator=old_collaborator_role, defaults={'modified_by': request.user.get_username()})
                            collaborator_role.role = old_collaborator_roles[old_collaborator_role]

                    continue

                document_prefix_with_space = str(item.catalog_number.lower()) + ' '
                documents_with_space = Document.objects.exclude(item=item).filter(filename__icontains = document_prefix_with_space)
                document_prefix_with_underscore = str(item.catalog_number.lower()) + '_'
                documents_with_underscore = Document.objects.exclude(item=item).filter(filename__icontains = document_prefix_with_underscore)
                matching_documents = documents_with_space.union(documents_with_underscore)
                if matching_documents:
                    for matching_document in matching_documents:
                        matching_document.item = item
                        matching_document.save()

                item.modified_by = request.user.get_username()
                item.save()
                if item_created:
                    messages.success(request, item.catalog_number + ' was added')
                else:
                    messages.success(request, item.catalog_number + ' was updated')

        elif re.search('collaborators', url_path, flags=re.I): # check if you are on the collaborator import page, based on URL

            collaborator_name_indexes = list_string_find_indices(headers,'^Collaborator Name$')
            if not is_valid_param(collaborator_name_indexes):
                messages.warning(request, 'Speadsheet must have a row with header "Collaborator Name"')
                csv_format_form = Csv_format_type()
                context = {
                    'order': prompt_message,
                    'csv_format_form' : csv_format_form,
                }
                return render(request, template, context)
            else:
                collaborator_name_value = row[collaborator_name_indexes[0]]

            collaborator_name_value = row[list_string_find_indices(headers,'^Collaborator Name$')[0]]
            if not collaborator_name_value:
                messages.warning(request, 'Warning: One line item had no collaborator name, and was ignored')
            else:

                if is_valid_param(collaborator_name_value):
                    collaborator_next_id = 1
                    try:
                        collaborator_next_id = Collaborator.objects.all().aggregate(Max('collaborator_id'))['collaborator_id__max'] + 1
                    except:
                        pass
                    collaborator, collaborator_created = Collaborator.objects.get_or_create(name=collaborator_name_value,
                        defaults={'collaborator_id': collaborator_next_id, 'modified_by': request.user.get_username()},)
                    # need to build warning here if there is a duplicate

#                original_collaborator = collaborator
#                collaborator.modified_by = request.user.get_username()
#                collaborator.save()

                if not collaborator_created:
                    dialect_instances_native = DialectInstance.objects.filter(collaborator_native=collaborator)
                    old_dialect_instances_native = { dialect_instance.language : Dialect.objects.filter(pk__in=list(dialect_instance.name.all().values_list('id', flat=True))) for dialect_instance in dialect_instances_native }
                    dialect_instances_other = DialectInstance.objects.filter(collaborator_other=collaborator)
                    old_dialect_instances_other = { dialect_instance.language : Dialect.objects.filter(pk__in=list(dialect_instance.name.all().values_list('id', flat=True))) for dialect_instance in dialect_instances_other }

                # native language and dialect, a manytomany one, with a double foreign key one
                import_native_language_field_success, return_object = import_language_field(request, headers, row, collaborator, model = 'Collaborator native')

                # other language and dialect, a manytomany one, with a double foreign key one
                import_other_language_field_success, return_object = import_language_field(request, headers, row, collaborator, model = 'Collaborator other')
                # collaborator.clean()
                # print("yeahn made it")
                ## the rest, in alphabetical order
                anonymous_success = import_field(request, 'anonymous', ('^Anonymous',), headers, row, collaborator, model = "Collaborator", yesno=True)
                birthdate_1part_success = import_date_field('birthdate', ('^Date of Birth$',), headers, row, collaborator)
                if not birthdate_1part_success:
                    messages.warning(request, "Date of Birth for " + collaborator.name + " (ID: " + str(collaborator.collaborator_id) + ") was not set, because it was an invalid format")
                import_field(request, 'clan_society', ('^Clan or Society$',), headers, row, collaborator, model = "Collaborator")
                import_field(request, 'collaborator_id', ('^Collaborator ID$',), headers, row, collaborator, model = "Collaborator")
                deathdate_1part_success = import_date_field('deathdate', ('^Date of Death$',), headers, row, collaborator)
                if not deathdate_1part_success:
                    messages.warning(request, "Date of Death for " + collaborator.name + " (ID: " + str(collaborator.collaborator_id) + ") was not set, because it was an invalid format")
                import_field(request, 'gender', ('^Gender$',), headers, row, collaborator, model = "Collaborator")
                import_field(request, 'gender', ('^Sex$',), headers, row, collaborator, model = "Collaborator")
                import_field(request, 'nickname', ('^Collaborator Nickname$',), headers, row, collaborator, model = "Collaborator")
                import_field(request, 'origin', ('^Collaborator Place of Origin$',), headers, row, collaborator, model = "Collaborator")
                import_field(request, 'other_info', ('^Other Collaborator Information$',), headers, row, collaborator, model = "Collaborator")
                import_field(request, 'other_info', ('^Other info$',), headers, row, collaborator, model = "Collaborator")
                import_field(request, 'other_names', ('^Other Names$',), headers, row, collaborator, model = "Collaborator")
                import_field(request, 'tribal_affiliations', ('^Tribal Affiliation(s)$',), headers, row, collaborator, model = "Collaborator")

                import_success = ( import_native_language_field_success and
                                   import_other_language_field_success and
                                   anonymous_success and
                                   birthdate_1part_success and
                                   deathdate_1part_success )

                if not import_success:
                    if collaborator_created:
                        collaborator.delete()
                    else:
                        dialect_instances_native.delete()
                        dialect_instances_other.delete()

                        for old_dialect_instance_native in old_dialect_instances_native:
                            dialect_instance, created = DialectInstance.objects.get_or_create(collaborator_native=collaborator, language=old_dialect_instance_native, defaults={'modified_by': request.user.get_username()})
                            dialect_instance.name.set(old_dialect_instances_native[old_dialect_instance_native], clear=True)
                        for old_dialect_instance_other in old_dialect_instances_other:
                            dialect_instance, created = DialectInstance.objects.get_or_create(collaborator_other=collaborator, language=old_dialect_instance_other, defaults={'modified_by': request.user.get_username()})
                            dialect_instance.name.set(old_dialect_instances_other[old_dialect_instance_other], clear=True)
                    continue

                collaborator.modified_by = request.user.get_username()
                collaborator.save()
                if collaborator_created:
                    messages.success(request, collaborator.name + ' (ID: ' + str(collaborator.collaborator_id) + ') was added')
                else:
                    messages.success(request, collaborator.name + ' (ID: ' + str(collaborator.collaborator_id) + ') was updated')

        else: # check if you are on the language import page, based on URL

            ## redirect to the language page (disable import) while I am converting to glottocodes
            return redirect('languages/')


            # Get the path to the CSV file
            csv_path = os.path.join(settings.STATIC_ROOT, 'codelist.csv')
            # Open the CSV file and read its contents
            global languoids
            languoids = []
            with open(csv_path, 'r') as file:
                reader = csv.DictReader(file)
                for languiod in reader:
                    languoids.append(dict(languiod))            



            import_language_field_success, return_object = import_language_field(request, headers, row, None, model = 'Language')
            if not import_language_field_success:
                continue
            if not return_object:
                continue

            glottocode_success = import_field(request, 'glottocode', ('^Glottocode$',), headers, row, return_object, model = 'Language', validate_glottocode="single")
            import_field(request, 'family', ('^Family$',), headers, row, return_object, model = "Language")
            import_field(request, 'family', ('^Language family$',), headers, row, return_object, model = "Language")
            import_field(request, 'family_abbrev', ('^Family abbreviation$',), headers, row, return_object, model = "Language")
            family_glottocode_success = import_field(request, 'family_id', ('^Family glottocode$',), headers, row, return_object, model = "Language", validate_glottocode="single")
            import_field(request, 'pri_subgroup', ('^Primary subgroup$',), headers, row, return_object, model = "Language")
            import_field(request, 'pri_subgroup_abbrev', ('^Primary subgroup abbreviation$',), headers, row, return_object, model = "Language")
            pri_subgroup_glottocode_success = import_field(request, 'pri_subgroup_id', ('^Primary subgroup glottocode$',), headers, row, return_object, model = "Language", validate_glottocode="single")
            import_field(request, 'sec_subgroup_abbrev', ('^Secondary subgroup abbreviation$',), headers, row, return_object, model = "Language")
            import_field(request, 'sec_subgroup', ('^Secondary subgroup$',), headers, row, return_object, model = "Language")
            sec_subgroup_glottocode_success = import_field(request, 'sec_subgroup_id', ('^Secondary subgroup glottocode$',), headers, row, return_object, model = "Language", validate_glottocode="single")
            import_field(request, 'alt_name', ('^Alternate name\(s\)$',), headers, row, return_object, model = "Language")
            import_field(request, 'alt_name', ('^Alternative name\(s\)$',), headers, row, return_object, model = "Language")
            import_field(request, 'dialects', ('^Dialects$',), headers, row, return_object, model = "Language")
            dialect_ids_success = import_field(request, 'dialects_ids', ('^Dialect glottocodes$',), headers, row, return_object, model = "Language", validate_glottocode="multiple")
            import_field(request, 'region', ('^Region$',), headers, row, return_object, model = "Language")
            latitude_success = import_field(request, 'latitude', ('^Latitude$',), headers, row, return_object, model = "Language", validate_coord=True)
            longitude_success = import_field(request, 'longitude', ('^Longitude$',), headers, row, return_object, model = "Language", validate_coord=True)
            import_field(request, 'tribes', ('^Tribes$',), headers, row, return_object, model = "Language")
            import_field(request, 'notes', ('^Notes$',), headers, row, return_object, model = "Language")

            import_success = ( glottocode_success and
                                family_glottocode_success and
                                pri_subgroup_glottocode_success and
                                sec_subgroup_glottocode_success and
                                dialect_ids_success and
                                latitude_success and
                                longitude_success )

            if not import_success:
                continue

            return_object.level = 'Language'
            return_object.modified_by = request.user.get_username()
            return_object.save()
            messages.success(request, 'Language ' + return_object.name + ' (ISO Indicator: ' + return_object.iso + ')' + ' was updated')

    context = {
        'prompt': prompt_message,
        'csv_format_form' : csv_format_form,
    }

    return render(request, template, context)

class document_upload(UserPassesTestMixin, FormView):
    def test_func(self):
        return self.request.user.groups.filter(name="Archivist").exists()
    def handle_no_permission(self):
        return redirect('/no-permission')
    form_class = UploadDocumentForm
    template_name = 'document_upload.html'  # Replace with your template.
    success_url = '/documents/?form_control_sort=updated'  # Replace with your URL or reverse().
    #success_url = '/documents/import/'

    def post(self, request, *args, **kwargs):
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        files = request.FILES.getlist('file')
        if form.is_valid():
            for this_file in files:
                instance, created = Document.objects.get_or_create(filename=this_file.name)

                if this_file.name.endswith(('.aac', '.AAC')):
                    try:
                        audio_instance = mutagen.File(this_file)
                        instance.duration=float('%.3f'%(audio_instance.info.length))
                    except Exception as e_audio:
                        try:
                            video_instance = Video(file=this_file)
                            video_instance.save()
                            instance.duration=float('%.3f'%(video_instance.duration))
                            video_instance.delete()
                        except Exception as e_video:
                            messages.warning(request, str(this_file.name) + " was not added/updated. (Error message for admin: Mutagen: " + str(e_audio) + ' / ' + str(e_video) + ")")
                            try:
                                instance.delete()
                            except:
                                pass
                            try:
                                video_instance.delete()
                            except:
                                pass

                            continue

                elif this_file.name.endswith(('.mpg', '.MPG', '.mp4', '.MP4', '.m4v', '.M4V', '.mov', '.MOV')):
                    try:
                        video_instance = Video(file=this_file)
                        video_instance.save()
                    except Exception as e:
                        messages.warning(request, str(this_file.name) + " was not added/updated. (Error message for admin: video_encoding: " + str(e) + ")")
                        try:
                            video_instance.delete()
                        except Exception as ee:
                            messages.warning(request, str(this_file.name) + " was not added/updated. (Error message for admin: video_encoding: " + str(ee) + ")")
                        instance.delete()
                        continue
                    instance.duration=float('%.3f'%(video_instance.duration))
                    video_instance.delete()

                elif this_file.name.endswith('.wav'):
                    try:
                        wav_duration = librosa.get_duration(filename=this_file)
                        # with wave.open(this_file,'rb') as f:
                        #     frames = f.getnframes()
                        #     rate = f.getframerate()
                        #     wav_duration = frames / float(rate)
                    except Exception as e:
                        messages.warning(request, str(this_file.name) + " was not added/updated. (Error message for admin: Wave: " + str(e) + ")")
                        instance.delete()
                        continue
                    instance.duration=float('%.3f'%(wav_duration))

                elif this_file.name.endswith(('.aiff','.AIFF','.m4a','.M4A','.mp3','.MP3')):
                    try:
                        audio_instance = mutagen.File(this_file)
                    except Exception as e:
                        messages.warning(request, str(this_file.name) + " was not added/updated. (Error message for admin: Mutagen: " + str(e) + ")")
                        instance.delete()
                        continue
                    instance.duration=float('%.3f'%(audio_instance.info.length))
                else:
                    instance.duration=None

                filename_core, file_extension = os.path.splitext(instance.filename)
                instance.filetype = file_extension.lower()

                instance.filesize = this_file.size


                filename_modified = str(instance.filename.lower()).replace(' ','_')
                filename_prefix = filename_modified.split("_")[0]
                matching_items = Item.objects.filter(catalog_number__iexact=filename_prefix)
                if is_valid_param(matching_items):
                    existing_item = Item.objects.get(catalog_number=matching_items[0])
                    instance.item = existing_item

                instance.modified_by = self.request.user.get_username()
                instance.save()
                messages.success(request, str(this_file.name) + " successfully added/updated.")


            return self.form_valid(form)
        else:
            return self.form_invalid(form)


# Accession Date: simple date, has approximates
# collection date: 3 part
# creation date: 3 part
# modified date: single, don't need approximates, just fill in's
# deposit Date: simple date, has approximates
# cataloged date: 3 part
# collaborator date of birth: simple date, has approximates
# collaborator date of death: simple date, has approximates
